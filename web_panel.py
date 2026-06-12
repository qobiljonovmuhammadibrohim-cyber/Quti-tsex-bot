"""
web_panel.py — To'liq web boshqaruv paneli (aiohttp, Jinja2 yo'q, to'g'ridan HTML).
Sahifalar:
  GET  /web/           → Dashboard
  GET  /web/login      → Login sahifasi
  POST /web/login      → Login
  GET  /web/logout     → Chiqish
  GET  /web/warehouse  → Ombor (qidiruv + filtr + kategoriya)
  POST /web/warehouse/add        → Mahsulot qo'shish
  POST /web/warehouse/kirim      → Kirim
  POST /web/warehouse/chiqim     → Chiqim
  POST /web/warehouse/thresholds → Chegara sozlash
  GET  /web/warehouse/export     → Excel eksport
  GET  /web/warehouse/logs       → Kirim/Chiqim tarixi
  GET  /web/workers              → Ishchilar
  GET  /web/workers/{id}/report  → Ishchi to'liq hisoboti
  POST /web/workers/{id}/toggle  → Aktiv/deaktiv
  POST /web/workers/{id}/delete  → O'chirish
  GET  /web/avans                → Avans tarixi
  POST /web/avans/add            → Avans berish
  GET  /web/penalties            → Jarimalar
  POST /web/penalties/add        → Jarima qo'shish
  GET  /web/reports              → Ish hisobotlari
  GET  /web/ombor-report         → Ombor harakati hisoboti
  GET  /web/maosh-report         → Maosh hisoboti
  GET  /web/salary               → Maosh paneli
  POST /web/salary/confirm/{id}  → Tasdiqlash
  POST /web/salary/confirm-all   → Barchasini tasdiqlash
  GET  /web/prices               → Narxlar
  POST /web/prices/set           → Narx o'rnatish
  GET  /web/reports/download/{t} → Excel yuklab olish
"""

import io
import logging
from datetime import date, datetime, timedelta

from aiohttp import web

from sqlalchemy import select, func, extract, desc, case as sa_case
import sqlalchemy.sql.functions as sf

from config.settings import WEB_HOST, WEB_PORT, WEB_PASSWORD, SECRET_KEY

import re as _re
import csv
import io

def _normalize_razmer(val) -> str | None:
    """
    Razmer normalizatsiyasi:
    "90 x 110" → "90x110", " 90X110 " → "90x110"
    """
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    s = _re.sub(r'\s*[xх×]\s*', 'x', s)
    s = _re.sub(r'\s+', ' ', s).strip()
    return s

from database.db import AsyncSessionLocal
from database.models import (
    User, UserRole, WarehouseProduct, WarehouseLog, ProductCategory, WorkEntry, WorkStatus, WorkType, WorkPrice, SalaryReport, Penalty, PenaltyType, Advance, Attendance, AttendanceType, WorkSession, QualityGrade,
)
from database.queries import (
    get_all_active_users, get_users_by_role,
    get_all_products, get_product_by_id,
    update_product_miqdor, search_products,
    get_all_prices, set_price,
    calculate_and_save_salary, get_monthly_reports,
    create_advance, get_advance_sum,
    get_warehouse_logs_paged,
    get_dashboard_stats, get_user_by_id,
    get_penalty_sum,
)
from utils.reports import generate_daily_excel, generate_weekly_excel, generate_monthly_excel

logger = logging.getLogger(__name__)

# ─── Konstantalar ───────────────────────────────────────────────────────────

CAT_NAMES = {
    "rulon": "🌀 Rulonlar",
    "gofra": "📋 Go'fralar",
    "gofra_zagatovka": "✂️ Zagatovkalar",
    "xromazes": "🖨 Xromazeslar",
    "laminat_xromazes": "✨ Laminat",
    "yarim_tayyor": "⚙️ Yarim tayyor",
    "qolip": "🔲 Qoliplar",
    "tayyor_mahsulot": "📦 Tayyor",
    "adyol_zapchast": "🧩 Adyol zapchast",
    "uskuna_zapchast": "🔧 Uskuna zapchast",
}
ROL_ICONS = {
    "superadmin": "👑", "admin": "⚙️", "omborchi": "🏭",
    "nazoratchi": "🔍", "ishchi": "👷",
}
MONTHS_UZ = [
    "", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr",
]
WORK_TYPE_LABELS = {
    "tiger_kesish": "✂️ Tiger Kesish",
    "gofra_kiley": "🔨 Go'fra Kiley",
    "gofra_ishlab": "🌀 Go'fra Ishlab",
    "list_qogoz": "📄 List Qog'oz",
    "laminatsiya": "✨ Laminatsiya",
    "zagatovka": "📦 Zagatovka",
    "stepler_tikish": "📌 Stepler Tikish",
    "rulon_orash": "🔄 Rulon O'rash",
    "rulonga_salafan": "🎁 Rulonga Salafan",
    "yopishtirma": "🔗 Yopishtirma",
    "adyol_tikish": "🧵 Adyol Tikish",
    "diplomat_tikish": "💼 Diplomat Tikish",
    "adyol_qoqish": "📫 Adyol Qoqish",
    "pastel_qoqish": "📬 Pastel Qoqish",
}
STATUS_LABELS = {
    "pending": "⏳ Kutmoqda",
    "approved": "✅ Tasdiqlangan",
    "adjusted": "🔧 Tuzatilgan",
    "rejected": "❌ Rad etilgan",
    "edit_requested": "✏️ O'zgartirish",
}
STATUS_CLS = {
    "pending": "by", "approved": "bg", "adjusted": "bc",
    "rejected": "br", "edit_requested": "bp",
}
PENALTY_LABELS = {
    "jarima": "💸 Jarima",
    "xaypsan1": "⚠️ 1-xaypsan",
    "xaypsan2": "🚫 2-xaypsan",
}
AVANS_MAX_PER_MONTH = 8

# ─── Yordamchi funksiyalar ──────────────────────────────────────────────────

def fmt(n) -> str:
    try:
        return f"{int(float(n)):,}".replace(",", " ")
    except Exception:
        return str(n)

def months_list() -> list:
    now = datetime.now()
    result = []
    for i in range(12):
        d = now.replace(day=1) - timedelta(days=30 * i)
        val = f"{d.year}-{d.month:02d}"
        result.append({"val": val, "label": f"{MONTHS_UZ[d.month]} {d.year}"})
    return result

def parse_month(oy_str: str) -> tuple:
    try:
        parts = oy_str.split("-")
        return int(parts[1]), int(parts[0])
    except Exception:
        now = datetime.now()
        return now.month, now.year

def stock_icon(p) -> str:
    if float(p.miqdor) <= float(p.min_threshold): return "🔴"
    if float(p.miqdor) <= float(p.yellow_threshold): return "🟡"
    return "🟢"

def stock_cls(p) -> str:
    if float(p.miqdor) <= float(p.min_threshold): return "cv-red"
    if float(p.miqdor) <= float(p.yellow_threshold): return "cv-yellow"
    return "cv-green"

def h(text) -> str:
    """HTML escape"""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# ─── CSS & HTML BASE ───────────────────────────────────────────────────────

CSS = """
:root{
  --bg:#F4F6FA;--bg2:#FFFFFF;--bg3:#F6F7FB;--bg4:#EDF0F7;--bg5:#E4E8F2;
  --border:#ECEFF6;--border2:#DDE2EF;
  --accent:#4F46E5;--accent-dim:#EEF0FF;
  --green:#0E9F6E;--green-dim:#E6F7F1;
  --yellow:#D97706;--yellow-dim:#FEF3E2;
  --red:#E02D2D;--red-dim:#FDECEC;
  --purple:#7C3AED;--purple-dim:#F3E8FF;
  --cyan:#0891B2;--cyan-dim:#E0F4F9;
  --text:#15192B;--text2:#7A82A0;--text3:#535C78;
  --r:16px;
  --sh:0 1px 2px rgba(21,25,43,.04),0 6px 20px rgba(21,25,43,.06);
  --sh2:0 10px 30px rgba(21,25,43,.12);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Plus Jakarta Sans','Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh;display:flex}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
.sidebar{width:228px;min-height:100vh;background:var(--bg2);border-right:1px solid var(--border);
  display:flex;flex-direction:column;position:fixed;left:0;top:0;bottom:0;z-index:100;overflow-y:auto}
.sb-logo{padding:16px 14px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.sb-logo-icon{width:38px;height:38px;background:linear-gradient(135deg,#4F46E5,#8B5CF6);
  border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:17px;color:#fff;
  box-shadow:0 6px 14px rgba(79,70,229,.3)}
.sb-logo h1{font-size:13px;font-weight:800;color:var(--text)}
.sb-logo span{font-size:10.5px;color:var(--text2);font-weight:600}
.sb-nav{flex:1;padding:8px 8px}
.nav-sec{margin-bottom:14px}
.nav-sec-label{font-size:9.5px;font-weight:800;text-transform:uppercase;letter-spacing:.12em;
  color:var(--text2);padding:0 10px;margin-bottom:4px;display:block}
.nav-item{display:flex;align-items:center;gap:9px;padding:9px 11px;border-radius:11px;
  color:var(--text2);text-decoration:none;transition:all .15s;margin-bottom:2px;font-size:13px;font-weight:700}
.nav-item:hover{background:var(--bg3);color:var(--text)}
.nav-item.active{background:var(--accent-dim);color:var(--accent)}
.sb-footer{padding:10px 14px;border-top:1px solid var(--border);font-size:10.5px;color:var(--text2);font-weight:600}
.main{margin-left:228px;flex:1;display:flex;flex-direction:column;min-height:100vh}
.topbar{height:54px;background:rgba(244,246,250,.85);backdrop-filter:blur(8px);border-bottom:1px solid var(--border);
  display:flex;align-items:center;padding:0 18px;gap:10px;position:sticky;top:0;z-index:50}
.topbar-title{font-size:15px;font-weight:800;flex:1}
.topbar-right{display:flex;align-items:center;gap:8px}
.content{flex:1;padding:16px 18px}
.stats-grid{display:grid;gap:12px;margin-bottom:14px}
.sg-3{grid-template-columns:repeat(3,1fr)}
.sg-4{grid-template-columns:repeat(4,1fr)}
.sg-5{grid-template-columns:repeat(5,1fr)}
.stat-card{background:var(--bg2);border:none;border-radius:var(--r);padding:15px 16px;box-shadow:var(--sh)}
.stat-card.cl-green{border-left:none;box-shadow:var(--sh),inset 4px 0 0 var(--green)}
.stat-card.cl-yellow{box-shadow:var(--sh),inset 4px 0 0 var(--yellow)}
.stat-card.cl-red{box-shadow:var(--sh),inset 4px 0 0 var(--red)}
.stat-card.cl-blue{box-shadow:var(--sh),inset 4px 0 0 var(--accent)}
.stat-card.cl-purple{box-shadow:var(--sh),inset 4px 0 0 var(--purple)}
.stat-card.cl-cyan{box-shadow:var(--sh),inset 4px 0 0 var(--cyan)}
.s-label{font-size:10.5px;color:var(--text2);font-weight:800;margin-bottom:6px;text-transform:uppercase;letter-spacing:.05em}
.s-value{font-size:23px;font-weight:800;line-height:1;font-family:'Plus Jakarta Sans',monospace}
.s-sub{font-size:11px;color:var(--text2);margin-top:4px;font-weight:600}
.cv-green{color:var(--green)}.cv-yellow{color:var(--yellow)}
.cv-red{color:var(--red)}.cv-blue{color:var(--accent)}
.cv-purple{color:var(--purple)}.cv-cyan{color:var(--cyan)}
.card{background:var(--bg2);border:none;border-radius:var(--r);padding:16px;margin-bottom:12px;box-shadow:var(--sh)}
.card-hd{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:8px}
.card-title{font-size:14px;font-weight:800;display:flex;align-items:center;gap:7px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}
.grid-main{display:grid;grid-template-columns:2fr 1fr;gap:12px}
.tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px;border-bottom:none;padding-bottom:4px}
.tab{padding:8px 16px;border-radius:99px;font-size:12.5px;font-weight:800;cursor:pointer;
  border:none;background:var(--bg2);color:var(--text2);text-decoration:none;transition:all .15s;box-shadow:var(--sh)}
.tab:hover{color:var(--text)}
.tab.active{background:var(--accent);color:#fff;box-shadow:0 6px 14px rgba(79,70,229,.3)}
.cat-tabs{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px}
.cat-tab{padding:7px 14px;border-radius:99px;background:var(--bg2);border:none;box-shadow:var(--sh);
  color:var(--text2);text-decoration:none;font-size:12px;font-weight:800;transition:all .15s}
.cat-tab:hover{color:var(--text)}
.cat-tab.active{background:var(--accent);color:#fff;box-shadow:0 6px 14px rgba(79,70,229,.3)}
.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:12px}
.toolbar-left{display:flex;gap:7px;flex:1;flex-wrap:wrap;align-items:center}
.search-wrap{position:relative;flex:1;min-width:160px;max-width:300px}
.search-wrap input{width:100%;background:var(--bg2);border:1px solid var(--border);
  border-radius:12px;padding:9px 12px 9px 32px;color:var(--text);font-size:13px;outline:none;box-shadow:var(--sh)}
.search-wrap input:focus{border-color:var(--accent)}
.si{position:absolute;left:11px;top:50%;transform:translateY(-50%);color:var(--text2);font-size:13px}
.fsel{background:var(--bg2);border:1px solid var(--border);border-radius:12px;
  padding:9px 12px;color:var(--text);font-size:13px;outline:none;cursor:pointer;font-weight:600;box-shadow:var(--sh)}
.fsel:focus{border-color:var(--accent)}
.fsel option{background:var(--bg2)}
.tbl-wrap{overflow-x:auto;border-radius:12px}
.ml4{margin-left:4px}
.mb12{margin-bottom:12px}
table{width:100%;border-collapse:collapse;font-size:12.5px}
th{background:var(--bg3);color:var(--text2);font-size:10px;font-weight:800;
  text-transform:uppercase;letter-spacing:.07em;padding:9px 12px;text-align:left;
  border-bottom:1px solid var(--border);white-space:nowrap}
th:first-child{border-radius:10px 0 0 0}
th:last-child{border-radius:0 10px 0 0}
td{padding:9px 12px;border-bottom:1px solid var(--border);vertical-align:middle}
tr:last-child td{border-bottom:none}
tr:hover td{background:var(--bg3)}
.td-n{font-family:'Plus Jakarta Sans',monospace;font-weight:700}
.item-card{background:var(--bg2);border:none;border-radius:var(--r);
  padding:14px 16px;margin-bottom:9px;transition:all .15s;box-shadow:var(--sh)}
.item-card:hover{transform:translateY(-1px);box-shadow:var(--sh2)}
.ic-hd{display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:7px;gap:8px}
.ic-title{font-size:14px;font-weight:800;display:flex;align-items:center;gap:6px;flex-wrap:wrap}
.ic-sub{font-size:12px;color:var(--text2);margin-top:2px;font-weight:600}
.ic-meta{display:flex;gap:12px;flex-wrap:wrap;font-size:12px;color:var(--text2);margin-bottom:8px;font-weight:600}
.ic-meta strong{color:var(--text)}
.ic-actions{display:flex;gap:6px;flex-wrap:wrap}
.badge{display:inline-flex;align-items:center;gap:3px;padding:3px 10px;
  border-radius:99px;font-size:11px;font-weight:800;white-space:nowrap;background:var(--bg4);color:var(--text3)}
.badge.red{background:var(--red-dim);color:var(--red)}
.badge.green{background:var(--green-dim);color:var(--green)}
.badge.yellow{background:var(--yellow-dim);color:var(--yellow)}
.bg{background:var(--green-dim);color:var(--green)}
.by{background:var(--yellow-dim);color:var(--yellow)}
.br{background:var(--red-dim);color:var(--red)}
.bb{background:var(--accent-dim);color:var(--accent)}
.bp{background:var(--purple-dim);color:var(--purple)}
.bc{background:var(--cyan-dim);color:var(--cyan)}
.bgr{background:var(--bg4);color:var(--text2)}
.btn{display:inline-flex;align-items:center;gap:5px;padding:8px 15px;border-radius:11px;
  font-size:12.5px;font-weight:800;cursor:pointer;border:none;text-decoration:none;
  transition:all .15s;white-space:nowrap}
.btn:active{transform:scale(.96)}
.btn-p{background:var(--accent);color:#fff;box-shadow:0 5px 12px rgba(79,70,229,.28)}
.btn-p:hover{background:#4338CA}
.btn-g{background:var(--green);color:#fff;box-shadow:0 5px 12px rgba(14,159,110,.28)}
.btn-g:hover{background:#0B8459}
.btn-s{background:var(--bg2);color:var(--text);border:1px solid var(--border2);box-shadow:var(--sh)}
.btn-s:hover{background:var(--bg3)}
.btn-d{background:var(--red-dim);color:var(--red)}
.btn-d:hover{background:#FAD9D9}
.btn-w{background:var(--yellow-dim);color:var(--yellow)}
.btn-cy{background:var(--cyan-dim);color:var(--cyan)}
.btn-pu{background:var(--purple-dim);color:var(--purple)}
.btn-sm{padding:5px 11px!important;font-size:11px!important;border-radius:9px!important}
.btn-xs{padding:3px 8px;font-size:10.5px;border-radius:8px}
.btn-green{background:var(--green)!important;color:#fff!important}
.btn-green:hover{background:#0B8459!important}
.btn-red{background:var(--red)!important;color:#fff!important}
.btn-red:hover{background:#C42222!important}
.btn-outline{background:var(--bg2);color:var(--text2);border:1px solid var(--border2)}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.fg{margin-bottom:11px}
.fl{display:block;font-size:10px;font-weight:800;color:var(--text2);
  margin-bottom:4px;text-transform:uppercase;letter-spacing:.06em}
input[type=text],input[type=number],input[type=date],
input[type=password],select,textarea{
  width:100%;background:var(--bg2);border:1px solid var(--border2);border-radius:12px;
  padding:9px 12px;color:var(--text);font-size:13px;outline:none;transition:all .15s;font-weight:600;
  font-family:inherit}
input:focus,select:focus,textarea:focus{border-color:var(--accent);box-shadow:0 0 0 3px var(--accent-dim)}
select option{background:var(--bg2)}
textarea{resize:vertical;min-height:64px}
.input-hint{font-size:10.5px;color:var(--text2);margin-top:3px;font-weight:600}
.overlay{display:none;position:fixed;inset:0;background:rgba(21,25,43,.45);
  z-index:400;align-items:center;justify-content:center;backdrop-filter:blur(5px)}
.overlay.open{display:flex}
.modal{background:var(--bg2);border:none;border-radius:20px;
  padding:22px;max-width:480px;width:94%;box-shadow:var(--sh2);
  max-height:90vh;overflow-y:auto}
.modal.wide{max-width:640px}
.modal.full{max-width:860px}
.modal-title{font-size:16px;font-weight:800;margin-bottom:4px;display:flex;align-items:center;gap:8px}
.modal-sub{font-size:12.5px;color:var(--text2);margin-bottom:15px;font-weight:600}
.modal-footer{display:flex;justify-content:flex-end;gap:8px;
  margin-top:15px;padding-top:13px;border-top:1px solid var(--border)}
.alert{padding:10px 13px;border-radius:12px;font-size:12.5px;margin-bottom:10px;
  display:flex;align-items:center;gap:8px;font-weight:700}
.alert-w{background:var(--yellow-dim);border:none;color:var(--yellow)}
.alert-r{background:var(--red-dim);border:none;color:var(--red)}
.alert-g{background:var(--green-dim);border:none;color:var(--green)}
.alert-b{background:var(--accent-dim);border:none;color:var(--accent)}
.pagination{display:flex;align-items:center;gap:4px;padding-top:10px;flex-wrap:wrap}
.pagination a{padding:6px 11px;border-radius:9px;font-size:11.5px;border:none;
  background:var(--bg2);color:var(--text2);text-decoration:none;font-weight:800;box-shadow:var(--sh)}
.pagination a:hover{color:var(--text)}
.pagination a.act{background:var(--accent);color:#fff}
.pinfo{font-size:11px;color:var(--text2);padding:0 6px;font-weight:600}
.mono{font-family:'Plus Jakarta Sans',monospace}
.fw7{font-weight:700}.fw8{font-weight:800}
.t-sm{font-size:11.5px}.t-xs{font-size:10.5px}
.t-muted{color:var(--text2)}.t-dim{color:var(--text3)}
.tag{display:inline-flex;background:var(--bg4);border:none;
  padding:2px 8px;border-radius:7px;font-size:10px;color:var(--text3);font-weight:700}
.divider{border:none;border-top:1px solid var(--border);margin:10px 0}
.empty-state{text-align:center;padding:44px 20px;color:var(--text2);font-weight:600}
.stock-bar{display:flex;align-items:center;gap:7px}
.bar-track{flex:1;background:var(--bg4);border-radius:99px;height:6px;
  overflow:hidden;min-width:50px;max-width:90px;border:none}
.bar-fill{height:100%;border-radius:99px}
.worker-avatar{width:36px;height:36px;border-radius:12px;
  background:linear-gradient(135deg,var(--accent),var(--purple));
  display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:800;color:#fff}
.toast{position:fixed;bottom:20px;right:20px;z-index:999;display:flex;flex-direction:column;gap:8px}
.toast-item{background:var(--bg2);border:none;border-radius:14px;
  padding:12px 18px;font-size:13px;font-weight:800;box-shadow:var(--sh2);
  animation:slideIn .25s ease;display:flex;align-items:center;gap:9px}
@keyframes slideIn{from{opacity:0;transform:translateX(20px)}to{opacity:1;transform:translateX(0)}}

/* ═══ MOBIL MOSLASHTIRISH (telefon va planshet) ═══ */
@media(max-width:480px){
  body{font-size:13px}
  .sidebar{display:none}
  .sidebar.open{display:flex;position:fixed;top:0;left:0;width:82%;height:100vh;z-index:1000;box-shadow:var(--sh2)}
  .main{margin-left:0;padding:12px 8px}
  .hamburger{display:flex !important;position:fixed;top:12px;left:12px;z-index:999;
    background:var(--accent);color:white;width:42px;height:42px;border-radius:13px;
    align-items:center;justify-content:center;border:none;cursor:pointer;font-size:20px;
    box-shadow:0 6px 14px rgba(79,70,229,.35)}
  .card{padding:13px;border-radius:14px}
  h1{font-size:18px !important}
  h2{font-size:15px !important}
  h3{font-size:14px !important}
  table{font-size:11px}
  th,td{padding:7px 5px !important}
  .btn{padding:7px 11px;font-size:11.5px}
  .btn-sm{padding:4px 9px;font-size:10.5px}
  .stat-card{padding:11px}
  .stat-num{font-size:20px !important}
  .stat-label{font-size:10px}
  .filters{flex-direction:column;gap:6px}
  .filters > *{width:100%}
  .grid-2,.grid-3,.grid-4{grid-template-columns:1fr !important}
  .modal{max-width:96% !important;padding:15px !important;font-size:13px}
  .header-bar{flex-direction:column;gap:8px;padding:10px 12px}
  .search-wrap{max-width:100% !important;width:100%}
  .badge{font-size:9.5px;padding:2px 7px}
  .nav-section{padding:6px 10px}
  .nav-link{padding:8px 12px;font-size:12px}
  .pg-num{padding:4px 8px;font-size:11px}
  .col-hide-mobile{display:none}
  .row-stack > *{display:block;width:100%;margin-bottom:6px}
}

/* Hamburger menyusi (faqat mobile uchun) */
.hamburger{display:none}

/* Tablet uchun */
@media(min-width:481px) and (max-width:768px){
  .sidebar{width:0;overflow:hidden}
  .main{margin-left:0}
  .sg-5,.sg-4{grid-template-columns:repeat(2,1fr)}
  .form-row{grid-template-columns:1fr}
}
"""


import hashlib, time

# ── SIMPLE SESSION AUTH ───────────────────────────────────────────────────────
_sessions: dict = {}   # {token: {"expires": timestamp}}
_SESSION_TTL = 60 * 60 * 8   # 8 soat

def _make_token(ip: str) -> str:
    raw = f"{ip}{time.time()}{SECRET_KEY}"
    return hashlib.sha256(raw.encode()).hexdigest()

def _check_session(request: web.Request) -> bool:
    token = request.cookies.get("wpsession")
    if not token:
        return False
    sess  = _sessions.get(token)
    if not sess:
        return False
    if time.time() > sess["expires"]:
        _sessions.pop(token, None)
        return False
    return True

def _current(request: web.Request) -> dict:
    """Joriy sessiya ma'lumoti: {user_id, role} yoki {}."""
    token = request.cookies.get("wpsession")
    if not token:
        return {}
    return _sessions.get(token, {})

def _require_auth(handler):
    """Decorator — kirish tekshiruvi (har qanday rol)."""
    async def wrapper(request: web.Request):
        if not _check_session(request):
            raise web.HTTPFound("/web/login")
        return await handler(request)
    return wrapper

def _require_role(*allowed_roles):
    """Decorator — faqat ruxsat etilgan rollar uchun.
    allowed_roles: 'admin', 'omborchi', 'nazoratchi' ...
    Admin/superadmin doim ruxsat etiladi."""
    def deco(handler):
        async def wrapper(request: web.Request):
            if not _check_session(request):
                raise web.HTTPFound("/web/login")
            sess = _current(request)
            role = sess.get("role", "")
            if role in ("admin", "superadmin"):
                return await handler(request)
            if allowed_roles and role not in allowed_roles:
                # Ruxsat yo'q — o'z paneliga yo'naltirish
                raise web.HTTPFound(_role_home(role))
            return await handler(request)
        return wrapper
    return deco

def _role_home(role: str) -> str:
    """Rolga mos bosh sahifa."""
    return {
        "omborchi":   "/web/ombor-panel",
        "nazoratchi": "/web/inspektor-panel",
        "admin":      "/web/",
        "superadmin": "/web/",
    }.get(role, "/web/login")



def _base(title: str, active: str, content: str, extra_js: str = "", role: str = "admin") -> str:
    nav_items = [
        ("dashboard",        "🏠", "Bosh sahifa",      "/web/"),
        ("ombor-home",       "🏠", "Ombor bosh",       "/web/ombor-panel"),
        ("ombor-ops",        "➕", "Kirim/Chiqim",     "/web/ombor-panel/operatsiya"),
        ("insp-home",        "🏠", "Nazorat bosh",     "/web/inspektor-panel"),
        ("insp-pending",     "⏳", "Tezkor tasdiq",    "/web/inspektor-panel/pending"),
        ("insp-review",      "🔍", "Bittalab ko'rish", "/web/inspektor-panel/review"),
        ("production",       "🏭", "Ishlab chiqarish", "/web/production"),
        ("rulon-hisobot",    "🌀", "Rulon hisobot",    "/web/rulon-hisobot"),
        ("inventory_health", "🏥", "Ombor salomatligi","/web/health"),
        ("zero-stock",       "⚠️", "Tugaganlar",         "/web/zero-stock"),
        ("tayyor-chiqim",    "📤", "Tayyor chiqimi",     "/web/tayyor-chiqim"),
        ("activity",         "📡", "Real-time oqim",  "/web/activity"),
        ("quality",          "✅", "Sifat boshqaruvi", "/web/quality"),
        ("workers",          "👥", "Ishchilar",         "/web/workers"),
        ("topshiriqlar",     "📋", "Topshiriqlar",      "/web/topshiriqlar"),
        ("avans",       "💳", "Avans",             "/web/avans"),
        ("penalties",   "⚠️", "Jarimalar",         "/web/penalties"),
        ("warehouse",   "📦", "Mahsulotlar",       "/web/warehouse"),
        ("ombor-logs",  "📋", "Kirim/Chiqim",      "/web/warehouse/logs"),
        ("ombor-cats",  "🏭", "Ombor bo'limlari",  "/web/ombor"),
        ("rulon",       "🌀", "Rulonlar",           "/web/ombor/rulon"),
        ("gofra",       "📋", "Gofralar",           "/web/ombor/gofra"),
        ("zagatovka",   "✂️", "Zagatovka",          "/web/ombor/gofra_zagatovka"),
        ("xromazes",    "🖨️", "Xromazeslar",        "/web/ombor/xromazes"),
        ("laminat",     "✨", "Laminat xromazes",   "/web/ombor/laminat_xromazes"),
        ("yarim",       "🧩", "Yarim tayyor",       "/web/ombor/yarim_tayyor"),
        ("qolip",       "🔲", "Qoliplar",           "/web/ombor/qolip"),
        ("adyol-zp",    "🧩", "Adyol zapchast",     "/web/ombor/adyol_zapchast"),
        ("stanok-zp",   "🔧", "Stanok ehtiyot",     "/web/ombor/uskuna_zapchast"),
        ("salary",      "💰", "Maoshlar",           "/web/salary"),
        ("salary_projection", "🔮", "Maosh prognozi",  "/web/salary-projection"),
        ("prices",      "⚙️", "Narxlar",            "/web/prices"),
        ("reports",     "📊", "Ish hisobotlari",    "/web/reports"),
        ("ombor-report","🏭", "Ombor hisoboti",     "/web/ombor-report"),
        ("maosh-report","💼", "Maosh hisoboti",     "/web/maosh-report"),
        ("stats-adv",   "📊", "Statistika",         "/web/stats-advanced"),
        ("mat-flow",    "🔄", "Material oqimi",     "/web/material-flow"),
        ("notif",       "🔔", "Bildirishnomalar",   "/web/notifications"),  # bildirishnoma markazi
        ("help",        "❓", "Yordam",              "/web/help"),
        ("system",      "ℹ️", "Tizim",              "/web/system"),
        ("analytics",   "📈", "Ishchilar tahlili",  "/web/analytics"),
        ("quick",       "⚡", "Tezkor amallar",     "/web/quick"),
    ]
    if role == "omborchi":
        # Omborchi — faqat ombor bilan bog'liq
        sections = [
            ("📦 Ombor", ["ombor-home", "ombor-ops", "warehouse", "zero-stock", "tayyor-chiqim"]),
            ("Ombor bo'limlari", ["rulon","gofra","zagatovka","xromazes","laminat","yarim","qolip","adyol-zp","stanok-zp"]),
            ("📋 Operatsiyalar", ["ombor-logs", "ombor-report"]),
        ]
    elif role == "nazoratchi":
        # Nazoratchi — tekshirish, sifat, ishchilar
        sections = [
            ("✅ Nazorat", ["insp-home", "insp-review", "insp-pending", "quality"]),
            ("👥 Ishchilar", ["workers", "analytics"]),
            ("📊 Hisobot", ["reports", "activity"]),
        ]
    else:
        # Admin — to'liq
        sections = [
            ("📊 Analytics", ["dashboard", "production", "rulon-hisobot", "inventory_health", "activity", "quality"]),
            ("👥 Xodimlar", ["workers", "topshiriqlar", "avans", "penalties"]),
            ("Ombor umumiy", ["warehouse", "ombor-logs", "ombor-cats"]),
            ("Ombor bo'limlari", ["rulon","gofra","zagatovka","xromazes","laminat","yarim","qolip","adyol-zp","stanok-zp"]),
            ("Moliya", ["salary", "salary_projection", "prices"]),
            ("Hisobotlar", ["reports", "ombor-report", "maosh-report"]),
            ("Tahlil", ["stats-adv", "mat-flow", "analytics"]),
            ("Boshqa", ["quick", "notif", "help", "system"]),
        ]
    ni_map = {k: (ic, lbl, url) for k, ic, lbl, url in nav_items}
    nav_html = ""
    for sec_name, keys in sections:
        nav_html += f'<div class="nav-sec"><span class="nav-sec-label">{sec_name}</span>'
        for k in keys:
            ic, lbl, url = ni_map[k]
            cls = "nav-item active" if k == active else "nav-item"
            nav_html += f'<a href="{url}" class="{cls}"><span>{ic}</span>{lbl}</a>'
        nav_html += '</div>'
    now = datetime.now()
    return f"""<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{h(title)} — Quti Tsexi</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}


/* ═══ KUCHLI MOBIL RESPONSIVE ═══ */

/* Planshet va kichik ekranlar (≤ 1024px) */
@media (max-width: 1024px) {{
  .sidebar {{ width: 180px !important; }}
  .main {{ padding: 16px !important; }}
  .stats-grid {{ grid-template-columns: repeat(2, 1fr) !important; }}
}}

/* Telefon (≤ 768px) — sidebar yuqorida horizontal nav */
@media (max-width: 768px) {{
  body {{ font-size: 14px !important; }}
  .sidebar {{
    width: 100% !important; height: auto !important;
    position: relative !important;
    border-right: none !important;
    border-bottom: 1px solid #1e293b !important;
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch !important;
  }}
  .sidebar h1, .sidebar .brand {{ font-size: 16px !important; padding: 10px !important; text-align: center !important; }}
  .sidebar nav {{
    display: flex !important; flex-wrap: nowrap !important;
    gap: 4px !important; padding: 8px !important;
    overflow-x: auto !important;
  }}
  .sidebar nav a {{
    padding: 8px 10px !important;
    font-size: 12px !important;
    border-radius: 8px !important;
    flex: 0 0 auto !important;
    white-space: nowrap !important;
    text-align: center !important;
  }}
  .main {{ margin-left: 0 !important; padding: 12px !important; max-width: 100% !important; }}
  .card {{ padding: 12px !important; margin-bottom: 10px !important; border-radius: 10px !important; }}
  h1 {{ font-size: 18px !important; line-height: 1.3 !important; }}
  h2 {{ font-size: 15px !important; line-height: 1.3 !important; }}
  h3 {{ font-size: 14px !important; }}
  p  {{ font-size: 13px !important; }}

  /* Jadvallar — gorizontal scroll */
  .table-wrap {{ overflow-x: auto !important; -webkit-overflow-scrolling: touch !important; width: 100% !important; }}
  table {{ font-size: 12px !important; min-width: 100% !important; }}
  table th, table td {{ padding: 6px 4px !important; white-space: nowrap !important; }}

  /* Statistika kartochkalari — 2 ustun */
  .stats-grid {{ grid-template-columns: repeat(2, 1fr) !important; gap: 8px !important; }}
  .stat-card {{ padding: 12px !important; }}
  .stat-value {{ font-size: 20px !important; }}
  .stat-label {{ font-size: 10px !important; }}
  .stat-trend {{ font-size: 10px !important; }}

  /* Charts — kichikroq */
  .chart-card {{ min-height: 280px !important; }}
  .charts-row {{ grid-template-columns: 1fr !important; gap: 10px !important; }}

  /* Tugmalar */
  .btn {{
    padding: 10px 14px !important;
    font-size: 13px !important;
    width: 100% !important;
    margin-bottom: 4px !important;
  }}
  .btn-group .btn {{ width: auto !important; flex: 1 !important; }}

  /* Formalar */
  input, select, textarea {{
    font-size: 16px !important;  /* iOS zoom emas */
    padding: 10px !important;
    width: 100% !important;
    box-sizing: border-box !important;
  }}
  .form-row {{ grid-template-columns: 1fr !important; }}

  /* Period switcher (kichkina tugmalar) */
  .period-switcher {{ flex-wrap: wrap !important; gap: 4px !important; }}
  .btn-period {{ padding: 6px 10px !important; font-size: 12px !important; flex: 1 1 auto !important; }}

  /* Alert/notification */
  .alert, .alert-r {{ padding: 10px !important; font-size: 13px !important; }}

  /* Modal/overlay */
  .overlay {{ padding: 10px !important; }}
  .modal {{ width: 95% !important; max-width: 95% !important; padding: 16px !important; }}
}}

/* Juda kichik telefonlar (≤ 480px) */
@media (max-width: 480px) {{
  body {{ font-size: 13px !important; }}
  .main {{ padding: 8px !important; }}
  .card {{ padding: 10px !important; border-radius: 8px !important; }}
  h1 {{ font-size: 16px !important; }}
  h2 {{ font-size: 14px !important; }}

  /* Statistika — 1 ustun */
  .stats-grid {{ grid-template-columns: 1fr !important; }}
  .stat-card {{ padding: 10px !important; }}
  .stat-value {{ font-size: 18px !important; }}

  /* Jadvallar yana kichik */
  table {{ font-size: 11px !important; }}
  table th, table td {{ padding: 4px 3px !important; }}

  /* Sidebar nav — to'liq yashir kichik tugmalari */
  .sidebar nav a {{ padding: 6px 8px !important; font-size: 11px !important; }}

  /* Form */
  input, select, textarea {{ padding: 8px !important; font-size: 14px !important; }}

  /* Rank list */
  .rank-row {{ grid-template-columns: 28px 1fr auto !important; gap: 6px !important; padding: 6px !important; }}
  .rank-stat {{ display: none !important; }}

  /* Chain card */
  .chain-card {{ padding: 8px !important; }}
  .chain-count {{ font-size: 18px !important; }}
}}

/* Touch interactivlik */
@media (hover: none) and (pointer: coarse) {{
  .btn:active {{ transform: scale(0.97) !important; }}
  .card:active {{ background: rgba(99,102,241,0.05) !important; }}
}}


/* ═══ KUCHAYTIRILGAN ANIMATSIYALAR VA STILLAR ═══ */
@keyframes slideIn {{
  from {{ opacity: 0; transform: translateY(10px); }}
  to   {{ opacity: 1; transform: translateY(0); }}
}}
@keyframes pulse {{
  0%, 100% {{ opacity: 1 }}
  50%      {{ opacity: 0.6 }}
}}
.card, .stat-card, .notif-row, .health-row {{
  animation: slideIn 0.3s ease-out;
}}
.alert-pulse {{ animation: pulse 2s infinite }}

/* Tooltip */
[data-tooltip] {{ position: relative }}
[data-tooltip]:hover::after {{
  content: attr(data-tooltip);
  position: absolute;
  bottom: 100%;
  left: 50%;
  transform: translateX(-50%);
  background: #0f172a;
  color: #fff;
  padding: 6px 10px;
  border-radius: 6px;
  font-size: 11px;
  white-space: nowrap;
  border: 1px solid #334155;
  z-index: 1000;
}}

/* Tab system */
.tabs {{
  display: flex;
  gap: 4px;
  background: rgba(30,41,59,0.5);
  padding: 4px;
  border-radius: 10px;
  margin-bottom: 16px;
  overflow-x: auto;
}}
.tab-btn {{
  padding: 8px 16px;
  border-radius: 6px;
  font-size: 13px;
  font-weight: 600;
  color: var(--muted);
  background: transparent;
  border: none;
  cursor: pointer;
  white-space: nowrap;
  transition: all .15s;
}}
.tab-btn:hover {{ background: rgba(99,102,241,0.1); color: var(--fg) }}
.tab-btn.active {{ background: #6366f1; color: #fff }}

/* Modal */
.modal-overlay {{
  position: fixed; inset: 0; background: rgba(0,0,0,0.7);
  display: none; align-items: center; justify-content: center;
  z-index: 9999; padding: 20px;
}}
.modal-overlay.show {{ display: flex }}
.modal-box {{
  background: #0f172a; border: 1px solid #1e293b;
  border-radius: 16px; padding: 24px; max-width: 500px; width: 100%;
  max-height: 90vh; overflow-y: auto;
}}
.modal-close {{
  float: right; background: none; border: none; color: var(--muted);
  font-size: 24px; cursor: pointer; padding: 0 8px;
}}

/* Toast */
.toast-container {{
  position: fixed; top: 20px; right: 20px; z-index: 9999;
  display: flex; flex-direction: column; gap: 8px;
}}
.toast {{
  background: #0f172a; border: 1px solid #1e293b; border-radius: 10px;
  padding: 12px 16px; min-width: 250px; max-width: 350px;
  display: flex; align-items: center; gap: 10px;
  animation: slideIn 0.3s ease-out;
  box-shadow: 0 10px 25px rgba(0,0,0,0.3);
}}
.toast-success {{ border-left: 4px solid var(--green) }}
.toast-error   {{ border-left: 4px solid var(--red) }}
.toast-info    {{ border-left: 4px solid #3b82f6 }}

/* Progress ring */
.progress-ring {{
  position: relative;
  width: 80px; height: 80px;
}}
.progress-ring svg {{
  transform: rotate(-90deg);
  width: 100%; height: 100%;
}}
.progress-ring circle {{
  fill: transparent;
  stroke-width: 8;
  stroke-linecap: round;
}}
.progress-ring .bg {{ stroke: rgba(30,41,59,0.6) }}
.progress-ring .fg {{ stroke: #6366f1; transition: stroke-dashoffset 0.5s }}
.progress-ring .label {{
  position: absolute; inset: 0;
  display: flex; align-items: center; justify-content: center;
  font-weight: 800; font-size: 18px;
}}

/* Badge */
.badge-dot {{
  display: inline-block; width: 8px; height: 8px;
  border-radius: 50%; vertical-align: middle; margin-right: 4px;
}}
.badge-dot.green  {{ background: var(--green) }}
.badge-dot.red    {{ background: var(--red) }}
.badge-dot.yellow {{ background: #f59e0b }}
.badge-dot.blue   {{ background: #3b82f6 }}

/* Floating action button */
.fab {{
  position: fixed; bottom: 20px; right: 20px;
  width: 56px; height: 56px; border-radius: 28px;
  background: linear-gradient(135deg, #6366f1, #8b5cf6);
  color: #fff; font-size: 24px;
  display: flex; align-items: center; justify-content: center;
  cursor: pointer; border: none;
  box-shadow: 0 8px 25px rgba(99,102,241,0.4);
  transition: transform .15s;
  z-index: 999;
}}
.fab:hover {{ transform: scale(1.1) }}

@media (max-width: 768px) {{
  .modal-box {{ padding: 16px; max-height: 100vh; border-radius: 8px; }}
  .toast-container {{ right: 10px; left: 10px; top: 10px; }}
  .toast {{ max-width: none }}
  .fab {{ bottom: 80px; right: 16px; width: 48px; height: 48px; font-size: 20px; border-radius: 24px; }}
}}



/* Top bar */
.topbar {{
  display: flex;
  gap: 12px;
  margin-bottom: 20px;
  align-items: center;
  flex-wrap: wrap;
}}
.topbar-search {{
  flex: 1;
  display: flex;
  gap: 4px;
  background: rgba(30,41,59,0.5);
  border-radius: 10px;
  padding: 4px;
  min-width: 200px;
}}
.topbar-search input {{
  flex: 1;
  background: transparent;
  border: none;
  color: var(--fg);
  padding: 8px 12px;
  font-size: 13px;
  outline: none;
}}
.topbar-search-btn {{
  background: #6366f1;
  color: #fff;
  border: none;
  border-radius: 6px;
  padding: 6px 14px;
  cursor: pointer;
  font-size: 12px;
  font-weight: 600;
}}
.topbar-actions {{
  display: flex;
  gap: 6px;
}}
.topbar-icon {{
  width: 36px; height: 36px;
  display: flex; align-items: center; justify-content: center;
  background: rgba(30,41,59,0.5);
  border-radius: 8px;
  text-decoration: none;
  color: var(--fg);
  font-weight: 700;
  transition: all .15s;
}}
.topbar-icon:hover {{
  background: rgba(99,102,241,0.2);
  transform: translateY(-2px);
}}

@media (max-width: 768px) {{
  .topbar {{ gap: 6px }}
  .topbar-search input {{ font-size: 12px; padding: 6px 10px }}
  .topbar-search-btn {{ padding: 6px 10px; font-size: 11px }}
  .topbar-icon {{ width: 32px; height: 32px; font-size: 12px }}
}}



/* Print friendly */
@media print {{
  .sidebar, .topbar, .fab, .btn, .modal-overlay, .toast-container {{
    display: none !important;
  }}
  body, .main, .card {{
    background: white !important;
    color: black !important;
    box-shadow: none !important;
  }}
  .main {{ margin-left: 0 !important; padding: 0 !important; }}
  .card {{ border: 1px solid #ddd !important; page-break-inside: avoid; }}
  h1, h2, h3 {{ color: black !important; }}
  table {{ font-size: 11px !important; }}
  th {{ background: #f1f5f9 !important; }}
}}

/* Hover effects */
.btn-primary {{
  background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%);
  color: #fff;
  border: none;
  padding: 10px 20px;
  border-radius: 10px;
  font-weight: 700;
  cursor: pointer;
  transition: all .2s;
  box-shadow: 0 4px 12px rgba(99,102,241,0.3);
}}
.btn-primary:hover {{
  transform: translateY(-2px);
  box-shadow: 0 8px 20px rgba(99,102,241,0.4);
}}
.btn-primary:active {{ transform: translateY(0) }}

/* Card hover */
.card-clickable {{
  cursor: pointer;
  transition: all .2s;
}}
.card-clickable:hover {{
  transform: translateY(-4px);
  box-shadow: 0 12px 30px rgba(0,0,0,0.3);
  border-color: rgba(99,102,241,0.4);
}}

/* Skeleton loader */
@keyframes shimmer {{
  0% {{ background-position: -200% 0 }}
  100% {{ background-position: 200% 0 }}
}}
.skeleton {{
  background: linear-gradient(90deg, rgba(30,41,59,0.4) 0%, rgba(51,65,85,0.6) 50%, rgba(30,41,59,0.4) 100%);
  background-size: 200% 100%;
  animation: shimmer 1.5s infinite;
  border-radius: 6px;
  height: 16px;
}}

/* Gradient text */
.gradient-text {{
  background: linear-gradient(135deg, #6366f1, #8b5cf6, #ec4899);
  -webkit-background-clip: text;
  background-clip: text;
  -webkit-text-fill-color: transparent;
  font-weight: 800;
}}

/* Glass effect */
.glass {{
  background: rgba(15,23,42,0.6);
  backdrop-filter: blur(10px);
  -webkit-backdrop-filter: blur(10px);
  border: 1px solid rgba(99,102,241,0.2);
}}

/* Status indicators */
.status-pill {{
  display: inline-flex;
  align-items: center;
  gap: 6px;
  padding: 4px 10px;
  border-radius: 20px;
  font-size: 11px;
  font-weight: 700;
}}
.status-pill::before {{
  content: '';
  width: 6px; height: 6px;
  border-radius: 50%;
  display: inline-block;
}}
.status-online::before  {{ background: var(--green); box-shadow: 0 0 6px var(--green); }}
.status-offline::before {{ background: var(--red) }}
.status-pending::before {{ background: #f59e0b; animation: pulse 2s infinite }}

/* Counter big */
.counter-big {{
  font-size: 36px;
  font-weight: 900;
  line-height: 1;
  margin: 8px 0;
}}

/* Trend arrow */
.trend-up {{
  display: inline-flex;
  align-items: center;
  gap: 2px;
  color: var(--green);
  font-weight: 700;
  font-size: 12px;
}}
.trend-up::before {{ content: 'up'; font-size: 10px }}
.trend-down {{
  display: inline-flex;
  align-items: center;
  gap: 2px;
  color: var(--red);
  font-weight: 700;
  font-size: 12px;
}}
.trend-down::before {{ content: 'down'; font-size: 10px }}

/* Mini stat cards (one-line) */
.mini-stat {{
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 8px 12px;
  background: rgba(30,41,59,0.4);
  border-radius: 8px;
  font-size: 12px;
}}
.mini-stat-icon {{ font-size: 16px }}
.mini-stat-label {{ color: var(--muted) }}
.mini-stat-value {{ font-weight: 700; margin-left: auto }}

/* Sidebar nav-sec collapsible */
.nav-sec {{
  margin-bottom: 8px;
}}
.nav-sec-label {{
  display: block;
  font-size: 10px;
  font-weight: 700;
  color: var(--muted);
  text-transform: uppercase;
  letter-spacing: 1px;
  padding: 8px 14px 4px;
  opacity: 0.7;
}}
.nav-item {{
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 7px 14px;
  margin: 0 6px;
  border-radius: 8px;
  font-size: 12.5px;
  color: var(--fg);
  text-decoration: none;
  transition: all .12s;
}}
.nav-item:hover {{
  background: rgba(99,102,241,0.1);
  transform: translateX(2px);
}}
.nav-item.active {{
  background: linear-gradient(90deg, rgba(99,102,241,0.2), rgba(139,92,246,0.1));
  border-left: 3px solid #6366f1;
  font-weight: 700;
}}
.nav-item > span:first-child {{ font-size: 16px }}

/* Sidebar header */
.sidebar-header {{
  padding: 16px 14px;
  border-bottom: 1px solid #1e293b;
  margin-bottom: 12px;
}}
.sidebar-logo {{
  font-size: 18px;
  font-weight: 800;
  background: linear-gradient(135deg, #6366f1, #8b5cf6);
  -webkit-background-clip: text;
  -webkit-text-fill-color: transparent;
}}
.sidebar-sub {{
  font-size: 10px;
  color: var(--muted);
  margin-top: 2px;
}}

/* Empty state */
.empty-state {{
  text-align: center;
  padding: 60px 20px;
  color: var(--muted);
}}
.empty-state-icon {{
  font-size: 48px;
  margin-bottom: 12px;
  opacity: 0.5;
}}
.empty-state-title {{
  font-size: 18px;
  font-weight: 700;
  margin-bottom: 8px;
  color: var(--fg);
}}

/* Loading spinner */
@keyframes spin {{
  to {{ transform: rotate(360deg) }}
}}
.spinner {{
  display: inline-block;
  width: 16px; height: 16px;
  border: 2px solid rgba(99,102,241,0.2);
  border-top-color: #6366f1;
  border-radius: 50%;
  animation: spin 0.6s linear infinite;
}}

/* Sticky header */
.sticky-header {{
  position: sticky;
  top: 0;
  z-index: 100;
  background: var(--bg);
  padding: 12px 0;
  margin: -12px 0 12px;
  border-bottom: 1px solid #1e293b;
}}

/* Data badge */
.data-badge {{
  display: inline-flex;
  align-items: center;
  gap: 4px;
  padding: 2px 8px;
  border-radius: 4px;
  font-size: 10px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}}
.badge-new      {{ background: rgba(16,185,129,0.15); color: #6ee7b7 }}
.badge-hot      {{ background: rgba(239,68,68,0.15);  color: #fca5a5 }}
.badge-pro      {{ background: rgba(139,92,246,0.15); color: #c4b5fd }}
.badge-warning  {{ background: rgba(245,158,11,0.15); color: #fcd34d }}

/* Quick filter bar */
.filter-bar {{
  display: flex;
  gap: 6px;
  padding: 8px;
  background: rgba(30,41,59,0.5);
  border-radius: 10px;
  margin-bottom: 16px;
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}}
.filter-bar a {{
  padding: 6px 12px;
  border-radius: 6px;
  font-size: 12px;
  font-weight: 600;
  color: var(--muted);
  text-decoration: none;
  white-space: nowrap;
  transition: all .15s;
}}
.filter-bar a:hover {{
  background: rgba(99,102,241,0.1);
  color: var(--fg);
}}
.filter-bar a.active {{
  background: #6366f1;
  color: #fff;
}}

/* Avatar */
.avatar {{
  width: 32px; height: 32px;
  border-radius: 50%;
  background: linear-gradient(135deg, #6366f1, #8b5cf6);
  display: inline-flex;
  align-items: center;
  justify-content: center;
  color: #fff;
  font-weight: 700;
  font-size: 13px;
  flex-shrink: 0;
}}
.avatar-lg {{ width: 60px; height: 60px; font-size: 24px }}
.avatar-sm {{ width: 24px; height: 24px; font-size: 11px }}

/* Tag */
.tag {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 12px;
  font-size: 11px;
  font-weight: 600;
  background: rgba(99,102,241,0.15);
  color: #4F46E5;
  margin: 2px;
}}
.tag-removable::after {{
  content: ' x';
  margin-left: 4px;
  font-weight: 700;
  cursor: pointer;
}}

/* Divider */
.divider {{
  height: 1px;
  background: linear-gradient(90deg, transparent, #334155, transparent);
  margin: 20px 0;
}}

/* Section heading */
.section-head {{
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 14px;
  padding-bottom: 8px;
  border-bottom: 1px solid #1e293b;
}}
.section-head h2 {{ margin: 0 }}
.section-head .actions {{
  display: flex;
  gap: 6px;
}}

/* Comparison table */
.compare-table {{
  width: 100%;
  border-collapse: separate;
  border-spacing: 0;
}}
.compare-table th {{
  background: rgba(30,41,59,0.5);
  padding: 12px;
  text-align: left;
  font-size: 12px;
  font-weight: 700;
  color: var(--muted);
  text-transform: uppercase;
  letter-spacing: 0.5px;
}}
.compare-table td {{
  padding: 10px 12px;
  border-bottom: 1px solid rgba(30,41,59,0.5);
  font-size: 13px;
}}
.compare-table tr:hover td {{
  background: rgba(99,102,241,0.05);
}}

/* Action menu */
.action-menu {{
  position: relative;
  display: inline-block;
}}
.action-menu-trigger {{
  background: rgba(30,41,59,0.5);
  border: none;
  color: var(--fg);
  padding: 6px 10px;
  border-radius: 6px;
  cursor: pointer;
  font-size: 14px;
}}
.action-menu-content {{
  position: absolute;
  top: 100%;
  right: 0;
  background: #0f172a;
  border: 1px solid #1e293b;
  border-radius: 8px;
  min-width: 160px;
  box-shadow: 0 10px 30px rgba(0,0,0,0.4);
  display: none;
  z-index: 100;
  margin-top: 4px;
}}
.action-menu.open .action-menu-content {{ display: block }}
.action-menu-item {{
  display: flex;
  align-items: center;
  gap: 8px;
  padding: 10px 14px;
  font-size: 13px;
  color: var(--fg);
  text-decoration: none;
  transition: background .12s;
}}
.action-menu-item:hover {{
  background: rgba(99,102,241,0.1);
}}
.action-menu-item.danger {{ color: var(--red) }}
.action-menu-item.danger:hover {{ background: rgba(239,68,68,0.1) }}

</style>
</head>
<body>
<button class="hamburger" onclick="document.querySelector('.sidebar').classList.toggle('open')">☰</button><aside class="sidebar">
  <div class="sb-logo">
    <div class="sb-logo-icon">📦</div>
    <div><h1>Quti Tsexi</h1><span>Boshqaruv paneli</span></div>
  </div>
  <nav class="sb-nav">{nav_html}</nav>
  <div class="sb-footer">
    {now.strftime('%d.%m.%Y %H:%M')}
  </div>
</aside>
<div class="main">
  <header class="topbar">
    <div class="topbar-title">{h(title)}</div>
    <div class="topbar-right">
      <span class="t-xs t-muted">{now.strftime('%d.%m.%Y %H:%M')}</span>
    </div>
  </header>
  <div class="content">{content}</div>
</div>
<div id="toast-container" class="toast"></div>
<script>
function toast(msg, type='g') {{
  const c = document.getElementById('toast-container');
  const d = document.createElement('div');
  const colors = {{g:'var(--green)',r:'var(--red)',b:'var(--accent)',y:'var(--yellow)'}};
  d.className = 'toast-item';
  d.style.borderLeftColor = colors[type] || colors.g;
  d.innerHTML = msg;
  c.appendChild(d);
  setTimeout(() => d.remove(), 3500);
}}
function openModal(id) {{
  document.getElementById(id).classList.add('open');
}}
function closeModal(id) {{
  document.getElementById(id).classList.remove('open');
}}
document.addEventListener('keydown', e => {{
  if(e.key === 'Escape') document.querySelectorAll('.overlay.open').forEach(m => m.classList.remove('open'));
}});
{extra_js}
</script>


</body>
</html>"""

# ─── DASHBOARD ─────────────────────────────────────────────────────────────

@_require_auth
async def dashboard(request: web.Request):
    """Kuchaytirilgan dashboard — diagrammalar, reyting, ogohlantirishlar."""
    async with AsyncSessionLocal() as db:
        stats = await get_dashboard_stats(db)

        # Oxirgi 30 kun ishlab chiqarish
        today = date.today()
        daily_data = []
        for i in range(29, -1, -1):
            d = today - timedelta(days=i)
            r = await db.execute(
                select(
                    func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                    func.count(WorkEntry.id),
                ).where(
                    WorkEntry.work_date == d,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            row = r.one()
            daily_data.append({
                "date": d.strftime("%d.%m"),
                "income": float(row[0] or 0),
                "count": int(row[1] or 0),
            })

        # Top 10 ishchi (oylik)
        month_start = today.replace(day=1)
        r_top = await db.execute(
            select(
                User.full_name,
                func.count(WorkEntry.id).label("works"),
                func.coalesce(func.sum(WorkEntry.jami_summa), 0).label("income"),
            )
            .join(WorkEntry, WorkEntry.worker_id == User.id)
            .where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status == WorkStatus.approved,
            )
            .group_by(User.id, User.full_name)
            .order_by(func.coalesce(func.sum(WorkEntry.jami_summa), 0).desc())
            .limit(10)
        )
        top_workers = [{"name": r[0], "works": r[1], "income": float(r[2])} for r in r_top.all()]

        # Ish turlari bo'yicha bo'linish (oylik)
        r_types = await db.execute(
            select(
                WorkEntry.work_type,
                func.count(WorkEntry.id),
                func.coalesce(func.sum(WorkEntry.jami_summa), 0),
            )
            .where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status == WorkStatus.approved,
            )
            .group_by(WorkEntry.work_type)
            .order_by(func.count(WorkEntry.id).desc())
        )
        work_types_data = [{"type": r[0].value if r[0] else "?", "count": r[1], "income": float(r[2])} for r in r_types.all()]

        # Ombor holatlari (qizil/sariq/yashil)
        r_stock = await db.execute(
            select(
                sf.sum(sa_case((WarehouseProduct.miqdor <= WarehouseProduct.min_threshold, 1), else_=0)),
                sf.sum(sa_case(((WarehouseProduct.miqdor > WarehouseProduct.min_threshold) &
                                (WarehouseProduct.miqdor <= WarehouseProduct.yellow_threshold), 1), else_=0)),
                sf.sum(sa_case((WarehouseProduct.miqdor > WarehouseProduct.yellow_threshold, 1), else_=0)),
            ).where(WarehouseProduct.is_active == True)
        )
        stock_row = r_stock.one()
        stock_red    = int(stock_row[0] or 0)
        stock_yellow = int(stock_row[1] or 0)
        stock_green  = int(stock_row[2] or 0)

        # Kechagi kun bilan taqqoslash
        yesterday = today - timedelta(days=1)
        r_yest = await db.execute(
            select(
                func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                func.count(WorkEntry.id),
            ).where(
                WorkEntry.work_date == yesterday,
                WorkEntry.status == WorkStatus.approved,
            )
        )
        yest_row = r_yest.one()
        yest_income = float(yest_row[0] or 0)
        yest_works  = int(yest_row[1] or 0)

        today_income_val = stats.get("today_income", 0)
        today_works_val  = stats.get("today_works", 0)

        income_change = ((today_income_val - yest_income) / yest_income * 100) if yest_income > 0 else 0
        works_change  = ((today_works_val - yest_works) / yest_works * 100) if yest_works > 0 else 0

    today_income = fmt(today_income_val)
    today_works = today_works_val
    pending = stats.get("pending_works", 0)
    approved = stats.get("approved_works", 0)
    rejected = stats.get("rejected_works", 0)
    low_stock = stats.get("low_stock", 0)
    workers_count = stats.get("workers_count", 0)
    monthly_total = fmt(stats.get("monthly_total", 0))
    open_sessions = stats.get("open_sessions", 0)
    edit_req = stats.get("edit_requested", 0)

    # Diagramma uchun JSON
    import json
    daily_labels  = json.dumps([d["date"]   for d in daily_data])
    daily_incomes = json.dumps([d["income"] for d in daily_data])
    daily_counts  = json.dumps([d["count"]  for d in daily_data])

    types_labels = json.dumps([d["type"] for d in work_types_data])
    types_counts = json.dumps([d["count"] for d in work_types_data])
    types_colors = json.dumps([
        "#6366f1","#8b5cf6","#ec4899","#f43f5e","#f59e0b",
        "#10b981","#06b6d4","#3b82f6","#a855f7","#84cc16",
        "#f97316","#14b8a6","#0ea5e9","#d946ef","#22c55e",
    ])

    # Top ishchilar HTML
    top_html = ""
    medals = ["🥇","🥈","🥉"] + ["🏅"] * 7
    for i, w in enumerate(top_workers):
        medal = medals[i] if i < len(medals) else f"{i+1}."
        top_html += (
            f'<div class="rank-row">'
            f'<div class="rank-medal">{medal}</div>'
            f'<div class="rank-name">{h(w["name"])}</div>'
            f'<div class="rank-stat">{w["works"]} ish</div>'
            f'<div class="rank-income">{fmt(w["income"])}</div>'
            f'</div>'
        )
    if not top_html:
        top_html = '<p style="color:var(--muted);text-align:center;padding:20px">Bu oyda ish yo\'q</p>'

    # Change indicators
    inc_arrow = "↑" if income_change > 0 else ("↓" if income_change < 0 else "—")
    inc_color = "var(--green)" if income_change > 0 else ("var(--red)" if income_change < 0 else "var(--muted)")
    works_arrow = "↑" if works_change > 0 else ("↓" if works_change < 0 else "—")
    works_color = "var(--green)" if works_change > 0 else ("var(--red)" if works_change < 0 else "var(--muted)")

    low_warn = f'<div class="alert alert-r">🔴 {low_stock} ta mahsulot kam qoldi! <a href="/web/warehouse" style="color:var(--red);text-decoration:underline">Ko\'rish →</a></div>' if low_stock else ""

    content = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>

<h1 style="margin-bottom:4px">📊 Boshqaruv paneli</h1>
<p style="color:var(--muted);margin-bottom:20px">Real vaqt holati va statistika — {today.strftime('%d.%m.%Y')}</p>

{low_warn}

<!-- ASOSIY KO'RSATKICHLAR -->
<div class="stats-grid" style="margin-bottom:24px">
  <div class="stat-card">
    <div class="stat-label">💰 Bugungi daromad</div>
    <div class="stat-value">{today_income}</div>
    <div class="stat-trend" style="color:{inc_color}">{inc_arrow} {abs(income_change):.1f}% vs kecha</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">📋 Bugungi ishlar</div>
    <div class="stat-value">{today_works}</div>
    <div class="stat-trend" style="color:{works_color}">{works_arrow} {abs(works_change):.1f}% vs kecha</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">💵 Oylik jami</div>
    <div class="stat-value">{monthly_total}</div>
    <div class="stat-trend" style="color:var(--muted)">{today.strftime('%B')} oyi</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">👷 Faol smenalar</div>
    <div class="stat-value">{open_sessions}</div>
    <div class="stat-trend" style="color:var(--muted)">jami {workers_count} ishchi</div>
  </div>
</div>

<!-- STATUS KO'RSATKICHLARI -->
<div class="stats-grid" style="margin-bottom:24px;grid-template-columns:repeat(auto-fit,minmax(140px,1fr))">
  <div class="stat-mini" style="border-left:4px solid #f59e0b">
    <div class="stat-mini-label">⏳ Kutilmoqda</div>
    <div class="stat-mini-value">{pending}</div>
  </div>
  <div class="stat-mini" style="border-left:4px solid var(--green)">
    <div class="stat-mini-label">✅ Tasdiqlangan</div>
    <div class="stat-mini-value">{approved}</div>
  </div>
  <div class="stat-mini" style="border-left:4px solid var(--red)">
    <div class="stat-mini-label">❌ Rad etilgan</div>
    <div class="stat-mini-value">{rejected}</div>
  </div>
  <div class="stat-mini" style="border-left:4px solid #8b5cf6">
    <div class="stat-mini-label">✏️ Tahrir so'rovi</div>
    <div class="stat-mini-value">{edit_req}</div>
  </div>
</div>

<!-- DIAGRAMMALAR -->
<div class="charts-row">
  <div class="card chart-card">
    <h2 style="margin-bottom:12px">📈 Oxirgi 30 kun — daromad</h2>
    <div style="height:280px;position:relative"><canvas id="incomeChart"></canvas></div>
  </div>
  <div class="card chart-card">
    <h2 style="margin-bottom:12px">🥧 Ish turlari (oylik)</h2>
    <div style="height:280px;position:relative"><canvas id="typesChart"></canvas></div>
  </div>
</div>

<!-- OMBOR HOLATI VA RATING -->
<div class="charts-row" style="margin-top:20px">
  <div class="card">
    <h2 style="margin-bottom:12px">📦 Ombor holati</h2>
    <div class="stock-bars">
      <div class="stock-bar-row">
        <span class="stock-icon" style="color:var(--red)">🔴</span>
        <span class="stock-text">Kam qoldi (kritik)</span>
        <span class="stock-num">{stock_red}</span>
      </div>
      <div class="stock-bar-row">
        <span class="stock-icon" style="color:#f59e0b">🟡</span>
        <span class="stock-text">Sariq zonada</span>
        <span class="stock-num">{stock_yellow}</span>
      </div>
      <div class="stock-bar-row">
        <span class="stock-icon" style="color:var(--green)">🟢</span>
        <span class="stock-text">Yetarli</span>
        <span class="stock-num">{stock_green}</span>
      </div>
    </div>
    <div style="height:140px;position:relative;margin-top:12px"><canvas id="stockChart"></canvas></div>
    <a href="/web/warehouse" class="btn" style="margin-top:12px;display:block;text-align:center">📦 Ombor ko'rish</a>
  </div>

  <div class="card">
    <h2 style="margin-bottom:12px">🏆 Top ishchilar (oylik)</h2>
    <div class="rank-list">{top_html}</div>
  </div>
</div>

<style>
.stat-card {{
  background:linear-gradient(135deg,rgba(99,102,241,0.1) 0%,rgba(99,102,241,0.03) 100%);
  border:1px solid rgba(99,102,241,0.2);
  border-radius:14px;
  padding:18px;
  transition:transform .2s, box-shadow .2s;
}}
.stat-card:hover {{
  transform:translateY(-2px);
  box-shadow:0 10px 25px rgba(0,0,0,0.2);
}}
.stat-label {{ color:var(--muted); font-size:12px; font-weight:600; text-transform:uppercase; letter-spacing:.5px }}
.stat-value {{ font-size:28px; font-weight:800; margin:6px 0; color:var(--fg) }}
.stat-trend {{ font-size:11px; font-weight:600 }}

.stat-mini {{
  background:rgba(30,41,59,0.5);
  border-radius:10px;
  padding:12px 14px;
}}
.stat-mini-label {{ color:var(--muted); font-size:11px; font-weight:600; margin-bottom:4px }}
.stat-mini-value {{ font-size:22px; font-weight:800 }}

.charts-row {{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:16px;
}}
@media (max-width:768px) {{
  .charts-row {{ grid-template-columns:1fr }}
}}

.chart-card {{ min-height:340px }}

.rank-list {{ display:flex; flex-direction:column; gap:6px }}
.rank-row {{
  display:grid;
  grid-template-columns:36px 1fr auto auto;
  gap:10px;
  align-items:center;
  padding:8px 10px;
  background:rgba(30,41,59,0.4);
  border-radius:8px;
  transition:background .15s;
}}
.rank-row:hover {{ background:rgba(99,102,241,0.15) }}
.rank-medal {{ font-size:20px; text-align:center }}
.rank-name {{ font-weight:600; font-size:13px }}
.rank-stat {{ color:var(--muted); font-size:11px }}
.rank-income {{ color:var(--green); font-weight:700; font-size:13px }}

.stock-bars {{ display:flex; flex-direction:column; gap:8px }}
.stock-bar-row {{
  display:grid;
  grid-template-columns:24px 1fr auto;
  gap:10px;
  align-items:center;
  padding:8px;
  background:rgba(30,41,59,0.3);
  border-radius:6px;
}}
.stock-num {{ font-weight:800; font-size:18px }}
</style>

<script>
// Chart.js global config
Chart.defaults.color = '#94a3b8';
Chart.defaults.borderColor = 'rgba(148,163,184,0.1)';
Chart.defaults.font.family = 'system-ui, -apple-system, sans-serif';

// 1. Daromad chart
new Chart(document.getElementById('incomeChart'), {{
  type: 'line',
  data: {{
    labels: {daily_labels},
    datasets: [{{
      label: 'Daromad (so\'m)',
      data: {daily_incomes},
      borderColor: '#6366f1',
      backgroundColor: 'rgba(99,102,241,0.15)',
      borderWidth: 2,
      tension: 0.35,
      fill: true,
      pointRadius: 3,
      pointHoverRadius: 6,
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ beginAtZero: true, ticks: {{ callback: v => (v/1000).toFixed(0)+'k' }} }},
      x: {{ ticks: {{ maxRotation: 0, autoSkip: true, maxTicksLimit: 10 }} }},
    }},
  }},
}});

// 2. Ish turlari pie chart
new Chart(document.getElementById('typesChart'), {{
  type: 'doughnut',
  data: {{
    labels: {types_labels},
    datasets: [{{
      data: {types_counts},
      backgroundColor: {types_colors},
      borderWidth: 0,
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{
      legend: {{ position: 'right', labels: {{ font: {{ size: 11 }}, padding: 8 }} }},
    }},
  }},
}});

// 3. Ombor stock chart
new Chart(document.getElementById('stockChart'), {{
  type: 'bar',
  data: {{
    labels: ['Kritik', 'Sariq', 'Yetarli'],
    datasets: [{{
      data: [{stock_red}, {stock_yellow}, {stock_green}],
      backgroundColor: ['#ef4444', '#f59e0b', '#10b981'],
      borderRadius: 6,
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    indexAxis: 'y',
    plugins: {{ legend: {{ display: false }} }},
    scales: {{ x: {{ beginAtZero: true }} }},
  }},
}});
</script>
"""
    return web.Response(text=_base("Bosh sahifa", "dashboard", content), content_type="text/html")




@_require_auth
async def production_analytics(request: web.Request):
    """Ishlab chiqarish chuqur tahlili — qadamlar bo'yicha, vaqt bo'yicha."""
    period   = request.query.get("period", "month")  # week | month | quarter | year
    today    = date.today()

    if period == "week":
        start_date  = today - timedelta(days=7)
        period_name = "Oxirgi 7 kun"
    elif period == "quarter":
        start_date  = today - timedelta(days=90)
        period_name = "Oxirgi 3 oy"
    elif period == "year":
        start_date  = today - timedelta(days=365)
        period_name = "Oxirgi yil"
    else:
        start_date  = today.replace(day=1)
        period_name = "Joriy oy"

    async with AsyncSessionLocal() as db:
        # Ishlab chiqarish zanjiri bo'yicha
        chain_steps = [
            ("gofra_ishlab",    "📐 Gofra ishlab",   "#3b82f6"),
            ("laminatsiya",     "✨ Laminatsiya",    "#8b5cf6"),
            ("zagatovka",       "📦 Zagatovka",      "#ec4899"),
            ("gofra_kiley",     "🔨 Gofra kley",     "#f59e0b"),
            ("tiger_kesish",    "✂️ Tiger kesish",   "#10b981"),
            ("list_qogoz",      "📄 List kesish",    "#06b6d4"),
            ("stepler_tikish",  "📌 Stepler",        "#84cc16"),
            ("yopishtirma",     "🔗 Yopishtirma",    "#f97316"),
            ("rulon_orash",     "🌀 Rulon o'rash",   "#a855f7"),
            ("rulonga_salafan", "🎁 Rulonga salafan","#d946ef"),
            ("adyol_tikish",    "🧵 Adyol tikish",   "#22c55e"),
            ("diplomat_tikish", "💼 Pastel tikish",  "#0ea5e9"),
            ("adyol_qoqish",    "🛏 Adyol qoqish",   "#14b8a6"),
            ("pastel_qoqish",   "📫 Pastel qoqish",  "#fb7185"),
        ]

        chain_data = []
        for wt_key, wt_label, color in chain_steps:
            try:
                wt_enum = WorkType(wt_key)
            except ValueError:
                continue
            r = await db.execute(
                select(
                    func.count(WorkEntry.id),
                    func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                    func.coalesce(func.sum(WorkEntry.soni), 0),
                ).where(
                    WorkEntry.work_type == wt_enum,
                    WorkEntry.work_date >= start_date,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            row = r.one()
            chain_data.append({
                "key":    wt_key,
                "label":  wt_label,
                "color":  color,
                "count":  int(row[0] or 0),
                "income": float(row[1] or 0),
                "soni":   float(row[2] or 0),
            })

        # Kunlik trend
        days_count = (today - start_date).days + 1
        if days_count > 60:
            # Haftalik agregatsiya
            trend_label = "Haftalik"
            buckets = []
            current = start_date
            while current <= today:
                week_end = min(current + timedelta(days=6), today)
                r = await db.execute(
                    select(
                        func.count(WorkEntry.id),
                        func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                    ).where(
                        WorkEntry.work_date >= current,
                        WorkEntry.work_date <= week_end,
                        WorkEntry.status == WorkStatus.approved,
                    )
                )
                row = r.one()
                buckets.append({
                    "label":  current.strftime("%d.%m"),
                    "count":  int(row[0] or 0),
                    "income": float(row[1] or 0),
                })
                current = week_end + timedelta(days=1)
            trend_data = buckets
        else:
            trend_label = "Kunlik"
            trend_data = []
            current = start_date
            while current <= today:
                r = await db.execute(
                    select(
                        func.count(WorkEntry.id),
                        func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                    ).where(
                        WorkEntry.work_date == current,
                        WorkEntry.status == WorkStatus.approved,
                    )
                )
                row = r.one()
                trend_data.append({
                    "label":  current.strftime("%d.%m"),
                    "count":  int(row[0] or 0),
                    "income": float(row[1] or 0),
                })
                current += timedelta(days=1)

        # Kvalitet (sifat) — rad etilganlar foizi
        r_qa = await db.execute(
            select(
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            ).where(
                WorkEntry.work_date >= start_date,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            )
        )
        qa_row     = r_qa.one()
        total_qa   = int(qa_row[0] or 0)
        rejected_n = int(qa_row[1] or 0)
        qa_rate    = ((total_qa - rejected_n) / total_qa * 100) if total_qa > 0 else 100

        # Inspektorlar reytingi
        r_insp = await db.execute(
            select(
                User.full_name,
                func.count(WorkEntry.id),
            )
            .join(WorkEntry, WorkEntry.inspector_id == User.id)
            .where(
                WorkEntry.work_date >= start_date,
                WorkEntry.finished_at.is_not(None),
            )
            .group_by(User.id, User.full_name)
            .order_by(func.count(WorkEntry.id).desc())
            .limit(5)
        )
        inspectors = [{"name": r[0], "count": r[1]} for r in r_insp.all()]

    import json
    chain_labels = json.dumps([d["label"] for d in chain_data])
    chain_counts = json.dumps([d["count"] for d in chain_data])
    chain_colors = json.dumps([d["color"] for d in chain_data])

    trend_labels  = json.dumps([d["label"]  for d in trend_data])
    trend_counts  = json.dumps([d["count"]  for d in trend_data])
    trend_incomes = json.dumps([d["income"] for d in trend_data])

    # Chain HTML cards
    chain_html = ""
    max_count = max((d["count"] for d in chain_data), default=1)
    for d in chain_data:
        pct = (d["count"] / max_count * 100) if max_count > 0 else 0
        chain_html += f"""
        <div class="chain-card" style="border-left:4px solid {d['color']}">
          <div class="chain-label">{d['label']}</div>
          <div class="chain-count">{d['count']}</div>
          <div class="chain-bar"><div class="chain-bar-fill" style="width:{pct:.0f}%;background:{d['color']}"></div></div>
          <div class="chain-meta">{fmt(d['income'])} so'm | {d['soni']:.0f} dona</div>
        </div>
        """

    insp_html = ""
    for i, ins in enumerate(inspectors):
        insp_html += f"""
        <div class="insp-row">
          <span class="insp-num">{i+1}</span>
          <span class="insp-name">{h(ins['name'])}</span>
          <span class="insp-count">{ins['count']} ish</span>
        </div>
        """
    if not insp_html:
        insp_html = '<p style="color:var(--muted);padding:16px;text-align:center">Tekshiruvchilar yo\'q</p>'

    content = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>

<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;flex-wrap:wrap;gap:12px">
  <div>
    <h1 style="margin-bottom:2px">🏭 Ishlab chiqarish tahlili</h1>
    <p style="color:var(--muted);margin:0">{period_name} | {start_date.strftime('%d.%m.%Y')} — {today.strftime('%d.%m.%Y')}</p>
  </div>
  <div class="period-switcher">
    <a href="?period=week"    class="btn-period {'active' if period == 'week' else ''}">Hafta</a>
    <a href="?period=month"   class="btn-period {'active' if period == 'month' else ''}">Oy</a>
    <a href="?period=quarter" class="btn-period {'active' if period == 'quarter' else ''}">3 oy</a>
    <a href="?period=year"    class="btn-period {'active' if period == 'year' else ''}">Yil</a>
  </div>
</div>

<!-- SIFAT KO'RSATKICHI -->
<div class="card" style="margin-bottom:20px">
  <h2>✅ Sifat ko'rsatkichi</h2>
  <div style="display:flex;align-items:center;gap:20px;margin-top:10px;flex-wrap:wrap">
    <div style="flex:0 0 auto">
      <div style="font-size:42px;font-weight:800;color:{('var(--green)' if qa_rate >= 95 else ('#f59e0b' if qa_rate >= 85 else 'var(--red)'))}">{qa_rate:.1f}%</div>
      <div style="color:var(--muted);font-size:12px">qabul qilingan ishlar</div>
    </div>
    <div style="flex:1 1 200px;min-width:200px">
      <div style="background:rgba(30,41,59,0.5);border-radius:12px;height:24px;overflow:hidden">
        <div style="background:linear-gradient(90deg,var(--green) 0%, #10b981 100%);height:100%;width:{qa_rate:.1f}%;transition:width .5s"></div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:6px;font-size:11px;color:var(--muted)">
        <span>{total_qa - rejected_n} qabul qilindi</span>
        <span>{rejected_n} rad etildi</span>
        <span>jami {total_qa}</span>
      </div>
    </div>
  </div>
</div>

<!-- ISHLAB CHIQARISH ZANJIRI -->
<div class="card" style="margin-bottom:20px">
  <h2>🔗 Ishlab chiqarish zanjiri</h2>
  <p style="color:var(--muted);font-size:12px;margin-bottom:14px">Har bosqichdagi ishlar soni va daromad</p>
  <div class="chain-grid">{chain_html}</div>
</div>

<!-- DIAGRAMMALAR -->
<div class="charts-row">
  <div class="card chart-card">
    <h2>📊 Bosqichlar bo'yicha hajmlar</h2>
    <div style="height:300px;position:relative;margin-top:10px"><canvas id="chainChart"></canvas></div>
  </div>
  <div class="card chart-card">
    <h2>📈 {trend_label} trend</h2>
    <div style="height:300px;position:relative;margin-top:10px"><canvas id="trendChart"></canvas></div>
  </div>
</div>

<!-- INSPEKTORLAR -->
<div class="card" style="margin-top:20px">
  <h2>👁 Top tekshiruvchilar</h2>
  <div class="insp-list">{insp_html}</div>
</div>

<style>
.period-switcher {{ display:flex; gap:4px; background:rgba(30,41,59,0.5); padding:4px; border-radius:10px }}
.btn-period {{
  padding:6px 14px; border-radius:6px; font-size:13px; font-weight:600;
  color:var(--muted); text-decoration:none; transition:all .15s;
}}
.btn-period:hover {{ background:rgba(99,102,241,0.1); color:var(--fg) }}
.btn-period.active {{ background:#6366f1; color:#fff }}

.chain-grid {{
  display:grid;
  grid-template-columns:repeat(auto-fill,minmax(180px,1fr));
  gap:10px;
}}
.chain-card {{
  background:rgba(30,41,59,0.5);
  border-radius:10px;
  padding:12px;
  border-left:4px solid #6366f1;
}}
.chain-label {{ font-size:12px; color:var(--muted); font-weight:600; margin-bottom:4px }}
.chain-count {{ font-size:24px; font-weight:800; margin-bottom:6px }}
.chain-bar {{ background:rgba(15,23,42,0.6); border-radius:4px; height:6px; overflow:hidden; margin-bottom:6px }}
.chain-bar-fill {{ height:100%; border-radius:4px; transition:width .5s }}
.chain-meta {{ font-size:10px; color:var(--muted) }}

.insp-list {{ display:flex; flex-direction:column; gap:6px }}
.insp-row {{
  display:grid;
  grid-template-columns:30px 1fr auto;
  gap:10px;
  align-items:center;
  padding:8px 12px;
  background:rgba(30,41,59,0.4);
  border-radius:8px;
}}
.insp-num {{
  background:#6366f1; color:#fff; border-radius:50%;
  width:24px; height:24px; display:flex; align-items:center;
  justify-content:center; font-size:11px; font-weight:700;
}}
.insp-name {{ font-weight:600; font-size:13px }}
.insp-count {{ color:var(--muted); font-size:12px; font-weight:600 }}
</style>

<script>
Chart.defaults.color = '#94a3b8';
Chart.defaults.borderColor = 'rgba(148,163,184,0.1)';

new Chart(document.getElementById('chainChart'), {{
  type: 'bar',
  data: {{
    labels: {chain_labels},
    datasets: [{{
      label: 'Ishlar soni',
      data: {chain_counts},
      backgroundColor: {chain_colors},
      borderRadius: 4,
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    indexAxis: 'y',
    plugins: {{ legend: {{ display: false }} }},
  }},
}});

new Chart(document.getElementById('trendChart'), {{
  type: 'line',
  data: {{
    labels: {trend_labels},
    datasets: [
      {{
        label: 'Ishlar soni',
        data: {trend_counts},
        borderColor: '#6366f1',
        backgroundColor: 'rgba(99,102,241,0.1)',
        yAxisID: 'y',
        tension: 0.35,
      }},
      {{
        label: 'Daromad (k so\'m)',
        data: {trend_incomes}.map(v => v/1000),
        borderColor: '#10b981',
        backgroundColor: 'rgba(16,185,129,0.1)',
        yAxisID: 'y1',
        tension: 0.35,
      }},
    ],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    interaction: {{ mode: 'index', intersect: false }},
    scales: {{
      y:  {{ type: 'linear', position: 'left',  beginAtZero: true }},
      y1: {{ type: 'linear', position: 'right', beginAtZero: true, grid: {{ drawOnChartArea: false }} }},
    }},
  }},
}});
</script>
"""
    return web.Response(text=_base("Ishlab chiqarish tahlili", "production", content), content_type="text/html")






@_require_auth
async def inventory_health(request: web.Request):
    """Ombor salomatligi — predictive alerts."""
    today    = date.today()
    week_ago = today - timedelta(days=7)

    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(WarehouseProduct)
            .where(WarehouseProduct.is_active == True)
            .order_by(WarehouseProduct.miqdor.asc())
        )
        all_products = r.scalars().all()

        consumption = []
        for p in all_products:
            r_c = await db.execute(
                select(func.coalesce(func.sum(func.abs(WarehouseLog.delta)), 0))
                .where(
                    WarehouseLog.product_id == p.id,
                    WarehouseLog.delta < 0,
                    WarehouseLog.created_at >= week_ago,
                )
            )
            week_use  = float(r_c.scalar() or 0)
            daily_use = week_use / 7 if week_use > 0 else 0
            days_left = (float(p.miqdor) / daily_use) if daily_use > 0 else 999

            status = "ok"
            if days_left < 3:    status = "critical"
            elif days_left < 7:  status = "warning"
            elif days_left < 14: status = "caution"

            consumption.append({
                "p": p, "week_use": week_use, "daily_use": daily_use,
                "days_left": days_left, "status": status,
            })

        consumption.sort(key=lambda x: x["days_left"])
        critical = [d for d in consumption if d["status"] == "critical"]
        warning  = [d for d in consumption if d["status"] == "warning"]
        caution  = [d for d in consumption if d["status"] == "caution"]
        ok_count = len(consumption) - len(critical) - len(warning) - len(caution)

    # HTML build (string concat — f-string emas)
    parts = []
    parts.append('<h1 style="margin-bottom:4px">🏥 Ombor salomatligi</h1>')
    parts.append('<p style="color:var(--muted);margin-bottom:20px">Predictive alerts — qaysi mahsulot qachon tugashi mumkin</p>')

    # Top cards
    parts.append('<div class="stats-grid" style="margin-bottom:20px">')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--red)"><div class="stat-label">🚨 Kritik</div><div class="stat-value" style="color:var(--red)">{len(critical)}</div><div class="stat-trend">3 kun ichida tugaydi</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid #f59e0b"><div class="stat-label">⚠️ Ogohlantirish</div><div class="stat-value" style="color:#f59e0b">{len(warning)}</div><div class="stat-trend">7 kun ichida tugaydi</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid #8b5cf6"><div class="stat-label">📅 Diqqat</div><div class="stat-value" style="color:#8b5cf6">{len(caution)}</div><div class="stat-trend">14 kun ichida tugaydi</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--green)"><div class="stat-label">✅ Yetarli</div><div class="stat-value" style="color:var(--green)">{ok_count}</div><div class="stat-trend">Hammasi OK</div></div>')
    parts.append('</div>')

    # Critical alerts
    parts.append('<div class="card" style="margin-bottom:16px">')
    parts.append('<h2 style="color:var(--red)">🚨 Kritik — darhol e\'tibor kerak</h2>')
    parts.append('<p style="color:var(--muted);font-size:12px;margin-bottom:10px">Bu mahsulotlar 3 kun ichida tugashi mumkin</p>')
    parts.append('<div class="alerts-list">')

    if not critical:
        parts.append('<p style="color:var(--muted);padding:12px;text-align:center">🎉 Tanqidiy holatda mahsulot yo\'q</p>')
    else:
        for d in critical[:15]:
            p    = d["p"]
            name = h(p.name)
            if p.razmer: name += " | " + h(p.razmer)
            if p.rang:   name += " | " + h(p.rang)
            days_txt = "TUGADI" if p.miqdor <= 0 else f"{d['days_left']:.1f} kun"
            parts.append('<div class="alert-row alert-critical">')
            parts.append('<div class="alert-icon">🚨</div>')
            parts.append('<div class="alert-content">')
            parts.append(f'<div class="alert-name">{name}</div>')
            parts.append(f'<div class="alert-meta">Qoldi: <b>{p.miqdor:.0f}</b> {p.birlik or "dona"} | Kunlik istemol: {d["daily_use"]:.1f}</div>')
            parts.append('</div>')
            parts.append(f'<div class="alert-days">{days_txt}</div>')
            parts.append('</div>')

    parts.append('</div></div>')

    # Warning alerts
    parts.append('<div class="card" style="margin-bottom:16px">')
    parts.append('<h2 style="color:#f59e0b">⚠️ Ogohlantirish — yetkazib berish kerak</h2>')
    parts.append('<div class="alerts-list">')

    if not warning:
        parts.append('<p style="color:var(--muted);padding:12px;text-align:center">Ogohlantirish yo\'q</p>')
    else:
        for d in warning[:15]:
            p    = d["p"]
            name = h(p.name)
            if p.razmer: name += " | " + h(p.razmer)
            parts.append('<div class="alert-row alert-warning">')
            parts.append('<div class="alert-icon">⚠️</div>')
            parts.append(f'<div class="alert-content"><div class="alert-name">{name}</div><div class="alert-meta">Qoldi: <b>{p.miqdor:.0f}</b> | {d["daily_use"]:.1f}/kun</div></div>')
            parts.append(f'<div class="alert-days">{d["days_left"]:.1f} kun</div>')
            parts.append('</div>')

    parts.append('</div></div>')

    # CSS
    parts.append('<style>')
    parts.append('.alerts-list { display:flex; flex-direction:column; gap:6px }')
    parts.append('.alert-row { display:grid; grid-template-columns:36px 1fr auto; gap:10px; align-items:center; padding:10px 12px; border-radius:8px; background:rgba(30,41,59,0.4); border-left:4px solid #6366f1 }')
    parts.append('.alert-critical { border-left-color:var(--red); background:rgba(239,68,68,0.08) }')
    parts.append('.alert-warning { border-left-color:#f59e0b; background:rgba(245,158,11,0.08) }')
    parts.append('.alert-icon { font-size:22px; text-align:center }')
    parts.append('.alert-name { font-weight:600; font-size:13px; margin-bottom:2px }')
    parts.append('.alert-meta { color:var(--muted); font-size:11px }')
    parts.append('.alert-days { font-weight:800; font-size:14px; padding:4px 10px; background:rgba(15,23,42,0.6); border-radius:6px; white-space:nowrap }')
    parts.append('</style>')

    content = "\n".join(parts)
    return web.Response(text=_base("Ombor salomatligi", "health", content), content_type="text/html")




@_require_auth
async def activity_feed(request: web.Request):
    """Real-time amallar oqimi — kim, qachon, nima qildi."""
    today = date.today()
    
    async with AsyncSessionLocal() as db:
        # Bugungi ish kirimlari
        r_works = await db.execute(
            select(WorkEntry, User)
            .join(User, User.id == WorkEntry.worker_id)
            .where(WorkEntry.work_date == today)
            .order_by(WorkEntry.created_at.desc())
            .limit(50)
        )
        works = r_works.all()
        
        # Bugungi ombor harakatlari
        r_logs = await db.execute(
            select(WarehouseLog, WarehouseProduct)
            .join(WarehouseProduct, WarehouseProduct.id == WarehouseLog.product_id)
            .where(WarehouseLog.created_at >= datetime.combine(today, datetime.min.time()))
            .order_by(WarehouseLog.created_at.desc())
            .limit(30)
        )
        logs = r_logs.all()

    parts = []
    parts.append('<h1 style="margin-bottom:4px">📡 Real-time aktivlik</h1>')
    parts.append(f'<p style="color:var(--muted);margin-bottom:20px">Bugun: {today.strftime("%d.%m.%Y")} | {len(works)} ish kirimi | {len(logs)} ombor amali</p>')
    
    parts.append('<div class="charts-row">')
    
    # Ishlar oqimi
    parts.append('<div class="card">')
    parts.append('<h2>📋 Bugungi ishlar oqimi</h2>')
    parts.append('<div class="feed-list">')
    
    if not works:
        parts.append('<p style="color:var(--muted);padding:20px;text-align:center">Bugun ish yo\'q</p>')
    else:
        for we, u in works:
            status_icons = {"approved": "✅", "rejected": "❌", "pending": "⏳"}
            icon = status_icons.get(we.status.value, "•")
            time_str = we.created_at.strftime("%H:%M") if we.created_at else "—"
            wt_label = we.work_type.value.replace("_", " ").title() if we.work_type else "?"
            
            parts.append('<div class="feed-item">')
            parts.append(f'<div class="feed-time">{time_str}</div>')
            parts.append(f'<div class="feed-icon">{icon}</div>')
            parts.append('<div class="feed-content">')
            parts.append(f'<div class="feed-title"><a href="/web/workers/{u.id}" style="color:#4F46E5;text-decoration:none">{h(u.full_name or "?")}</a></div>')
            parts.append(f'<div class="feed-desc">{wt_label} — {we.soni:.0f} dona</div>')
            parts.append('</div>')
            parts.append(f'<div class="feed-value">{fmt(we.jami_summa)}</div>')
            parts.append('</div>')
    
    parts.append('</div></div>')
    
    # Ombor oqimi
    parts.append('<div class="card">')
    parts.append('<h2>📦 Ombor harakati</h2>')
    parts.append('<div class="feed-list">')
    
    if not logs:
        parts.append('<p style="color:var(--muted);padding:20px;text-align:center">Bugun ombor amali yo\'q</p>')
    else:
        for log, p in logs:
            time_str = log.created_at.strftime("%H:%M") if log.created_at else "—"
            delta = float(log.delta or 0)
            
            if delta > 0:
                icon = "📥"
                color = "var(--green)"
                delta_str = f"+{delta:.0f}"
            else:
                icon = "📤"
                color = "var(--red)"
                delta_str = f"{delta:.0f}"
            
            name = h(p.name)
            if p.razmer: name += f" | {h(p.razmer)}"
            
            parts.append('<div class="feed-item">')
            parts.append(f'<div class="feed-time">{time_str}</div>')
            parts.append(f'<div class="feed-icon">{icon}</div>')
            parts.append('<div class="feed-content">')
            parts.append(f'<div class="feed-title">{name}</div>')
            parts.append(f'<div class="feed-desc">{h(log.amal or "")}</div>')
            parts.append('</div>')
            parts.append(f'<div class="feed-value" style="color:{color};font-weight:800">{delta_str}</div>')
            parts.append('</div>')
    
    parts.append('</div></div>')
    parts.append('</div>')
    
    # CSS
    parts.append('<style>')
    parts.append('.feed-list { display:flex; flex-direction:column; gap:4px; max-height:600px; overflow-y:auto }')
    parts.append('.feed-item { display:grid; grid-template-columns:50px 30px 1fr auto; gap:10px; align-items:center; padding:8px 10px; background:rgba(30,41,59,0.4); border-radius:8px; transition:background .15s }')
    parts.append('.feed-item:hover { background:rgba(99,102,241,0.1) }')
    parts.append('.feed-time { color:var(--muted); font-size:11px; font-family:ui-monospace,monospace }')
    parts.append('.feed-icon { font-size:18px; text-align:center }')
    parts.append('.feed-title { font-weight:600; font-size:13px }')
    parts.append('.feed-desc { color:var(--muted); font-size:11px; margin-top:1px }')
    parts.append('.feed-value { font-weight:700; font-size:13px; white-space:nowrap }')
    parts.append('</style>')
    
    # Auto-refresh skript
    parts.append('<script>setTimeout(()=>location.reload(), 30000);</script>')
    
    content = "\n".join(parts)
    return web.Response(text=_base("Real-time aktivlik", "activity", content), content_type="text/html")




@_require_auth
async def quality_dashboard(request: web.Request):
    """Sifat tahlili — rad etish stavkalari, problem ishchilar."""
    today = date.today()
    month_start = today.replace(day=1)
    
    async with AsyncSessionLocal() as db:
        # Umumiy sifat
        r = await db.execute(
            select(
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.approved, 1), else_=0)),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            ).where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            )
        )
        row = r.one()
        total    = int(row[0] or 0)
        approved = int(row[1] or 0)
        rejected = int(row[2] or 0)
        accept_rate = (approved / total * 100) if total > 0 else 100
        
        # Ishchilar bo'yicha
        r_w = await db.execute(
            select(
                User.id, User.full_name,
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            )
            .join(WorkEntry, WorkEntry.worker_id == User.id)
            .where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            )
            .group_by(User.id, User.full_name)
            .having(func.count(WorkEntry.id) > 0)
        )
        workers_qa = []
        for r2 in r_w.all():
            w_total = int(r2[2] or 0)
            w_rej   = int(r2[3] or 0)
            rate    = ((w_total - w_rej) / w_total * 100) if w_total > 0 else 0
            workers_qa.append({
                "id": r2[0], "name": r2[1], 
                "total": w_total, "rejected": w_rej, "rate": rate,
            })
        # Eng yomon - eng yuqori rad etish foizi
        workers_qa.sort(key=lambda x: x["rate"])
        
        # Ish turlari bo'yicha
        r_t = await db.execute(
            select(
                WorkEntry.work_type,
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            ).where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            ).group_by(WorkEntry.work_type)
        )
        types_qa = []
        for r3 in r_t.all():
            t_total = int(r3[1] or 0)
            t_rej   = int(r3[2] or 0)
            rate    = ((t_total - t_rej) / t_total * 100) if t_total > 0 else 0
            types_qa.append({
                "type":  r3[0].value if r3[0] else "?",
                "total": t_total, "rejected": t_rej, "rate": rate,
            })
        types_qa.sort(key=lambda x: x["rate"])
        
        # Eng ko'p ishlatilgan rad etish sabablari (oxirgi 30 kun)
        from sqlalchemy import distinct
        r_reasons = await db.execute(
            select(
                WorkEntry.izoh,
                func.count(WorkEntry.id),
            ).where(
                WorkEntry.work_date >= month_start,
                WorkEntry.status == WorkStatus.rejected,
                WorkEntry.izoh.is_not(None),
                WorkEntry.izoh != "",
            ).group_by(WorkEntry.izoh)
            .order_by(func.count(WorkEntry.id).desc())
            .limit(10)
        )
        top_reasons = [(r[0] or "—", r[1]) for r in r_reasons.all()]

    parts = []
    parts.append('<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>')
    parts.append('<h1 style="margin-bottom:4px">✅ Sifat boshqaruvi</h1>')
    parts.append(f'<p style="color:var(--muted);margin-bottom:20px">Bu oyda: {total} ish, {rejected} ta rad etilgan</p>')
    
    # Asosiy ko'rsatkichlar
    qa_color = "var(--green)" if accept_rate >= 95 else ("#f59e0b" if accept_rate >= 85 else "var(--red)")
    parts.append('<div class="stats-grid" style="margin-bottom:20px">')
    parts.append(f'<div class="stat-card" style="border-left:4px solid {qa_color}"><div class="stat-label">✅ Sifat darajasi</div><div class="stat-value" style="color:{qa_color}">{accept_rate:.1f}%</div><div class="stat-trend">qabul qilingan</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--green)"><div class="stat-label">👍 Qabul qilindi</div><div class="stat-value">{approved}</div><div class="stat-trend">ish</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--red)"><div class="stat-label">👎 Rad etildi</div><div class="stat-value" style="color:var(--red)">{rejected}</div><div class="stat-trend">ish</div></div>')
    parts.append(f'<div class="stat-card"><div class="stat-label">📊 Jami</div><div class="stat-value">{total}</div><div class="stat-trend">tekshirilgan</div></div>')
    parts.append('</div>')
    
    # Sifat darajasi bar
    parts.append('<div class="card" style="margin-bottom:16px">')
    parts.append('<h3 style="margin-bottom:10px">Sifat darajasi</h3>')
    parts.append('<div style="background:rgba(30,41,59,0.5);border-radius:12px;height:30px;overflow:hidden;position:relative">')
    parts.append(f'<div style="background:linear-gradient(90deg,{qa_color},var(--green));height:100%;width:{accept_rate:.1f}%;display:flex;align-items:center;justify-content:flex-end;padding-right:12px;color:#fff;font-weight:700">{accept_rate:.1f}%</div>')
    parts.append('</div>')
    parts.append('</div>')
    
    # Worker QA jadval
    parts.append('<div class="charts-row">')
    parts.append('<div class="card">')
    parts.append('<h2>👥 Ishchilar sifati</h2>')
    parts.append('<p style="color:var(--muted);font-size:12px;margin-bottom:10px">Eng pastdan yuqoriga</p>')
    parts.append('<div class="qa-list">')
    
    if not workers_qa:
        parts.append('<p style="color:var(--muted);padding:20px;text-align:center">Ma\'lumot yo\'q</p>')
    else:
        for w in workers_qa[:15]:
            rate_color = "var(--green)" if w["rate"] >= 95 else ("#f59e0b" if w["rate"] >= 85 else "var(--red)")
            parts.append('<div class="qa-row">')
            parts.append(f'<div class="qa-name"><a href="/web/workers/{w["id"]}" style="color:inherit;text-decoration:none">{h(w["name"] or "—")}</a></div>')
            parts.append(f'<div class="qa-bar"><div class="qa-bar-fill" style="width:{w["rate"]:.0f}%;background:{rate_color}"></div></div>')
            parts.append(f'<div class="qa-stats">{w["total"] - w["rejected"]}/{w["total"]}</div>')
            parts.append(f'<div class="qa-rate" style="color:{rate_color}">{w["rate"]:.0f}%</div>')
            parts.append('</div>')
    
    parts.append('</div></div>')
    
    # Ish turlari QA
    parts.append('<div class="card">')
    parts.append('<h2>🔧 Ish turlari bo\'yicha</h2>')
    parts.append('<div class="qa-list">')
    
    if not types_qa:
        parts.append('<p style="color:var(--muted);padding:20px;text-align:center">Ma\'lumot yo\'q</p>')
    else:
        for t in types_qa:
            rate_color = "var(--green)" if t["rate"] >= 95 else ("#f59e0b" if t["rate"] >= 85 else "var(--red)")
            t_label = t["type"].replace("_", " ").title()
            parts.append('<div class="qa-row">')
            parts.append(f'<div class="qa-name">{h(t_label)}</div>')
            parts.append(f'<div class="qa-bar"><div class="qa-bar-fill" style="width:{t["rate"]:.0f}%;background:{rate_color}"></div></div>')
            parts.append(f'<div class="qa-stats">{t["total"] - t["rejected"]}/{t["total"]}</div>')
            parts.append(f'<div class="qa-rate" style="color:{rate_color}">{t["rate"]:.0f}%</div>')
            parts.append('</div>')
    
    parts.append('</div></div>')
    parts.append('</div>')
    
    # Top rad etish sabablari
    if top_reasons:
        parts.append('<div class="card" style="margin-top:20px">')
        parts.append('<h2>📝 Eng ko\'p rad etish sabablari</h2>')
        parts.append('<div class="reasons-list">')
        max_count = max(c for _, c in top_reasons) if top_reasons else 1
        for reason, cnt in top_reasons:
            pct = (cnt / max_count * 100) if max_count > 0 else 0
            parts.append('<div class="reason-row">')
            parts.append(f'<div class="reason-text">{h(reason[:80])}</div>')
            parts.append('<div class="reason-bar-wrap">')
            parts.append(f'<div class="reason-bar" style="width:{pct:.0f}%"></div>')
            parts.append(f'<span class="reason-count">{cnt}</span>')
            parts.append('</div>')
            parts.append('</div>')
        parts.append('</div></div>')
    
    # CSS
    parts.append('<style>')
    parts.append('.qa-list { display:flex; flex-direction:column; gap:6px }')
    parts.append('.qa-row { display:grid; grid-template-columns:1fr 100px 60px 50px; gap:10px; align-items:center; padding:8px 10px; background:rgba(30,41,59,0.4); border-radius:8px }')
    parts.append('.qa-name { font-weight:600; font-size:13px }')
    parts.append('.qa-bar { background:rgba(15,23,42,0.6); border-radius:4px; height:8px; overflow:hidden }')
    parts.append('.qa-bar-fill { height:100%; border-radius:4px; transition:width .5s }')
    parts.append('.qa-stats { color:var(--muted); font-size:11px; text-align:right }')
    parts.append('.qa-rate { font-weight:800; font-size:14px; text-align:right }')
    parts.append('.reasons-list { display:flex; flex-direction:column; gap:6px; margin-top:10px }')
    parts.append('.reason-row { display:grid; grid-template-columns:1fr 200px; gap:14px; align-items:center; padding:8px 10px; background:rgba(30,41,59,0.4); border-radius:8px }')
    parts.append('.reason-text { font-size:13px }')
    parts.append('.reason-bar-wrap { display:flex; align-items:center; gap:10px; position:relative }')
    parts.append('.reason-bar { height:8px; background:linear-gradient(90deg,#ef4444,#f59e0b); border-radius:4px; flex:0 0 auto; max-width:170px }')
    parts.append('.reason-count { color:var(--muted); font-size:11px; font-weight:700 }')
    parts.append('</style>')
    
    content = "\n".join(parts)
    return web.Response(text=_base("Sifat boshqaruvi", "quality", content), content_type="text/html")


@_require_auth
async def warehouse(request: web.Request):
    cat_filter = request.rel_url.query.get("cat", "")
    query = request.rel_url.query.get("q", "")
    holat = request.rel_url.query.get("holat", "")
    page = int(request.rel_url.query.get("page", 1))
    per_page = 30

    async with AsyncSessionLocal() as db:
        cat_enum = None
        if cat_filter:
            try:
                cat_enum = ProductCategory(cat_filter)
            except ValueError:
                pass
        products, total = await search_products(
            db, query=query or None, category=cat_enum,
            holat=holat or None, limit=per_page, offset=(page - 1) * per_page
        )

    # Category tabs
    cat_tabs = '<a href="/web/warehouse" class="cat-tab' + (' active' if not cat_filter else '') + '">🌐 Barchasi</a>'
    for cv, cn in CAT_NAMES.items():
        active = " active" if cat_filter == cv else ""
        params = f"?cat={cv}" + (f"&q={h(query)}" if query else "")
        cat_tabs += f'<a href="/web/warehouse{params}" class="cat-tab{active}">{h(cn)}</a>'

    # Products table
    rows = ""
    for p in products:
        icon = stock_icon(p)
        cls = stock_cls(p)
        razmer = h(p.razmer or "—")
        rang = h(p.rang or "—")
        tur = h(p.tur or "—")
        rows += f"""<tr>
<td><strong>{icon} {h(p.name)}</strong><br><span class="t-xs t-muted">{CAT_NAMES.get(p.category.value if hasattr(p.category,'value') else str(p.category), str(p.category))}</span></td>
<td class="t-xs t-muted">{razmer}</td>
<td class="t-xs t-muted">{rang}</td>
<td class="t-xs t-muted">{tur}</td>
<td><span class="{cls} fw7 mono">{p.miqdor} {h(p.birlik)}</span></td>
<td class="t-xs t-muted">{p.min_threshold} / {p.yellow_threshold}</td>
<td>
  <button class="btn btn-g btn-xs" onclick="openKirim({p.id}, '{h(p.name)}', '{h(p.birlik)}')">+Kirim</button>
  <button class="btn btn-d btn-xs" onclick="openChiqim({p.id}, '{h(p.name)}', {p.miqdor}, '{h(p.birlik)}')">-Chiqim</button>
  <button class="btn btn-cy btn-xs" onclick="openThresh({p.id}, '{h(p.name)}', {p.min_threshold}, {p.yellow_threshold})">⚙️</button>
  <button class="btn btn-xs {"btn-g" if getattr(p,"alert_enabled",True) else "bgr"}" style='{"" if getattr(p,"alert_enabled",True) else "opacity:0.55"}' title='{"Xabar yoqilgan — o&#39;chirish uchun bosing" if getattr(p,"alert_enabled",True) else "Xabar o&#39;chirilgan — yoqish uchun bosing"}' onclick="toggleAlert({p.id}, this)">{"🔔" if getattr(p,"alert_enabled",True) else "🔕"}</button>
  <a href="/web/warehouse/delete-confirm/{p.id}" class="btn btn-d btn-xs" title="O'chirish" onclick="return confirm('Ushbu mahsulotni butunlay ochirmoqchimisiz?')">🗑</a>
</td>
</tr>"""

    if not rows:
        rows = '<tr><td colspan="7" class="empty-state">📭 Mahsulot topilmadi</td></tr>'

    # Pagination
    total_pages = max(1, (total + per_page - 1) // per_page)
    pag = ""
    if total_pages > 1:
        base_url = f"/web/warehouse?cat={cat_filter}&q={query}&holat={holat}"
        pag = '<div class="pagination">'
        for i in range(1, total_pages + 1):
            active = " act" if i == page else ""
            pag += f'<a href="{base_url}&page={i}" class="{active}">{i}</a>'
        pag += f'<span class="pinfo">{total} ta mahsulot</span></div>'

    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">📦 Ombor boshqaruvi</span>
  <div style="display:flex;gap:7px">
    <button class="btn btn-p btn-sm" onclick="openModal('modal-add')">➕ Yangi mahsulot</button>
    <a href="/web/warehouse/export" class="btn btn-cy btn-sm">📥 Excel</a>
    <a href="/web/warehouse/logs" class="btn btn-s btn-sm">📋 Tarix</a>
  </div>
</div>
<div class="cat-tabs">{cat_tabs}</div>
<div class="toolbar">
  <div class="toolbar-left">
    <form method="GET" style="display:flex;gap:6px;flex-wrap:wrap">
      <input type="hidden" name="cat" value="{h(cat_filter)}">
      <div class="search-wrap">
        <span class="si">🔍</span>
        <input type="text" name="q" value="{h(query)}" placeholder="Qidirish...">
      </div>
      <select name="holat" class="fsel">
        <option value="">Barcha holat</option>
        <option value="kam" {'selected' if holat=='kam' else ''}>🔴 Kam qolgan</option>
        <option value="cheklangan" {'selected' if holat=='cheklangan' else ''}>🟡 Cheklangan</option>
        <option value="yetarli" {'selected' if holat=='yetarli' else ''}>🟢 Yetarli</option>
      </select>
      <button type="submit" class="btn btn-s btn-sm">Filtr</button>
    </form>
  </div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Mahsulot</th><th>Razmer</th><th>Rang</th><th>Tur</th>
        <th>Qoldiq</th><th>Min/Sariq</th><th>Amallar</th>
      </tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
  {pag}
</div>

<!-- Modal: Yangi mahsulot -->
<div class="overlay" id="modal-add" onclick="if(event.target===this)closeModal('modal-add')">
  <div class="modal">
    <div class="modal-title">➕ Yangi mahsulot qo'shish</div>
    <div class="modal-sub">Ombor bazasiga yangi mahsulot qo'shing</div>
    <form method="POST" action="/web/warehouse/add">
      <div class="form-row">
        <div class="fg">
          <label class="fl">Kategoriya *</label>
          <select name="category" required>
            {''.join(f'<option value="{cv}">{cn}</option>' for cv, cn in CAT_NAMES.items())}
          </select>
        </div>
        <div class="fg">
          <label class="fl">Nomi *</label>
          <input type="text" name="name" required placeholder="Mahsulot nomi">
        </div>
      </div>
      <div class="form-row">
        <div class="fg">
          <label class="fl">Razmer</label>
          <input type="text" name="razmer" placeholder="masalan: 600x600">
        </div>
        <div class="fg">
          <label class="fl">Rang</label>
          <input type="text" name="rang" placeholder="masalan: Oq">
        </div>
      </div>
      <div class="form-row">
        <div class="fg">
          <label class="fl">Tur</label>
          <input type="text" name="tur" placeholder="masalan: tiger_uchun">
        </div>
        <div class="fg">
          <label class="fl">Birlik</label>
          <select name="birlik">
            <option value="dona">dona</option>
            <option value="kg">kg</option>
            <option value="m">m</option>
            <option value="m²">m²</option>
            <option value="rulon">rulon</option>
          </select>
        </div>
      </div>
      <div class="form-row">
        <div class="fg">
          <label class="fl">Boshlang'ich miqdor</label>
          <input type="number" name="miqdor" value="0" min="0" step="0.01">
        </div>
        <div class="fg">
          <label class="fl">Gramaj (g/m²)</label>
          <input type="number" name="qalinlik" step="0.1" placeholder="ixtiyoriy">
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-add')">Bekor</button>
        <button type="submit" class="btn btn-p">✅ Saqlash</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal: Kirim -->
<div class="overlay" id="modal-kirim" onclick="if(event.target===this)closeModal('modal-kirim')">
  <div class="modal">
    <div class="modal-title">📥 Kirim</div>
    <div class="modal-sub" id="kirim-sub">Mahsulot nomi</div>
    <form method="POST" action="/web/warehouse/kirim">
      <input type="hidden" name="product_id" id="kirim-pid">
      <div class="fg">
        <label class="fl">Miqdor *</label>
        <input type="number" name="miqdor" id="kirim-miqdor" required min="0.001" step="0.001">
        <div class="input-hint" id="kirim-birlik"></div>
      </div>
      <div class="fg">
        <label class="fl">Izoh</label>
        <input type="text" name="izoh" placeholder="ixtiyoriy">
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-kirim')">Bekor</button>
        <button type="submit" class="btn btn-g">📥 Kirim</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal: Chiqim -->
<div class="overlay" id="modal-chiqim" onclick="if(event.target===this)closeModal('modal-chiqim')">
  <div class="modal">
    <div class="modal-title">📤 Chiqim</div>
    <div class="modal-sub" id="chiqim-sub">Mahsulot nomi</div>
    <form method="POST" action="/web/warehouse/chiqim">
      <input type="hidden" name="product_id" id="chiqim-pid">
      <div class="fg">
        <label class="fl">Miqdor *</label>
        <input type="number" name="miqdor" id="chiqim-miqdor" required min="0.001" step="0.001">
        <div class="input-hint" id="chiqim-birlik"></div>
      </div>
      <div class="fg">
        <label class="fl">Izoh</label>
        <input type="text" name="izoh" placeholder="ixtiyoriy">
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-chiqim')">Bekor</button>
        <button type="submit" class="btn btn-d">📤 Chiqim</button>
      </div>
    </form>
  </div>
</div>

<!-- Modal: Chegara sozlash -->
<div class="overlay" id="modal-thresh" onclick="if(event.target===this)closeModal('modal-thresh')">
  <div class="modal">
    <div class="modal-title">⚙️ Chegara sozlash</div>
    <div class="modal-sub" id="thresh-sub"></div>
    <form method="POST" action="/web/warehouse/thresholds">
      <input type="hidden" name="product_id" id="thresh-pid">
      <div class="form-row">
        <div class="fg">
          <label class="fl">🔴 Minimal chegara</label>
          <input type="number" name="min_threshold" id="thresh-min" min="0" step="0.1">
        </div>
        <div class="fg">
          <label class="fl">🟡 Sariq chegara</label>
          <input type="number" name="yellow_threshold" id="thresh-yellow" min="0" step="0.1">
        </div>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-thresh')">Bekor</button>
        <button type="submit" class="btn btn-w">⚙️ Saqlash</button>
      </div>
    </form>
  </div>
</div>
"""
    js = """
function deleteProduct(id) {
  if (!confirm("Ushbu mahsulotni butunlay ochirmoqchimisiz? Bu amalni qaytarib bolmaydi!")) return;
  fetch('/web/warehouse/delete', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({product_id: id})
  })
  .then(function(r) { return r.json(); })
  .then(function(d) {
    if (d.ok) { location.reload(); }
    else { alert('Xato: ' + (d.error || 'nomalum')); }
  })
  .catch(function(e) { alert('Tarmoq xatosi: ' + e); });
}


function toggleAlert(productId, btn) {
  var isOn = btn.textContent.trim() === '🔔';
  var newVal = isOn ? 0 : 1;
  fetch('/web/warehouse/alert-toggle', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({product_id: productId, alert_enabled: newVal})
  })
  .then(function(r) { return r.json(); })
  .then(function(d) {
    if (d.ok) {
      if (newVal === 1) {
        btn.textContent = '🔔';
        btn.className = btn.className.replace('bgr', 'btn-g');
        btn.style.opacity = '';
        btn.title = "Xabar yoqilgan \u2014 o\u02bbchirish uchun bosing";
      } else {
        btn.textContent = '🔕';
        btn.className = btn.className.replace('btn-g', 'bgr');
        btn.style.opacity = '0.55';
        btn.title = "Xabar o\u02bbchirilgan \u2014 yoqish uchun bosing";
      }
    } else {
      alert('Xato: ' + (d.error || 'nomalum'));
    }
  })
  .catch(function(e) { alert('Tarmoq xatosi: ' + e); });
}

function openKirim(id, name, birlik) {
  document.getElementById('kirim-pid').value = id;
  document.getElementById('kirim-sub').textContent = name;
  document.getElementById('kirim-birlik').textContent = 'Birlik: ' + birlik;
  openModal('modal-kirim');
}
function openChiqim(id, name, miq, birlik) {
  document.getElementById('chiqim-pid').value = id;
  document.getElementById('chiqim-sub').textContent = name + ' (joriy: ' + miq + ' ' + birlik + ')';
  document.getElementById('chiqim-birlik').textContent = 'Maksimal: ' + miq + ' ' + birlik;
  document.getElementById('chiqim-miqdor').max = miq;
  openModal('modal-chiqim');
}
function openThresh(id, name, mn, yr) {
  document.getElementById('thresh-pid').value = id;
  document.getElementById('thresh-sub').textContent = name;
  document.getElementById('thresh-min').value = mn;
  document.getElementById('thresh-yellow').value = yr;
  openModal('modal-thresh');
}
"""
    return web.Response(text=_base("📦 Ombor", "warehouse", content, js), content_type="text/html")

@_require_auth
async def warehouse_add(request: web.Request):
    data = await request.post()
    async with AsyncSessionLocal() as db:
        try:
            cat = ProductCategory(data.get("category", "rulon"))
        except ValueError:
            cat = ProductCategory.rulon
        try:
            miqdor = float(str(data.get("miqdor", "0")).replace(",", "."))
        except Exception:
            miqdor = 0.0
        qalinlik = None
        try:
            q = data.get("qalinlik", "").strip()
            if q:
                qalinlik = float(q.replace(",", "."))
        except Exception:
            pass
        birlik = data.get("birlik", "dona")
        product = WarehouseProduct(
            category=cat,
            name=data.get("name", "Noma'lum"),
            razmer=_normalize_razmer(data.get("razmer") or None),
            rang=data.get("rang") or None,
            tur=data.get("tur") or None,
            qalinlik=qalinlik,
            birlik=birlik,
            miqdor=miqdor,
        )
        db.add(product)
        await db.commit()
    raise web.HTTPFound("/web/warehouse")

@_require_auth


@_require_auth
async def warehouse_delete(request: web.Request):
    """Mahsulotni butunlay o'chirish (is_active=False)."""
    ct = request.content_type or ""
    try:
        if "application/json" in ct:
            data = await request.json()
            pid  = int(data.get("product_id", 0))
        else:
            form = await request.post()
            pid  = int(form.get("product_id", 0))
    except Exception:
        return web.json_response({"ok": False, "error": "Noto'g'ri ma'lumot"})

    if pid <= 0:
        return web.json_response({"ok": False, "error": "ID noto'g'ri"})

    async with AsyncSessionLocal() as db:
        product = await db.get(WarehouseProduct, pid)
        if not product:
            return web.json_response({"ok": False, "error": "Topilmadi"})
        product.is_active = False
        await db.commit()
    return web.json_response({"ok": True, "deleted": True})




# ═══ PDF DOWNLOADS ═══════════════════════════════════════════════════════════

@_require_auth
async def pdf_daily(request: web.Request):
    """Kunlik PDF hisobot."""
    from utils.pdf_reports import generate_daily_report_pdf
    d_str = request.query.get("date", "")
    try:
        d = datetime.strptime(d_str, "%Y-%m-%d").date() if d_str else date.today()
    except ValueError:
        d = date.today()
    async with AsyncSessionLocal() as db:
        pdf_bytes = await generate_daily_report_pdf(db, d)
    return web.Response(
        body=pdf_bytes,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="kunlik_{d.strftime("%Y-%m-%d")}.pdf"',
        },
    )


@_require_auth
async def pdf_monthly(request: web.Request):
    """Oylik PDF hisobot."""
    from utils.pdf_reports import generate_monthly_report_pdf
    oy_str = request.query.get("oy", "")
    try:
        if oy_str:
            year, month = map(int, oy_str.split("-")[:2])
        else:
            today = date.today()
            year, month = today.year, today.month
    except (ValueError, IndexError):
        today = date.today()
        year, month = today.year, today.month

    async with AsyncSessionLocal() as db:
        pdf_bytes = await generate_monthly_report_pdf(db, year, month)
    return web.Response(
        body=pdf_bytes,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="oylik_{year}-{month:02d}.pdf"',
        },
    )


@_require_auth
async def pdf_worker(request: web.Request):
    """Ishchi PDF hisobot."""
    from utils.pdf_reports import generate_worker_report_pdf
    try:
        wid = int(request.match_info.get("worker_id", "0"))
    except ValueError:
        wid = 0
    oy_str = request.query.get("oy", "")
    try:
        if oy_str:
            year, month = map(int, oy_str.split("-")[:2])
        else:
            today = date.today()
            year, month = today.year, today.month
    except (ValueError, IndexError):
        today = date.today()
        year, month = today.year, today.month

    if wid <= 0:
        raise web.HTTPFound("/web/workers")

    async with AsyncSessionLocal() as db:
        try:
            pdf_bytes = await generate_worker_report_pdf(db, wid, year, month)
        except ValueError:
            raise web.HTTPFound("/web/workers")

    return web.Response(
        body=pdf_bytes,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="ishchi_{wid}_{year}-{month:02d}.pdf"',
        },
    )


@_require_auth
async def pdf_warehouse(request: web.Request):
    """Ombor PDF hisobot."""
    from utils.pdf_reports import generate_warehouse_report_pdf
    async with AsyncSessionLocal() as db:
        pdf_bytes = await generate_warehouse_report_pdf(db)
    return web.Response(
        body=pdf_bytes,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="ombor_{date.today().strftime("%Y-%m-%d")}.pdf"',
        },
    )




# ═══ SYSTEM HEALTH ENDPOINTS ════════════════════════════════════════════════

async def health_check_route(request: web.Request):
    """Railway health check — JSON."""
    try:
        async with AsyncSessionLocal() as db:
            from utils.health_monitor import get_system_stats
            stats = await get_system_stats(db)
        return web.json_response({"status": "ok", **stats})
    except Exception as e:
        return web.json_response({"status": "error", "error": str(e)}, status=500)


@_require_auth
async def web_system_health(request: web.Request):
    """Tizim salomatligi — vizual sahifa."""
    async with AsyncSessionLocal() as db:
        from utils.health_monitor import get_system_stats, format_uptime
        stats = await get_system_stats(db)

    db_stats = stats.get("database", {})
    cache_stats = stats.get("cache", {})
    sys_stats = stats.get("system", {})
    uptime = format_uptime(stats["uptime_sec"])

    db_status = db_stats.get("status", "?")
    db_color = "var(--green)" if db_status == "ok" else ("#f59e0b" if db_status == "slow" else "var(--red)")
    db_emoji = "🟢" if db_status == "ok" else ("🟡" if db_status == "slow" else "🔴")

    parts = []
    parts.append('<h1 style="margin-bottom:4px">🩺 Tizim salomatligi</h1>')
    parts.append(f'<p style="color:var(--muted);margin-bottom:20px">Yangilangan: {datetime.now().strftime("%H:%M:%S")}</p>')

    # KPI
    parts.append('<div class="stats-grid" style="margin-bottom:20px">')
    parts.append(f'<div class="stat-card" style="border-left:4px solid {db_color}"><div class="stat-label">{db_emoji} Database</div><div class="stat-value">{db_stats.get("response_ms", 0)} ms</div><div class="stat-trend">{db_status}</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--green)"><div class="stat-label">⏱ Uptime</div><div class="stat-value" style="font-size:18px">{uptime}</div><div class="stat-trend">Bot ishlamoqda</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid #8b5cf6"><div class="stat-label">💾 Cache</div><div class="stat-value">{cache_stats.get("size", 0)}</div><div class="stat-trend">{cache_stats.get("hit_rate", "0%")} hit</div></div>')
    mem = sys_stats.get("memory_mb", 0)
    parts.append(f'<div class="stat-card" style="border-left:4px solid #06b6d4"><div class="stat-label">💻 Xotira</div><div class="stat-value">{mem} MB</div><div class="stat-trend">{sys_stats.get("cpu_time", "—")}</div></div>')
    parts.append('</div>')

    # Database detail
    parts.append('<div class="card" style="margin-bottom:16px">')
    parts.append('<h2>📊 Database statistikasi</h2>')
    parts.append('<table style="width:100%;margin-top:10px">')
    parts.append(f'<tr><td>👥 Foydalanuvchilar</td><td style="text-align:right"><b>{db_stats.get("users", 0)}</b></td></tr>')
    parts.append(f'<tr><td>📝 Ish yozuvlari</td><td style="text-align:right"><b>{db_stats.get("work_entries", 0)}</b></td></tr>')
    parts.append(f'<tr><td>📦 Mahsulotlar</td><td style="text-align:right"><b>{db_stats.get("products", 0)}</b></td></tr>')
    parts.append(f'<tr><td>⏱ Javob vaqti</td><td style="text-align:right"><b>{db_stats.get("response_ms", 0)} ms</b></td></tr>')
    parts.append('</table>')
    parts.append('</div>')

    # Cache detail
    parts.append('<div class="card">')
    parts.append('<h2>💾 Cache statistikasi</h2>')
    parts.append('<table style="width:100%;margin-top:10px">')
    parts.append(f'<tr><td>📦 Hozirgi hajm</td><td style="text-align:right"><b>{cache_stats.get("size", 0)}</b> yozuv</td></tr>')
    parts.append(f'<tr><td>✅ Topildi (hits)</td><td style="text-align:right;color:var(--green)"><b>{cache_stats.get("hits", 0)}</b></td></tr>')
    parts.append(f'<tr><td>❌ Topilmadi (misses)</td><td style="text-align:right;color:var(--red)"><b>{cache_stats.get("misses", 0)}</b></td></tr>')
    parts.append(f'<tr><td>📊 Hit rate</td><td style="text-align:right"><b>{cache_stats.get("hit_rate", "0%")}</b></td></tr>')
    parts.append('</table>')
    parts.append('</div>')

    parts.append('<script>setTimeout(()=>location.reload(), 30000);</script>')

    content = "\n".join(parts)
    return web.Response(text=_base("Tizim salomatligi", "health", content), content_type="text/html")




# ═══ QR KOD ENDPOINTS ════════════════════════════════════════════════════════

async def product_qr(request: web.Request):
    """Mahsulot QR kod — ommaviy ko'rinish (auth shart emas)."""
    try:
        pid = int(request.match_info.get("product_id", "0"))
    except ValueError:
        pid = 0

    async with AsyncSessionLocal() as db:
        p = await db.get(WarehouseProduct, pid)
        if not p:
            return web.Response(text="Mahsulot topilmadi", status=404)

        # So'nggi 5 ta operatsiya
        r = await db.execute(
            select(WarehouseLog)
            .where(WarehouseLog.product_id == pid)
            .order_by(WarehouseLog.created_at.desc())
            .limit(5)
        )
        logs = r.scalars().all()

    # Holat
    miqdor = float(p.miqdor or 0)
    min_th = float(p.min_threshold or 0)
    status_color = "#10b981" if miqdor > min_th else "#ef4444"
    status_text = "Yetarli" if miqdor > min_th else "Kam qoldi!"

    info_rows = [
        ("📁 Kategoriya", p.category.value if p.category else "—"),
        ("📐 Razmer",     p.razmer or "—"),
        ("🎨 Rang",       p.rang or "—"),
        ("📦 Tur",        p.tur or "—"),
        ("📐 Razmer tur", p.razmer_tur or "—"),
        ("📊 Birlik",     p.birlik or "dona"),
        ("⚠️ Min kerak",   f"{min_th:.0f}"),
    ]
    info_html = ""
    for label, val in info_rows:
        if val and val != "—":
            info_html += f'<tr><td style="color:#94a3b8;padding:6px 8px">{label}</td><td style="padding:6px 8px"><b>{val}</b></td></tr>'

    logs_html = ""
    for log in logs:
        amal = log.amal_turi.value if log.amal_turi else "?"
        d_str = log.created_at.strftime("%d.%m %H:%M") if log.created_at else "—"
        sign = "+" if (log.delta or 0) > 0 else ""
        col = "#10b981" if (log.delta or 0) > 0 else "#ef4444"
        logs_html += (
            f'<div style="padding:8px;border-left:3px solid {col};background:#1e293b;border-radius:6px;margin-bottom:6px">'
            f'<div style="font-size:12px;color:#94a3b8">{d_str} — {amal}</div>'
            f'<div style="font-weight:600;color:{col}">{sign}{log.delta:.0f} {p.birlik or "dona"}</div>'
            f'</div>'
        )

    html = f"""<!DOCTYPE html><html><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{p.name} — Quti Tsexi</title>
<style>
body {{
  font-family: system-ui, -apple-system, sans-serif;
  background: #0f172a; color: #e2e8f0;
  margin: 0; padding: 16px; min-height: 100vh;
}}
.container {{ max-width: 500px; margin: 0 auto }}
.card {{
  background: #1e293b; border-radius: 16px; padding: 20px;
  margin-bottom: 14px; border: 1px solid #334155;
}}
h1 {{ margin: 0 0 8px 0; font-size: 24px; color: #f1f5f9 }}
.qty {{
  font-size: 56px; font-weight: 800;
  color: {status_color}; text-align: center;
  margin: 14px 0;
}}
.status {{
  text-align: center;
  padding: 8px;
  background: {status_color}22;
  color: {status_color};
  border-radius: 12px;
  font-weight: 700;
}}
table {{ width: 100%; border-collapse: collapse }}
td {{ border-bottom: 1px solid #334155 }}
tr:last-child td {{ border-bottom: none }}
h2 {{ font-size: 16px; margin: 16px 0 10px; color: #4F46E5 }}
.footer {{ text-align: center; padding: 12px; color: #64748b; font-size: 12px }}
</style></head><body>
<div class="container">
  <div class="card">
    <h1>{p.name}</h1>
    <div class="qty">{miqdor:.0f}</div>
    <div style="text-align:center;color:#94a3b8;margin-bottom:8px">{p.birlik or 'dona'}</div>
    <div class="status">{status_text}</div>
  </div>

  <div class="card">
    <h2>📋 Ma'lumotlar</h2>
    <table>{info_html}</table>
  </div>

  {('<div class="card"><h2>📊 So\'nggi operatsiyalar</h2>' + logs_html + '</div>') if logs_html else ''}

  <div class="footer">
    🏭 Quti Tsexi — ID: {p.id}<br>
    {datetime.now().strftime('%d.%m.%Y %H:%M')}
  </div>
</div>
</body></html>"""

    return web.Response(text=html, content_type="text/html")




@_require_auth
async def worker_toggle(request: web.Request):
    """Ishchini faollashtirish/bloklash."""
    try:
        wid = int(request.match_info.get("id", "0"))
    except ValueError:
        wid = 0
    async with AsyncSessionLocal() as db:
        w = await db.get(User, wid)
        if w:
            w.is_active = not w.is_active
            await db.commit()
    raise web.HTTPFound("/web/workers")


@_require_auth
async def worker_delete(request: web.Request):
    """Ishchini o'chirish (is_active=False)."""
    try:
        wid = int(request.match_info.get("id", "0"))
    except ValueError:
        wid = 0
    async with AsyncSessionLocal() as db:
        w = await db.get(User, wid)
        if w:
            w.is_active = False
            await db.commit()
    raise web.HTTPFound("/web/workers")




@_require_auth
async def zero_stock_page(request: web.Request):
    """Tugagan mahsulotlar — saqlash yoki o'chirish."""
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(WarehouseProduct)
            .where(
                WarehouseProduct.is_active == True,
                WarehouseProduct.miqdor <= 0,
            )
            .order_by(WarehouseProduct.name)
        )
        products = r.scalars().all()

    parts = []
    parts.append('<h1 style="margin-bottom:4px">⚠️ Tugagan mahsulotlar</h1>')
    parts.append('<p style="color:var(--muted);margin-bottom:20px">0 ga tushgan mahsulotlar — saqlash yoki butunlay o\'chirish</p>')

    if not products:
        parts.append('<div class="card" style="text-align:center;padding:40px">')
        parts.append('<h2 style="color:var(--green)">🎉 Tugagan mahsulot yo\'q!</h2>')
        parts.append('<p style="color:var(--muted)">Ombor holati yaxshi.</p>')
        parts.append('</div>')
    else:
        parts.append('<div class="card" style="background:rgba(245,158,11,0.1);border-left:4px solid #f59e0b;margin-bottom:16px">')
        parts.append('<b>⚠️ {} ta mahsulot tugagan!</b><br>Saqlash yoki o\'chirish kerakligini hal qiling.'.format(len(products)))
        parts.append('</div>')

        for p in products:
            name = h(p.name)
            cat_name = p.category.value if p.category else ""
            tur = (" — " + h(p.tur)) if p.tur else ""

            badges = []
            if p.razmer:     badges.append('<span style="background:rgba(99,102,241,.15);color:#4F46E5;padding:2px 8px;border-radius:8px;font-size:11px">' + h(p.razmer) + '</span>')
            if p.razmer_tur: badges.append('<span style="background:rgba(139,92,246,.15);color:#c4b5fd;padding:2px 8px;border-radius:8px;font-size:11px">' + h(p.razmer_tur) + '</span>')
            if p.rang:       badges.append('<span style="background:rgba(236,72,153,.15);color:#f9a8d4;padding:2px 8px;border-radius:8px;font-size:11px">🎨 ' + h(p.rang) + '</span>')

            row = '<div class="card" style="margin-bottom:10px;border-left:4px solid var(--red)">'
            row += '<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px">'
            row += '<div>'
            row += '<div style="font-weight:700;font-size:15px;margin-bottom:4px">📦 ' + name + '</div>'
            row += '<div style="font-size:12px;color:var(--muted);margin-bottom:6px">' + cat_name + tur + '</div>'
            row += '<div style="display:flex;gap:6px;flex-wrap:wrap">' + " ".join(badges) + '</div>'
            row += '</div>'
            row += '<div style="display:flex;gap:8px;flex-wrap:wrap">'
            row += '<button class="btn" style="background:#10b981;color:#fff;padding:8px 14px" onclick="zeroKeep(' + str(p.id) + ', \'' + name + '\')">✅ Saqlash</button>'
            row += '<button class="btn" style="background:#ef4444;color:#fff;padding:8px 14px" onclick="zeroDelete(' + str(p.id) + ', \'' + name + '\')">❌ O\'chirish</button>'
            row += '</div></div></div>'
            parts.append(row)

    js = """
<script>
function zeroKeep(id, nom) {
  if (!confirm('"' + nom + '" saqlansinmi?')) return;
  fetch('/web/warehouse/zero-keep', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({product_id: id})
  }).then(r => r.json()).then(d => {
    if (d.ok) location.reload(); else alert('Xato');
  });
}
function zeroDelete(id, nom) {
  if (!confirm('"' + nom + '" butunlay o\'chirilsinmi?')) return;
  fetch('/web/warehouse/delete', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({product_id: id})
  }).then(r => r.json()).then(d => {
    if (d.ok) location.reload(); else alert('Xato');
  });
}
</script>
"""
    parts.append(js)

    content = "\n".join(parts)
    return web.Response(text=_base("Tugagan mahsulotlar", "zero-stock", content), content_type="text/html")


@_require_auth
async def warehouse_zero_keep(request: web.Request):
    """Mahsulotni saqlash (zero_notified = False)."""
    try:
        data = await request.json()
        pid = int(data.get("product_id", 0))
    except Exception:
        return web.json_response({"ok": False, "error": "Xato"})

    async with AsyncSessionLocal() as db:
        p = await db.get(WarehouseProduct, pid)
        if not p:
            return web.json_response({"ok": False, "error": "Topilmadi"})
        p.zero_notified = False
        await db.commit()
    return web.json_response({"ok": True})




@_require_auth
async def warehouse_delete_confirm(request: web.Request):
    """Mahsulotni o'chirish (GET link orqali) — keyin warehouse ga qaytadi."""
    try:
        pid = int(request.match_info.get("id", "0"))
    except ValueError:
        pid = 0
    cat = request.query.get("cat", "")
    async with AsyncSessionLocal() as db:
        p = await db.get(WarehouseProduct, pid)
        if p:
            p.is_active = False
            await db.commit()
    # Qaysi kategoriyaga qaytish
    if cat:
        raise web.HTTPFound(f"/web/warehouse?cat={cat}")
    raise web.HTTPFound("/web/warehouse")




@_require_auth
async def warehouse_kirim(request: web.Request):
    """JSON yoki form data — ikkalasini ham qabul qiladi."""
    ct = request.content_type or ""
    try:
        if "application/json" in ct:
            data = await request.json()
            pid    = int(data.get("product_id", 0))
            miqdor = float(str(data.get("miqdor", "0")).replace(",", "."))
            izoh   = data.get("izoh") or "Web panel — kirim"
        else:
            form   = await request.post()
            pid    = int(form.get("product_id", 0))
            miqdor = float(str(form.get("miqdor", "0")).replace(",", "."))
            izoh   = form.get("izoh") or "Web panel — kirim"
    except Exception:
        return web.json_response({"ok": False, "error": "Noto'g'ri ma'lumot"})

    if pid > 0 and miqdor > 0:
        async with AsyncSessionLocal() as db:
            product = await db.get(WarehouseProduct, pid)
            if not product:
                return web.json_response({"ok": False, "error": "Mahsulot topilmadi"})
            await update_product_miqdor(db, pid, miqdor, None, izoh=izoh)
            await db.commit()
            new_miq = float(product.miqdor) + miqdor
            if "application/json" in ct:
                return web.json_response({"ok": True, "new_miqdor": round(new_miq, 1)})
    if "application/json" in ct:
        return web.json_response({"ok": False, "error": "Miqdor 0 dan katta bo'lishi kerak"})
    raise web.HTTPFound("/web/warehouse")

@_require_auth
async def warehouse_chiqim(request: web.Request):
    """JSON yoki form data — ikkalasini ham qabul qiladi."""
    ct = request.content_type or ""
    try:
        if "application/json" in ct:
            data   = await request.json()
            pid    = int(data.get("product_id", 0))
            miqdor = float(str(data.get("miqdor", "0")).replace(",", "."))
            izoh   = data.get("izoh") or "Web panel — chiqim"
        else:
            form   = await request.post()
            pid    = int(form.get("product_id", 0))
            miqdor = float(str(form.get("miqdor", "0")).replace(",", "."))
            izoh   = form.get("izoh") or "Web panel — chiqim"
    except Exception:
        return web.json_response({"ok": False, "error": "Noto'g'ri ma'lumot"})

    if pid > 0 and miqdor > 0:
        async with AsyncSessionLocal() as db:
            product = await db.get(WarehouseProduct, pid)
            if not product:
                return web.json_response({"ok": False, "error": "Mahsulot topilmadi"})
            cur = float(product.miqdor)
            if cur < miqdor:
                return web.json_response({"ok": False, "error": f"Yetarli emas: {cur:.1f} ta bor"})
            await update_product_miqdor(db, pid, -miqdor, None, izoh=izoh)
            await db.commit()
            new_miq = cur - miqdor
            if "application/json" in ct:
                return web.json_response({"ok": True, "new_miqdor": round(new_miq, 1)})
    if "application/json" in ct:
        return web.json_response({"ok": False, "error": "Miqdor 0 dan katta bo'lishi kerak"})
    raise web.HTTPFound("/web/warehouse")

async def warehouse_thresholds(request: web.Request):
    data = await request.post()
    pid = int(data.get("product_id", 0))
    try:
        mn = float(str(data.get("min_threshold", "2")).replace(",", "."))
        yr = float(str(data.get("yellow_threshold", "5")).replace(",", "."))
    except Exception:
        mn, yr = 2.0, 5.0
    if pid > 0:
        async with AsyncSessionLocal() as db:
            product = await get_product_by_id(db, pid)
            if product:
                product.min_threshold = mn
                product.yellow_threshold = yr
                await db.commit()
    raise web.HTTPFound("/web/warehouse")


@_require_auth
async def warehouse_alert_toggle(request: web.Request):
    """Mahsulot uchun bildirishnoma yoq/o'chir."""
    try:
        data = await request.json()
        pid           = int(data.get("product_id", 0))
        alert_enabled = bool(int(data.get("alert_enabled", 1)))
    except Exception as e:
        return web.json_response({"ok": False, "error": f"Noto'g'ri ma'lumot: {e}"})

    if pid <= 0:
        return web.json_response({"ok": False, "error": "ID noto'g'ri"})

    async with AsyncSessionLocal() as db:
        product = await db.get(WarehouseProduct, pid)
        if not product:
            return web.json_response({"ok": False, "error": "Mahsulot topilmadi"})
        product.alert_enabled = alert_enabled
        await db.commit()

    status = "yoqildi 🔔" if alert_enabled else "o'chirildi 🔕"
    return web.json_response({
        "ok": True,
        "product_id": pid,
        "alert_enabled": alert_enabled,
        "message": f"Bildirishnoma {status}"
    })


async def warehouse_export(request: web.Request):
    async with AsyncSessionLocal() as db:
        products = await get_all_products(db)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ombor qoldig'i"
    headers = ["#", "Kategoriya", "Nomi", "Razmer", "Rang", "Tur", "Birlik", "Miqdor", "Min", "Sariq", "Holat"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for i, p in enumerate(products, 1):
        holat = "🔴 Kam" if float(p.miqdor) <= float(p.min_threshold) else "🟡 Cheklangan" if float(p.miqdor) <= float(p.yellow_threshold) else "🟢 Yetarli"
        ws.append([i, p.category.value, p.name, p.razmer or "", p.rang or "", p.tur or "",
                   p.birlik, float(p.miqdor), float(p.min_threshold), float(p.yellow_threshold), holat])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    today = date.today().strftime("%Y%m%d")
    return web.Response(
        body=out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ombor_{today}.xlsx"'},
    )

@_require_auth
async def warehouse_logs(request: web.Request):
    page = int(request.rel_url.query.get("page", 1))
    amal_f = request.rel_url.query.get("amal", "")
    per_page = 50
    async with AsyncSessionLocal() as db:
        logs, total = await get_warehouse_logs_paged(
            db, amal=amal_f or None, limit=per_page, offset=(page - 1) * per_page
        )
        rows = ""
        for lg in logs:
            product = await get_product_by_id(db, lg.product_id)
            pname = h(product.name) if product else f"id={lg.product_id}"
            amal_badge = '<span class="badge bg">📥 Kirim</span>' if lg.amal == "kirim" else '<span class="badge br">📤 Chiqim</span>'
            ts = lg.created_at.strftime("%d.%m.%Y %H:%M") if lg.created_at else "—"
            rows += f"""<tr>
<td class="t-xs t-muted">{ts}</td>
<td><strong>{pname}</strong></td>
<td>{amal_badge}</td>
<td class="td-n">{lg.miqdor}</td>
<td class="td-n t-muted">{lg.oldin} → {lg.keyin}</td>
<td class="t-xs t-muted">{h(lg.izoh or '—')}</td>
</tr>"""
    if not rows:
        rows = '<tr><td colspan="6" class="empty-state">📭 Tarix bo\'sh</td></tr>'
    total_pages = max(1, (total + per_page - 1) // per_page)
    pag = ""
    if total_pages > 1:
        pag = '<div class="pagination">'
        for i in range(1, min(total_pages + 1, 20)):
            active = " act" if i == page else ""
            pag += f'<a href="/web/warehouse/logs?page={i}&amal={amal_f}" class="{active}">{i}</a>'
        pag += f'<span class="pinfo">{total} ta yozuv</span></div>'
    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">📋 Kirim/Chiqim tarixi</span>
  <div style="display:flex;gap:7px">
    <a href="?amal=" class="btn btn-s btn-sm {'btn-p' if not amal_f else ''}">Barchasi</a>
    <a href="?amal=kirim" class="btn btn-s btn-sm {'btn-g' if amal_f=='kirim' else ''}">📥 Kirim</a>
    <a href="?amal=chiqim" class="btn btn-s btn-sm {'btn-d' if amal_f=='chiqim' else ''}">📤 Chiqim</a>
  </div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Vaqt</th><th>Mahsulot</th><th>Amal</th><th>Miqdor</th><th>Oldin → Keyin</th><th>Izoh</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
  {pag}
</div>
"""
    return web.Response(text=_base("📋 Kirim/Chiqim tarixi", "ombor-logs", content), content_type="text/html")

# ─── WORKERS ───────────────────────────────────────────────────────────────

@_require_auth
async def workers(request: web.Request):
    """Ishchilar ro'yxati — avatarlar, statistika, qidiruv, saralash."""
    search = (request.query.get("q") or "").strip().lower()
    sort   = request.query.get("sort", "income")  # income | works | name | recent
    today  = date.today()
    month_start = today.replace(day=1)

    async with AsyncSessionLocal() as db:
        all_users = await get_users_by_role(db, UserRole.ishchi)

        # Har bir ishchi uchun statistika
        worker_stats = []
        for u in all_users:
            if search and search not in (u.full_name or "").lower():
                continue

            # Oylik daromad
            r = await db.execute(
                select(
                    func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                    func.count(WorkEntry.id),
                ).where(
                    WorkEntry.worker_id == u.id,
                    WorkEntry.work_date >= month_start,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            row = r.one()
            income = float(row[0] or 0)
            works  = int(row[1] or 0)

            # So'nggi ish
            r_last = await db.execute(
                select(WorkEntry.created_at)
                .where(WorkEntry.worker_id == u.id)
                .order_by(WorkEntry.created_at.desc())
                .limit(1)
            )
            last_work = r_last.scalar_one_or_none()

            # Faol smena
            from database.models import WorkSession
            r_sess = await db.execute(
                select(WorkSession)
                .where(WorkSession.worker_id == u.id, WorkSession.closed_at.is_(None))
                .limit(1)
            )
            active_session = r_sess.scalar_one_or_none() is not None

            worker_stats.append({
                "u": u, "income": income, "works": works,
                "last_work": last_work, "active": active_session,
            })

        # Saralash
        if sort == "income":
            worker_stats.sort(key=lambda x: x["income"], reverse=True)
        elif sort == "works":
            worker_stats.sort(key=lambda x: x["works"], reverse=True)
        elif sort == "name":
            worker_stats.sort(key=lambda x: x["u"].full_name or "")
        elif sort == "recent":
            worker_stats.sort(key=lambda x: x["last_work"] or datetime.min, reverse=True)

        active_count = sum(1 for w in worker_stats if w["active"])
        total_income = sum(w["income"] for w in worker_stats)
        total_works  = sum(w["works"]  for w in worker_stats)

    # HTML build
    parts = []
    parts.append('<h1 style="margin-bottom:4px">👥 Ishchilar</h1>')
    parts.append(f'<p style="color:var(--muted);margin-bottom:20px">Jami {len(worker_stats)} ishchi | {active_count} faol smena</p>')

    # Statistika cards
    parts.append('<div class="stats-grid" style="margin-bottom:20px">')
    parts.append(f'<div class="stat-card"><div class="stat-label">👷 Jami ishchilar</div><div class="stat-value">{len(worker_stats)}</div><div class="stat-trend">{active_count} faol smena</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--green)"><div class="stat-label">📋 Oylik ishlar</div><div class="stat-value">{total_works}</div><div class="stat-trend">Tasdiqlangan</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid #f59e0b"><div class="stat-label">💰 Oylik daromad</div><div class="stat-value">{fmt(total_income)}</div><div class="stat-trend">{today.strftime("%B")} oyi</div></div>')
    avg_per = total_income / len(worker_stats) if worker_stats else 0
    parts.append(f'<div class="stat-card" style="border-left:4px solid #8b5cf6"><div class="stat-label">📊 O\'rtacha</div><div class="stat-value">{fmt(avg_per)}</div><div class="stat-trend">Bir ishchiga</div></div>')
    parts.append('</div>')

    # Qidiruv va saralash
    parts.append('<div class="card" style="margin-bottom:16px;padding:14px">')
    parts.append('<div style="display:flex;gap:10px;flex-wrap:wrap;align-items:center">')
    parts.append('<form method="get" style="flex:1;min-width:200px;display:flex;gap:8px">')
    parts.append(f'<input type="text" name="q" placeholder="🔍 Ishchi qidirish..." value="{h(search)}" class="finp" style="flex:1">')
    parts.append(f'<input type="hidden" name="sort" value="{sort}">')
    parts.append('<button type="submit" class="btn btn-p btn-sm">Qidirish</button>')
    parts.append('</form>')

    parts.append('<div style="display:flex;gap:4px;background:rgba(30,41,59,0.5);padding:4px;border-radius:10px">')
    for sk, sl in [("income", "💰 Daromad"), ("works", "📋 Ishlar"), ("name", "🔤 Nom"), ("recent", "🕒 So'nggi")]:
        q_str = f"?q={h(search)}&sort={sk}" if search else f"?sort={sk}"
        active = "active" if sort == sk else ""
        parts.append(f'<a href="{q_str}" class="btn-period {active}">{sl}</a>')
    parts.append('</div>')
    parts.append('</div></div>')

    # Worker cards grid
    parts.append('<div class="workers-grid">')
    if not worker_stats:
        parts.append('<p style="color:var(--muted);padding:40px;text-align:center;grid-column:1/-1">Ishchi topilmadi</p>')
    else:
        for w in worker_stats:
            u = w["u"]
            initial = (u.full_name[0] if u.full_name else "?").upper()
            block_badge = '<span class="block-badge">🚫 Bloklangan</span>' if not u.is_active else ""
            active_badge = '<span class="active-badge">🟢 Smenada</span>' if w["active"] else ""
            last_txt = ""
            if w["last_work"]:
                days_ago = (datetime.now() - w["last_work"]).days
                if days_ago == 0:
                    last_txt = "Bugun"
                elif days_ago == 1:
                    last_txt = "Kecha"
                elif days_ago < 7:
                    last_txt = f"{days_ago} kun oldin"
                else:
                    last_txt = w["last_work"].strftime("%d.%m.%Y")

            parts.append(f'<a href="/web/workers/{u.id}" class="worker-card">')
            parts.append(f'<div class="worker-avatar-big">{h(initial)}</div>')
            parts.append(f'<div class="worker-info">')
            parts.append(f'<div class="worker-name">{h(u.full_name or "—")} {block_badge}{active_badge}</div>')
            parts.append(f'<div class="worker-tg">TG: {u.telegram_id or "—"}</div>')
            parts.append(f'<div class="worker-stats">')
            parts.append(f'<span class="ws-item"><span class="ws-label">💰</span><b>{fmt(w["income"])}</b></span>')
            parts.append(f'<span class="ws-item"><span class="ws-label">📋</span><b>{w["works"]}</b> ish</span>')
            if last_txt:
                parts.append(f'<span class="ws-item ws-muted"><span class="ws-label">🕒</span>{last_txt}</span>')
            parts.append(f'</div>')
            parts.append(f'</div>')
            parts.append(f'</a>')

    parts.append('</div>')

    # CSS
    parts.append('<style>')
    parts.append('.workers-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(280px,1fr)); gap:12px }')
    parts.append('.worker-card { display:flex; gap:12px; padding:14px; background:rgba(30,41,59,0.5); border-radius:12px; text-decoration:none; color:inherit; transition:all .2s; border:1px solid transparent }')
    parts.append('.worker-card:hover { background:rgba(99,102,241,0.15); border-color:rgba(99,102,241,0.3); transform:translateY(-2px); box-shadow:0 8px 20px rgba(0,0,0,0.2) }')
    parts.append('.worker-avatar-big { width:48px; height:48px; border-radius:50%; background:linear-gradient(135deg,#6366f1,#8b5cf6); display:flex; align-items:center; justify-content:center; font-size:22px; font-weight:800; color:#fff; flex-shrink:0 }')
    parts.append('.worker-info { flex:1; min-width:0 }')
    parts.append('.worker-name { font-weight:700; font-size:14px; margin-bottom:2px; display:flex; align-items:center; gap:6px; flex-wrap:wrap }')
    parts.append('.worker-tg { color:var(--muted); font-size:11px; margin-bottom:6px }')
    parts.append('.worker-stats { display:flex; gap:10px; flex-wrap:wrap; font-size:12px }')
    parts.append('.ws-item { display:flex; align-items:center; gap:3px }')
    parts.append('.ws-label { opacity:0.7 }')
    parts.append('.ws-muted { color:var(--muted) }')
    parts.append('.block-badge { font-size:10px; padding:2px 6px; background:rgba(239,68,68,0.15); color:#fca5a5; border-radius:6px; font-weight:600 }')
    parts.append('.active-badge { font-size:10px; padding:2px 6px; background:rgba(16,185,129,0.15); color:#6ee7b7; border-radius:6px; font-weight:600 }')
    parts.append('</style>')

    content = "\n".join(parts)
    return web.Response(text=_base("Ishchilar", "workers", content), content_type="text/html")


@_require_auth
async def worker_detail(request: web.Request):
    """Ishchi haqida to'liq ma'lumot — grafiklar, statistika."""
    try:
        wid = int(request.match_info.get("worker_id", "0"))
    except ValueError:
        wid = 0
    if wid <= 0:
        raise web.HTTPFound("/web/workers")

    today = date.today()
    month_start = today.replace(day=1)

    async with AsyncSessionLocal() as db:
        worker = await db.get(User, wid)
        if not worker:
            raise web.HTTPFound("/web/workers")

        # Joriy oy statistikasi
        r = await db.execute(
            select(
                func.count(WorkEntry.id),
                func.coalesce(func.sum(WorkEntry.jami_summa), 0),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.approved, 1), else_=0)),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.pending,  1), else_=0)),
            ).where(
                WorkEntry.worker_id == wid,
                WorkEntry.work_date >= month_start,
            )
        )
        row = r.one()
        m_total, m_income = int(row[0] or 0), float(row[1] or 0)
        m_approved, m_rejected, m_pending = int(row[2] or 0), int(row[3] or 0), int(row[4] or 0)

        # 30 kunlik grafik
        days = []
        for i in range(29, -1, -1):
            d = today - timedelta(days=i)
            r_d = await db.execute(
                select(func.coalesce(func.sum(WorkEntry.jami_summa), 0)).where(
                    WorkEntry.worker_id == wid,
                    WorkEntry.work_date == d,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            days.append({"date": d.strftime("%d.%m"), "income": float(r_d.scalar() or 0)})

        # Ish turlari bo'yicha
        r_types = await db.execute(
            select(
                WorkEntry.work_type,
                func.count(WorkEntry.id),
                func.coalesce(func.sum(WorkEntry.jami_summa), 0),
            ).where(
                WorkEntry.worker_id == wid,
                WorkEntry.work_date >= month_start,
                WorkEntry.status == WorkStatus.approved,
            ).group_by(WorkEntry.work_type)
            .order_by(func.count(WorkEntry.id).desc())
        )
        types_breakdown = [(r[0].value if r[0] else "?", r[1], float(r[2])) for r in r_types.all()]

        # Jarima
        r_pen = await db.execute(
            select(func.coalesce(func.sum(Penalty.summa), 0))
            .where(
                Penalty.worker_id == wid,
                func.extract('month', Penalty.created_at) == today.month,
                func.extract('year',  Penalty.created_at) == today.year,
            )
        )
        m_penalty = float(r_pen.scalar() or 0)

        # Avans
        r_adv = await db.execute(
            select(func.coalesce(func.sum(Advance.summa), 0))
            .where(Advance.worker_id == wid, Advance.oy == today.month, Advance.yil == today.year)
        )
        m_advance = float(r_adv.scalar() or 0)

        sof_maosh = m_income - m_penalty - m_advance

        # Davomat — bu oy
        r_att = await db.execute(
            select(Attendance.tur, func.count(Attendance.id))
            .where(
                Attendance.worker_id == wid,
                func.extract('month', Attendance.sana) == today.month,
                func.extract('year',  Attendance.sana) == today.year,
            ).group_by(Attendance.tur)
        )
        attendance = {r[0].value if r[0] else "?": r[1] for r in r_att.all()}

        # So'nggi 10 ish
        r_recent = await db.execute(
            select(WorkEntry)
            .where(WorkEntry.worker_id == wid)
            .order_by(WorkEntry.created_at.desc())
            .limit(10)
        )
        recent = r_recent.scalars().all()

    import json
    days_labels  = json.dumps([d["date"] for d in days])
    days_incomes = json.dumps([d["income"] for d in days])

    types_html = ""
    for wt, cnt, inc in types_breakdown:
        types_html += f"""
        <tr>
          <td>{h(wt.replace('_', ' ').title())}</td>
          <td style="text-align:center">{cnt}</td>
          <td style="text-align:right">{fmt(inc)} so'm</td>
        </tr>
        """
    if not types_breakdown:
        types_html = '<tr><td colspan="3" style="text-align:center;color:var(--muted);padding:16px">Ish topilmadi</td></tr>'

    recent_html = ""
    for w in recent:
        status_color = {"approved": "var(--green)", "rejected": "var(--red)", "pending": "#f59e0b"}.get(w.status.value, "var(--muted)")
        status_label = {"approved": "✅", "rejected": "❌", "pending": "⏳"}.get(w.status.value, "?")
        recent_html += f"""
        <tr>
          <td>{w.work_date.strftime('%d.%m.%Y') if w.work_date else '-'}</td>
          <td>{h(w.work_type.value.replace('_', ' ').title())}</td>
          <td style="text-align:center">{w.soni:.0f}</td>
          <td style="text-align:right">{fmt(w.jami_summa)} so'm</td>
          <td style="text-align:center;color:{status_color}">{status_label}</td>
        </tr>
        """
    if not recent:
        recent_html = '<tr><td colspan="5" style="text-align:center;color:var(--muted);padding:16px">Ishlar yo\'q</td></tr>'

    att_ish     = attendance.get('ish', 0)
    att_kasal   = attendance.get('kasallik', 0)
    att_tatil   = attendance.get('tatil', 0)
    att_sababli = attendance.get('sababli', 0)
    att_sababsi = attendance.get('sababsiz', 0)

    content = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>

<div style="display:flex;align-items:center;gap:16px;margin-bottom:20px;flex-wrap:wrap">
  <div class="worker-avatar">{h(worker.full_name[0] if worker.full_name else '?')}</div>
  <div style="flex:1">
    <h1 style="margin:0">{h(worker.full_name or '—')}</h1>
    <p style="color:var(--muted);margin:4px 0 0 0">
      ID: {worker.id} | TG: {worker.telegram_id} |
      Rol: {worker.role.value if worker.role else '—'} |
      {('🟢 Faol' if worker.is_active else '🔴 Bloklangan')}
    </p>
  </div>
  <a href="/web/workers" class="btn">← Ortga</a>
</div>

<!-- ASOSIY KO'RSATKICHLAR -->
<div class="stats-grid" style="margin-bottom:20px">
  <div class="stat-card" style="border-left:4px solid var(--green)">
    <div class="stat-label">💰 Sof maosh ({today.strftime('%B')})</div>
    <div class="stat-value" style="color:var(--green)">{fmt(sof_maosh)} so'm</div>
    <div class="stat-trend">Daromad - Jarima - Avans</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">📋 Jami ishlar</div>
    <div class="stat-value">{m_total}</div>
    <div class="stat-trend">✅ {m_approved} | ❌ {m_rejected} | ⏳ {m_pending}</div>
  </div>
  <div class="stat-card" style="border-left:4px solid var(--red)">
    <div class="stat-label">⚠️ Jarima</div>
    <div class="stat-value" style="color:var(--red)">-{fmt(m_penalty)}</div>
    <div class="stat-trend">Bu oyda</div>
  </div>
  <div class="stat-card" style="border-left:4px solid #f59e0b">
    <div class="stat-label">💳 Avans</div>
    <div class="stat-value" style="color:#f59e0b">-{fmt(m_advance)}</div>
    <div class="stat-trend">Bu oyda</div>
  </div>
</div>

<!-- 30 KUN GRAFIK -->
<div class="card" style="margin-bottom:20px">
  <h2>📈 30 kunlik daromad</h2>
  <div style="height:280px;position:relative;margin-top:10px"><canvas id="incomeChart"></canvas></div>
</div>

<!-- DAVOMAT VA ISH TURLARI -->
<div class="charts-row">
  <div class="card">
    <h2>📅 Davomat ({today.strftime('%B')})</h2>
    <div class="att-grid" style="margin-top:12px">
      <div class="att-cell" style="border-color:var(--green)">
        <div class="att-icon">✅</div>
        <div class="att-num">{att_ish}</div>
        <div class="att-label">Ish kun</div>
      </div>
      <div class="att-cell" style="border-color:#f59e0b">
        <div class="att-icon">🤒</div>
        <div class="att-num">{att_kasal}</div>
        <div class="att-label">Kasallik</div>
      </div>
      <div class="att-cell" style="border-color:#3b82f6">
        <div class="att-icon">🌴</div>
        <div class="att-num">{att_tatil}</div>
        <div class="att-label">Ta'til</div>
      </div>
      <div class="att-cell" style="border-color:#8b5cf6">
        <div class="att-icon">📝</div>
        <div class="att-num">{att_sababli}</div>
        <div class="att-label">Sababli</div>
      </div>
      <div class="att-cell" style="border-color:var(--red)">
        <div class="att-icon">⛔</div>
        <div class="att-num">{att_sababsi}</div>
        <div class="att-label">Sababsiz</div>
      </div>
    </div>
  </div>

  <div class="card">
    <h2>📊 Ish turlari bo'yicha</h2>
    <table style="width:100%;margin-top:8px;font-size:13px">
      <thead><tr><th>Tur</th><th>Soni</th><th style="text-align:right">Daromad</th></tr></thead>
      <tbody>{types_html}</tbody>
    </table>
  </div>
</div>

<!-- SO'NGGI ISHLAR -->
<div class="card" style="margin-top:20px">
  <h2>📜 So'nggi 10 ish</h2>
  <div class="table-wrap" style="margin-top:8px">
    <table style="width:100%;font-size:13px">
      <thead>
        <tr>
          <th>Sana</th><th>Tur</th>
          <th style="text-align:center">Soni</th>
          <th style="text-align:right">Summa</th>
          <th style="text-align:center">Status</th>
        </tr>
      </thead>
      <tbody>{recent_html}</tbody>
    </table>
  </div>
</div>

<style>
.worker-avatar {{
  width:60px; height:60px; border-radius:50%;
  background:linear-gradient(135deg,#6366f1,#8b5cf6);
  display:flex; align-items:center; justify-content:center;
  font-size:28px; font-weight:800; color:#fff;
}}
.att-grid {{
  display:grid;
  grid-template-columns:repeat(auto-fill,minmax(80px,1fr));
  gap:8px;
}}
.att-cell {{
  background:rgba(30,41,59,0.5);
  border-radius:10px;
  padding:10px;
  text-align:center;
  border-top:3px solid #6366f1;
}}
.att-icon {{ font-size:18px }}
.att-num {{ font-size:20px; font-weight:800; margin:4px 0 }}
.att-label {{ font-size:10px; color:var(--muted) }}
</style>

<script>
Chart.defaults.color = '#94a3b8';
new Chart(document.getElementById('incomeChart'), {{
  type: 'bar',
  data: {{
    labels: {days_labels},
    datasets: [{{
      label: 'Daromad',
      data: {days_incomes},
      backgroundColor: 'rgba(99,102,241,0.6)',
      borderColor: '#6366f1',
      borderWidth: 1,
      borderRadius: 4,
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ beginAtZero: true, ticks: {{ callback: v => (v/1000).toFixed(0)+'k' }} }},
      x: {{ ticks: {{ maxRotation: 0, autoSkip: true, maxTicksLimit: 10 }} }},
    }},
  }},
}});
</script>
"""
    return web.Response(text=_base(f"{worker.full_name}", "workers", content), content_type="text/html")




@_require_auth
async def inventory_health(request: web.Request):
    """Ombor sog'lig'i — predictiv tahlil, trend, ogohlantirishlar."""
    today = date.today()

    async with AsyncSessionLocal() as db:
        # Kritik mahsulotlar (kam qoldi)
        r_critical = await db.execute(
            select(WarehouseProduct)
            .where(
                WarehouseProduct.is_active == True,
                WarehouseProduct.miqdor <= WarehouseProduct.min_threshold,
            )
            .order_by(WarehouseProduct.miqdor.asc())
            .limit(20)
        )
        critical = r_critical.scalars().all()

        # Sariq zonadagilar
        r_yellow = await db.execute(
            select(WarehouseProduct)
            .where(
                WarehouseProduct.is_active == True,
                WarehouseProduct.miqdor > WarehouseProduct.min_threshold,
                WarehouseProduct.miqdor <= WarehouseProduct.yellow_threshold,
            )
            .order_by(WarehouseProduct.miqdor.asc())
            .limit(20)
        )
        yellow_prods = r_yellow.scalars().all()

        # Kategoriya bo'yicha ombor qiymati va miqdor
        r_cat = await db.execute(
            select(
                WarehouseProduct.category,
                func.count(WarehouseProduct.id),
                func.coalesce(func.sum(WarehouseProduct.miqdor), 0),
            )
            .where(WarehouseProduct.is_active == True)
            .group_by(WarehouseProduct.category)
            .order_by(func.count(WarehouseProduct.id).desc())
        )
        cat_breakdown = [(r[0].value if r[0] else "?", r[1], float(r[2])) for r in r_cat.all()]

        # Oxirgi 14 kun harakat (kirim/chiqim)
        movement = []
        for i in range(13, -1, -1):
            d = today - timedelta(days=i)
            r_m = await db.execute(
                select(
                    sf.sum(sa_case((WarehouseLog.amal == "kirim",  WarehouseLog.miqdor), else_=0)),
                    sf.sum(sa_case((WarehouseLog.amal == "chiqim", WarehouseLog.miqdor), else_=0)),
                ).where(func.date(WarehouseLog.created_at) == d)
            )
            row = r_m.one()
            movement.append({
                "date":   d.strftime("%d.%m"),
                "kirim":  float(row[0] or 0),
                "chiqim": float(row[1] or 0),
            })

        # Eng faol mahsulotlar (chiqim)
        r_active = await db.execute(
            select(
                WarehouseProduct.name,
                func.count(WarehouseLog.id),
                func.sum(WarehouseLog.miqdor),
            )
            .join(WarehouseLog, WarehouseLog.product_id == WarehouseProduct.id)
            .where(
                WarehouseLog.amal == "chiqim",
                WarehouseLog.created_at >= today - timedelta(days=30),
            )
            .group_by(WarehouseProduct.id, WarehouseProduct.name)
            .order_by(func.sum(WarehouseLog.miqdor).desc())
            .limit(10)
        )
        most_used = [(r[0], r[1], float(r[2] or 0)) for r in r_active.all()]

    import json
    move_labels  = json.dumps([m["date"]   for m in movement])
    move_kirim   = json.dumps([m["kirim"]  for m in movement])
    move_chiqim  = json.dumps([m["chiqim"] for m in movement])

    cat_labels = json.dumps([c[0] for c in cat_breakdown])
    cat_counts = json.dumps([c[1] for c in cat_breakdown])

    # Kritik mahsulotlar HTML
    critical_html = ""
    for p in critical:
        days_left = "?"
        try:
            min_t = float(p.min_threshold or 0)
            critical_html += f"""
            <div class="health-row health-critical">
              <div class="health-name">
                <strong>{h(p.name)}</strong>
                {('<span class="badge">'+ h(p.razmer) +'</span>') if p.razmer else ''}
                {('<span class="badge">'+ h(p.rang)   +'</span>') if p.rang   else ''}
              </div>
              <div class="health-num" style="color:var(--red)">{p.miqdor:.0f} / {min_t:.0f} {p.birlik}</div>
              <div class="health-action">
                <button onclick="quickKirim({p.id},'{h(p.name).replace(chr(39), chr(92)+chr(39))}')" class="btn btn-sm btn-green">+ Kirim</button>
              </div>
            </div>
            """
        except Exception:
            pass

    if not critical_html:
        critical_html = '<p style="color:var(--muted);text-align:center;padding:20px">✅ Hozircha kritik mahsulot yo\'q!</p>'

    yellow_html = ""
    for p in yellow_prods:
        try:
            yellow_html += f"""
            <div class="health-row health-yellow">
              <div class="health-name">
                <strong>{h(p.name)}</strong>
                {('<span class="badge">'+ h(p.razmer) +'</span>') if p.razmer else ''}
              </div>
              <div class="health-num" style="color:#f59e0b">{p.miqdor:.0f} {p.birlik}</div>
            </div>
            """
        except Exception:
            pass
    if not yellow_html:
        yellow_html = '<p style="color:var(--muted);text-align:center;padding:20px">Bo\'sh</p>'

    most_used_html = ""
    max_used = max((u[2] for u in most_used), default=1)
    for name, cnt, total in most_used:
        pct = (total / max_used * 100) if max_used > 0 else 0
        most_used_html += f"""
        <div class="used-row">
          <div class="used-name">{h(name)}</div>
          <div class="used-bar"><div class="used-bar-fill" style="width:{pct:.0f}%"></div></div>
          <div class="used-num">{total:.0f}</div>
        </div>
        """
    if not most_used:
        most_used_html = '<p style="color:var(--muted);text-align:center;padding:20px">Ma\'lumot yo\'q</p>'

    critical_count = len(critical)
    yellow_count   = len(yellow_prods)

    content = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>

<h1 style="margin-bottom:4px">💊 Ombor sog'lig'i</h1>
<p style="color:var(--muted);margin-bottom:20px">Mahsulotlar holati, harakat trendi va kritik ogohlantirishlar</p>

<!-- OGOHLANTIRISHLAR -->
<div class="stats-grid" style="margin-bottom:20px">
  <div class="stat-card" style="border-left:4px solid var(--red)">
    <div class="stat-label">🔴 Kritik holatda</div>
    <div class="stat-value" style="color:var(--red)">{critical_count}</div>
    <div class="stat-trend">Darhol kirim kerak</div>
  </div>
  <div class="stat-card" style="border-left:4px solid #f59e0b">
    <div class="stat-label">🟡 Sariq zonada</div>
    <div class="stat-value" style="color:#f59e0b">{yellow_count}</div>
    <div class="stat-trend">Yetadi, lekin kuzating</div>
  </div>
  <div class="stat-card" style="border-left:4px solid var(--green)">
    <div class="stat-label">🟢 Yetarli</div>
    <div class="stat-value" style="color:var(--green)">{sum(c[1] for c in cat_breakdown) - critical_count - yellow_count}</div>
    <div class="stat-trend">Normal holatda</div>
  </div>
</div>

<!-- KRITIK MAHSULOTLAR -->
<div class="card" style="margin-bottom:20px">
  <h2 style="color:var(--red)">🔴 Kritik mahsulotlar (kirim kerak)</h2>
  <p style="color:var(--muted);font-size:12px;margin-bottom:12px">Minimal chegaradan kam qolgan mahsulotlar</p>
  <div class="health-list">{critical_html}</div>
</div>

<!-- SARIQ ZONA -->
<div class="card" style="margin-bottom:20px">
  <h2 style="color:#f59e0b">🟡 Sariq zonadagilar</h2>
  <p style="color:var(--muted);font-size:12px;margin-bottom:12px">Yetarli, lekin yaqin kelajakda tugaydi</p>
  <div class="health-list">{yellow_html}</div>
</div>

<!-- GRAFIKLAR -->
<div class="charts-row">
  <div class="card chart-card">
    <h2>📊 14 kunlik harakat</h2>
    <div style="height:280px;position:relative;margin-top:10px"><canvas id="moveChart"></canvas></div>
  </div>
  <div class="card chart-card">
    <h2>🥧 Kategoriya bo'yicha</h2>
    <div style="height:280px;position:relative;margin-top:10px"><canvas id="catChart"></canvas></div>
  </div>
</div>

<!-- ENG FAOL MAHSULOTLAR -->
<div class="card" style="margin-top:20px">
  <h2>🔥 Eng ko'p ishlatilgan (30 kun)</h2>
  <div class="used-list" style="margin-top:12px">{most_used_html}</div>
</div>

<style>
.health-list {{ display:flex; flex-direction:column; gap:6px }}
.health-row {{
  display:grid;
  grid-template-columns:1fr auto auto;
  gap:12px;
  align-items:center;
  padding:10px 14px;
  border-radius:8px;
}}
.health-critical {{ background:rgba(239,68,68,0.1); border-left:3px solid var(--red) }}
.health-yellow   {{ background:rgba(245,158,11,0.1); border-left:3px solid #f59e0b }}
.health-name {{ font-size:13px }}
.health-num {{ font-weight:700; font-size:14px }}
.badge {{
  display:inline-block; padding:2px 8px; border-radius:4px;
  background:rgba(99,102,241,0.2); color:#4F46E5;
  font-size:10px; font-weight:600; margin-left:6px;
}}

.used-list {{ display:flex; flex-direction:column; gap:6px }}
.used-row {{
  display:grid;
  grid-template-columns:180px 1fr 60px;
  gap:10px;
  align-items:center;
  padding:8px 12px;
  background:rgba(30,41,59,0.4);
  border-radius:6px;
}}
.used-name {{ font-size:12px; font-weight:600; overflow:hidden; text-overflow:ellipsis; white-space:nowrap }}
.used-bar {{ background:rgba(15,23,42,0.6); border-radius:4px; height:8px; overflow:hidden }}
.used-bar-fill {{
  height:100%;
  background:linear-gradient(90deg,#6366f1,#8b5cf6);
  border-radius:4px;
}}
.used-num {{ text-align:right; font-weight:700; font-size:13px }}

@media (max-width:768px) {{
  .health-row {{ grid-template-columns:1fr auto }}
  .health-action {{ grid-column:1/-1 }}
  .used-row {{ grid-template-columns:1fr 60px; row-gap:4px }}
  .used-bar {{ grid-column:1/-1 }}
}}
</style>

<script>
Chart.defaults.color = '#94a3b8';

new Chart(document.getElementById('moveChart'), {{
  type: 'line',
  data: {{
    labels: {move_labels},
    datasets: [
      {{
        label: 'Kirim',
        data: {move_kirim},
        borderColor: '#10b981',
        backgroundColor: 'rgba(16,185,129,0.1)',
        tension: 0.35,
      }},
      {{
        label: 'Chiqim',
        data: {move_chiqim},
        borderColor: '#ef4444',
        backgroundColor: 'rgba(239,68,68,0.1)',
        tension: 0.35,
      }},
    ],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    interaction: {{ mode: 'index', intersect: false }},
    scales: {{ y: {{ beginAtZero: true }} }},
  }},
}});

new Chart(document.getElementById('catChart'), {{
  type: 'pie',
  data: {{
    labels: {cat_labels},
    datasets: [{{
      data: {cat_counts},
      backgroundColor: ['#6366f1','#8b5cf6','#ec4899','#f43f5e','#f59e0b','#10b981','#06b6d4','#3b82f6','#a855f7','#84cc16','#f97316','#14b8a6'],
    }}],
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    plugins: {{ legend: {{ position: 'right', labels: {{ font: {{ size: 10 }}, padding: 6 }} }} }},
  }},
}});

function quickKirim(id, nom) {{
  const m = prompt(`+Kirim: ${{nom}}\nMiqdor:`);
  if (!m || isNaN(m) || +m <= 0) return;
  fetch('/web/warehouse/kirim', {{
    method:'POST', headers:{{"Content-Type":"application/json"}},
    body: JSON.stringify({{product_id: id, miqdor: +m}})
  }}).then(r=>r.json()).then(d=>{{
    if(d.ok) {{ alert('✅ Yangi qoldiq: ' + d.new_miqdor); location.reload() }}
    else alert('❌ ' + (d.error||''))
  }});
}}
</script>
"""
    return web.Response(text=_base("Ombor sog'lig'i", "inventory_health", content), content_type="text/html")




@_require_auth
async def quality_dashboard(request: web.Request):
    """Sifat tahlili — rad etilganlar, sabablar, ishchilar bo'yicha."""
    today = date.today()
    period = request.query.get("period", "month")

    if period == "week":
        start_date = today - timedelta(days=7)
        period_name = "Oxirgi 7 kun"
    elif period == "quarter":
        start_date = today - timedelta(days=90)
        period_name = "Oxirgi 3 oy"
    else:
        start_date = today.replace(day=1)
        period_name = "Joriy oy"

    async with AsyncSessionLocal() as db:
        # Umumiy QA
        r = await db.execute(
            select(
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.approved, 1), else_=0)),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.pending,  1), else_=0)),
            ).where(WorkEntry.work_date >= start_date)
        )
        row = r.one()
        total, app, rej, pen = int(row[0] or 0), int(row[1] or 0), int(row[2] or 0), int(row[3] or 0)

        approval_rate = (app / (app + rej) * 100) if (app + rej) > 0 else 0
        rejection_rate = (rej / (app + rej) * 100) if (app + rej) > 0 else 0

        # Ishchilar bo'yicha rad etilganlar
        r_workers = await db.execute(
            select(
                User.full_name,
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            )
            .join(WorkEntry, WorkEntry.worker_id == User.id)
            .where(
                WorkEntry.work_date >= start_date,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            )
            .group_by(User.id, User.full_name)
            .having(func.count(WorkEntry.id) >= 1)
            .order_by(sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)).desc())
            .limit(15)
        )
        workers_qa = []
        for r in r_workers.all():
            t, rj = int(r[1] or 0), int(r[2] or 0)
            rate = (rj / t * 100) if t > 0 else 0
            workers_qa.append({"name": r[0], "total": t, "rejected": rj, "rate": rate})

        # Ish turlari bo'yicha rad etish
        r_types = await db.execute(
            select(
                WorkEntry.work_type,
                func.count(WorkEntry.id),
                sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
            )
            .where(
                WorkEntry.work_date >= start_date,
                WorkEntry.status.in_([WorkStatus.approved, WorkStatus.rejected]),
            )
            .group_by(WorkEntry.work_type)
            .order_by(sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)).desc())
        )
        types_qa = []
        for r in r_types.all():
            t, rj = int(r[1] or 0), int(r[2] or 0)
            rate = (rj / t * 100) if t > 0 else 0
            types_qa.append({"type": r[0].value if r[0] else "?", "total": t, "rejected": rj, "rate": rate})

        # Kunlik trend
        daily_qa = []
        for i in range(min((today - start_date).days, 30), -1, -1):
            d = today - timedelta(days=i)
            r = await db.execute(
                select(
                    sf.sum(sa_case((WorkEntry.status == WorkStatus.approved, 1), else_=0)),
                    sf.sum(sa_case((WorkEntry.status == WorkStatus.rejected, 1), else_=0)),
                ).where(WorkEntry.work_date == d)
            )
            row = r.one()
            daily_qa.append({
                "date":     d.strftime("%d.%m"),
                "approved": int(row[0] or 0),
                "rejected": int(row[1] or 0),
            })

    import json
    daily_labels   = json.dumps([d["date"]     for d in daily_qa])
    daily_approved = json.dumps([d["approved"] for d in daily_qa])
    daily_rejected = json.dumps([d["rejected"] for d in daily_qa])

    types_labels = json.dumps([t["type"] for t in types_qa])
    types_rates  = json.dumps([t["rate"] for t in types_qa])

    # Ishchilar HTML
    workers_html = ""
    for w in workers_qa:
        color = "var(--red)" if w["rate"] > 15 else ("#f59e0b" if w["rate"] > 5 else "var(--green)")
        workers_html += f"""
        <div class="qa-row">
          <div class="qa-name">{h(w['name'])}</div>
          <div class="qa-stats">
            <span class="qa-total">{w['total']}</span> jami,
            <span style="color:var(--red)">{w['rejected']}</span> rad
          </div>
          <div class="qa-rate" style="color:{color}">{w['rate']:.1f}%</div>
          <div class="qa-bar">
            <div class="qa-bar-fill" style="width:{w['rate']:.0f}%;background:{color}"></div>
          </div>
        </div>
        """
    if not workers_qa:
        workers_html = '<p style="color:var(--muted);text-align:center;padding:20px">Ma\'lumot yo\'q</p>'

    types_html = ""
    for t in types_qa:
        color = "var(--red)" if t["rate"] > 15 else ("#f59e0b" if t["rate"] > 5 else "var(--green)")
        types_html += f"""
        <div class="qa-row">
          <div class="qa-name">{h(t['type'].replace('_', ' ').title())}</div>
          <div class="qa-stats">{t['total']} ishdan {t['rejected']} rad</div>
          <div class="qa-rate" style="color:{color}">{t['rate']:.1f}%</div>
          <div class="qa-bar">
            <div class="qa-bar-fill" style="width:{t['rate']:.0f}%;background:{color}"></div>
          </div>
        </div>
        """
    if not types_qa:
        types_html = '<p style="color:var(--muted);text-align:center;padding:20px">Ma\'lumot yo\'q</p>'

    content = f"""
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>

<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;flex-wrap:wrap;gap:12px">
  <div>
    <h1 style="margin-bottom:2px">🎯 Sifat boshqaruvi</h1>
    <p style="color:var(--muted);margin:0">{period_name}</p>
  </div>
  <div class="period-switcher">
    <a href="?period=week"    class="btn-period {'active' if period=='week' else ''}">Hafta</a>
    <a href="?period=month"   class="btn-period {'active' if period=='month' else ''}">Oy</a>
    <a href="?period=quarter" class="btn-period {'active' if period=='quarter' else ''}">3 oy</a>
  </div>
</div>

<!-- ASOSIY KO'RSATKICHLAR -->
<div class="stats-grid" style="margin-bottom:20px">
  <div class="stat-card" style="border-left:4px solid var(--green)">
    <div class="stat-label">✅ Tasdiqlangan</div>
    <div class="stat-value" style="color:var(--green)">{app}</div>
    <div class="stat-trend">{approval_rate:.1f}% qabul darajasi</div>
  </div>
  <div class="stat-card" style="border-left:4px solid var(--red)">
    <div class="stat-label">❌ Rad etilgan</div>
    <div class="stat-value" style="color:var(--red)">{rej}</div>
    <div class="stat-trend">{rejection_rate:.1f}% rad darajasi</div>
  </div>
  <div class="stat-card" style="border-left:4px solid #f59e0b">
    <div class="stat-label">⏳ Kutilmoqda</div>
    <div class="stat-value" style="color:#f59e0b">{pen}</div>
    <div class="stat-trend">Tekshirilmagan</div>
  </div>
  <div class="stat-card">
    <div class="stat-label">📊 Jami</div>
    <div class="stat-value">{total}</div>
    <div class="stat-trend">{period_name.lower()}da</div>
  </div>
</div>

<!-- GRAFIK -->
<div class="card" style="margin-bottom:20px">
  <h2>📈 Kunlik sifat trendi</h2>
  <div style="height:280px;position:relative;margin-top:10px"><canvas id="dailyChart"></canvas></div>
</div>

<!-- ISHCHILAR -->
<div class="charts-row">
  <div class="card">
    <h2>👷 Ishchilar bo'yicha (eng ko'p rad)</h2>
    <div class="qa-list" style="margin-top:12px">{workers_html}</div>
  </div>
  <div class="card">
    <h2>🔧 Ish turlari bo'yicha</h2>
    <div class="qa-list" style="margin-top:12px">{types_html}</div>
  </div>
</div>

<style>
.qa-list {{ display:flex; flex-direction:column; gap:8px }}
.qa-row {{
  display:grid;
  grid-template-columns:1fr auto auto;
  gap:8px;
  align-items:center;
  padding:10px 12px;
  background:rgba(30,41,59,0.4);
  border-radius:8px;
}}
.qa-name {{ font-weight:600; font-size:13px }}
.qa-stats {{ color:var(--muted); font-size:11px }}
.qa-rate {{ font-weight:800; font-size:14px; min-width:50px; text-align:right }}
.qa-bar {{
  grid-column:1 / -1;
  background:rgba(15,23,42,0.6);
  height:5px;
  border-radius:3px;
  overflow:hidden;
  margin-top:4px;
}}
.qa-bar-fill {{ height:100%; border-radius:3px; transition:width .5s }}
</style>

<script>
Chart.defaults.color = '#94a3b8';
new Chart(document.getElementById('dailyChart'), {{
  type: 'line',
  data: {{
    labels: {daily_labels},
    datasets: [
      {{ label: 'Tasdiqlangan', data: {daily_approved},
         borderColor:'#10b981', backgroundColor:'rgba(16,185,129,0.1)', tension:0.35, fill:true }},
      {{ label: 'Rad etilgan', data: {daily_rejected},
         borderColor:'#ef4444', backgroundColor:'rgba(239,68,68,0.1)', tension:0.35, fill:true }},
    ],
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    interaction:{{ mode:'index', intersect:false }},
    scales:{{ y:{{ beginAtZero:true }} }},
  }},
}});
</script>
"""
    return web.Response(text=_base("Sifat boshqaruvi", "quality", content), content_type="text/html")




@_require_auth
async def notifications_center(request: web.Request):
    """Bildirishnomalar markazi — kritik ogohlantirishlar, kutilayotgan ishlar."""
    today = date.today()

    async with AsyncSessionLocal() as db:
        # Kritik mahsulotlar
        r_crit = await db.execute(
            select(WarehouseProduct).where(
                WarehouseProduct.is_active == True,
                WarehouseProduct.miqdor <= WarehouseProduct.min_threshold,
            ).order_by(WarehouseProduct.miqdor.asc()).limit(10)
        )
        critical = r_crit.scalars().all()

        # Kutilayotgan ishlar
        r_pending = await db.execute(
            select(WorkEntry, User.full_name)
            .join(User, WorkEntry.worker_id == User.id)
            .where(WorkEntry.status == WorkStatus.pending)
            .order_by(WorkEntry.created_at.asc())
            .limit(15)
        )
        pending_works = r_pending.all()

        # Tahrir so'rovlari
        r_edit = await db.execute(
            select(WorkEntry, User.full_name)
            .join(User, WorkEntry.worker_id == User.id)
            .where(WorkEntry.status == WorkStatus.edit_requested)
            .order_by(WorkEntry.created_at.desc())
            .limit(10)
        )
        edit_works = r_edit.all()

        # 0 ga tushgan mahsulotlar (zero_notified)
        try:
            r_zero = await db.execute(
                select(WarehouseProduct).where(
                    WarehouseProduct.is_active == True,
                    WarehouseProduct.miqdor <= 0,
                ).order_by(WarehouseProduct.name).limit(20)
            )
            zero_prods = r_zero.scalars().all()
        except Exception:
            zero_prods = []

        # Bloklangan ishchilar
        r_blocked = await db.execute(
            select(User).where(
                User.role == UserRole.ishchi,
                User.is_active == False,
            ).limit(20)
        )
        blocked = r_blocked.scalars().all()

    # HTML qurish
    def render_notif_row(icon, title, subtitle, action_url, color):
        return f"""
        <a href="{action_url}" class="notif-row" style="border-left-color:{color}">
          <div class="notif-icon">{icon}</div>
          <div class="notif-body">
            <div class="notif-title">{h(title)}</div>
            <div class="notif-sub">{h(subtitle)}</div>
          </div>
          <div class="notif-arrow">→</div>
        </a>
        """

    notif_html = ""

    if critical:
        notif_html += '<div class="notif-section"><h3>🔴 Kritik holatlar</h3>'
        for p in critical:
            details = []
            if p.razmer:     details.append(p.razmer)
            if p.razmer_tur: details.append(p.razmer_tur)
            if p.rang:       details.append(p.rang)
            sub = f"Qoldi: {p.miqdor:.0f} {p.birlik} (min: {p.min_threshold or 0})"
            if details: sub = " | ".join(details) + " — " + sub
            notif_html += render_notif_row("🔴", p.name, sub, "/web/warehouse", "var(--red)")
        notif_html += '</div>'

    if zero_prods:
        notif_html += '<div class="notif-section"><h3>⚪ Tugagan mahsulotlar</h3>'
        for p in zero_prods:
            notif_html += render_notif_row("0", p.name, f"Omborda: 0 {p.birlik}", f"/web/warehouse", "#64748b")
        notif_html += '</div>'

    if pending_works:
        notif_html += '<div class="notif-section"><h3>⏳ Tasdiqlashni kutmoqda</h3>'
        for w, wname in pending_works:
            sub = f"{wname} — {w.work_type.value if w.work_type else '?'} — {w.soni:.0f} dona — {fmt(w.jami_summa)} so'm"
            d = w.work_date.strftime("%d.%m") if w.work_date else "?"
            notif_html += render_notif_row("⏳", f"#{w.id} | {d}", sub, "/web/reports?status=pending", "#f59e0b")
        notif_html += '</div>'

    if edit_works:
        notif_html += '<div class="notif-section"><h3>✏️ Tahrir so\'rovlari</h3>'
        for w, wname in edit_works:
            sub = f"{wname} — {w.edit_note or 'Sabab yo\'q'}"
            notif_html += render_notif_row("✏️", f"#{w.id} | {w.work_type.value if w.work_type else '?'}", sub,
                                          "/web/reports?status=edit_requested", "#8b5cf6")
        notif_html += '</div>'

    if blocked:
        notif_html += '<div class="notif-section"><h3>⛔ Bloklangan ishchilar</h3>'
        for u in blocked:
            notif_html += render_notif_row("⛔", u.full_name or "Noma'lum",
                                          f"TG ID: {u.telegram_id}", f"/web/workers/{u.id}", "#64748b")
        notif_html += '</div>'

    if not notif_html:
        notif_html = '<div style="text-align:center;padding:60px;color:var(--muted)"><div style="font-size:48px">✅</div><h2>Hammasi yaxshi!</h2><p>Hozircha hech qanday ogohlantirish yo\'q.</p></div>'

    total_alerts = len(critical) + len(pending_works) + len(edit_works) + len(zero_prods) + len(blocked)

    content = f"""
<h1 style="margin-bottom:4px">🔔 Bildirishnomalar markazi</h1>
<p style="color:var(--muted);margin-bottom:20px">
  Jami {total_alerts} ta diqqat talab qiluvchi xabar
</p>

<!-- Filtr tugmalar -->
<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px">
  <span class="filter-chip filter-red">🔴 {len(critical)} kritik</span>
  <span class="filter-chip filter-orange">⏳ {len(pending_works)} kutmoqda</span>
  <span class="filter-chip filter-purple">✏️ {len(edit_works)} tahrir</span>
  <span class="filter-chip filter-gray">0 {len(zero_prods)} tugagan</span>
  <span class="filter-chip filter-dark">⛔ {len(blocked)} bloklangan</span>
</div>

{notif_html}

<style>
.filter-chip {{
  display:inline-block;
  padding:6px 12px;
  border-radius:20px;
  font-size:12px;
  font-weight:700;
  background:rgba(30,41,59,0.5);
}}
.filter-red    {{ background:rgba(239,68,68,0.15); color:#fca5a5 }}
.filter-orange {{ background:rgba(245,158,11,0.15); color:#fcd34d }}
.filter-purple {{ background:rgba(139,92,246,0.15); color:#c4b5fd }}
.filter-gray   {{ background:rgba(100,116,139,0.15); color:#535C78 }}
.filter-dark   {{ background:rgba(15,23,42,0.5); color:#94a3b8 }}

.notif-section {{
  background:rgba(30,41,59,0.3);
  border-radius:12px;
  padding:16px;
  margin-bottom:14px;
}}
.notif-section h3 {{ margin:0 0 12px 0; font-size:14px; color:var(--muted) }}

.notif-row {{
  display:grid;
  grid-template-columns:36px 1fr 24px;
  gap:12px;
  align-items:center;
  padding:10px 14px;
  background:rgba(15,23,42,0.4);
  border-radius:8px;
  border-left:4px solid #6366f1;
  margin-bottom:6px;
  text-decoration:none;
  color:inherit;
  transition:all .15s;
}}
.notif-row:hover {{
  background:rgba(99,102,241,0.1);
  transform:translateX(4px);
}}
.notif-icon {{ font-size:20px; text-align:center }}
.notif-title {{ font-weight:600; font-size:13px; margin-bottom:2px }}
.notif-sub   {{ font-size:11px; color:var(--muted); overflow:hidden; text-overflow:ellipsis; white-space:nowrap }}
.notif-arrow {{ color:var(--muted); text-align:right; font-size:14px }}
</style>
"""
    return web.Response(text=_base("Bildirishnomalar", "notif", content), content_type="text/html")




@_require_auth
async def global_search(request):
    q = request.query.get("q", "").strip()
    results = {"products": [], "workers": [], "works": []}

    if q and len(q) >= 2:
        q_like = f"%{q.lower()}%"
        async with AsyncSessionLocal() as db:
            r = await db.execute(
                select(WarehouseProduct)
                .where(
                    WarehouseProduct.is_active == True,
                    func.lower(WarehouseProduct.name).like(q_like),
                ).limit(20)
            )
            results["products"] = r.scalars().all()

            r2 = await db.execute(
                select(User)
                .where(func.lower(User.full_name).like(q_like))
                .limit(20)
            )
            results["workers"] = r2.scalars().all()

            if q.isdigit():
                r3 = await db.execute(
                    select(WorkEntry, User.full_name)
                    .join(User, WorkEntry.worker_id == User.id)
                    .where(WorkEntry.id == int(q))
                    .limit(5)
                )
                results["works"] = r3.all()

    prod_html = ""
    for p in results["products"]:
        details = []
        if p.razmer:     details.append(p.razmer)
        if p.razmer_tur: details.append(p.razmer_tur)
        if p.tur:        details.append(p.tur)
        if p.rang:       details.append(p.rang)
        if p.qism:       details.append("qism: " + str(p.qism))
        if p.yonalish:   details.append("yo'nalish: " + str(p.yonalish))
        det = " | ".join(details) if details else "-"
        color = "var(--red)" if p.miqdor <= (p.min_threshold or 0) else "var(--green)"
        prod_html += (
            '<a href="/web/ombor/' + (p.category.value if p.category else "warehouse") + '" class="search-row">'
            '<div class="search-icon">P</div>'
            '<div class="search-body">'
            '<div class="search-title">' + h(p.name) + '</div>'
            '<div class="search-sub">' + h(det) + '</div>'
            '</div>'
            '<div class="search-num" style="color:' + color + '">' + f"{p.miqdor:.0f} {p.birlik}" + '</div>'
            '</a>'
        )

    work_html = ""
    for u in results["workers"]:
        role = u.role.value if u.role else "?"
        status_icon = "Active" if u.is_active else "Block"
        work_html += (
            '<a href="/web/workers/' + str(u.id) + '" class="search-row">'
            '<div class="search-icon">' + status_icon + '</div>'
            '<div class="search-body">'
            '<div class="search-title">' + h(u.full_name or "-") + '</div>'
            '<div class="search-sub">' + role + ' | TG: ' + str(u.telegram_id) + '</div>'
            '</div>'
            '<div class="search-num">arrow</div>'
            '</a>'
        )

    works_html = ""
    for w, wname in results["works"]:
        status = w.status.value if w.status else "?"
        status_color = {"approved": "var(--green)", "rejected": "var(--red)", "pending": "#f59e0b"}.get(status, "var(--muted)")
        works_html += (
            '<a href="/web/reports" class="search-row">'
            '<div class="search-icon">W</div>'
            '<div class="search-body">'
            '<div class="search-title">#' + str(w.id) + ' - ' + h(wname or "?") + '</div>'
            '<div class="search-sub">' + (w.work_type.value if w.work_type else "?") + ' | ' + f"{w.soni:.0f}" + ' dona</div>'
            '</div>'
            '<div class="search-num" style="color:' + status_color + '">' + status + '</div>'
            '</a>'
        )

    total = len(results["products"]) + len(results["workers"]) + len(results["works"])
    products_count = len(results["products"])
    workers_count  = len(results["workers"])
    works_count    = len(results["works"])

    # HTML komponentlari
    info_block = ""
    if q:
        info_block = (
            '<p style="color:var(--muted);margin-bottom:14px">Topildi: <b>' + str(total) +
            '</b> natija "<i>' + h(q) + '</i>" bo\'yicha</p>'
        )

    prod_block = ""
    if q:
        prod_block = (
            '<div class="card" style="margin-bottom:14px">'
            '<h2>Mahsulotlar (' + str(products_count) + ')</h2>'
            '<div class="search-list">' + (prod_html or '<p style="color:var(--muted);padding:10px">Topilmadi</p>') + '</div>'
            '</div>'
        )

    work_block = ""
    if q:
        work_block = (
            '<div class="card" style="margin-bottom:14px">'
            '<h2>Ishchilar (' + str(workers_count) + ')</h2>'
            '<div class="search-list">' + (work_html or '<p style="color:var(--muted);padding:10px">Topilmadi</p>') + '</div>'
            '</div>'
        )

    works_block = ""
    if q and q.isdigit():
        works_block = (
            '<div class="card" style="margin-bottom:14px">'
            '<h2>Ishlar (ID: ' + h(q) + ')</h2>'
            '<div class="search-list">' + (works_html or '<p style="color:var(--muted);padding:10px">Topilmadi</p>') + '</div>'
            '</div>'
        )

    empty_block = ""
    if not q:
        empty_block = (
            '<div style="text-align:center;padding:60px;color:var(--muted)">'
            '<div style="font-size:48px;margin-bottom:12px">?</div>'
            '<h2 style="font-size:18px">Qidirish uchun yozing</h2>'
            '<p>Mahsulot nomi, ishchi nomi yoki ish ID raqami (kamida 2 belgi)</p>'
            '</div>'
        )

    content = (
        '<h1 style="margin-bottom:14px">Qidiruv</h1>'
        '<form method="GET" style="margin-bottom:20px">'
        '<div style="display:flex;gap:8px">'
        '<input type="text" name="q" value="' + h(q) + '" placeholder="Mahsulot, ishchi yoki ish ID..." '
        'autofocus style="flex:1;padding:12px 16px;font-size:14px;border-radius:10px;'
        'border:1px solid #1e293b;background:rgba(15,23,42,0.6);color:var(--fg)">'
        '<button class="btn" type="submit">Qidirish</button>'
        '</div></form>'
        + info_block + prod_block + work_block + works_block + empty_block
        + '<style>'
        '.search-list { display:flex; flex-direction:column; gap:4px }'
        '.search-row { display:grid; grid-template-columns:32px 1fr auto; gap:10px; '
        'align-items:center; padding:10px 12px; background:rgba(30,41,59,0.4); '
        'border-radius:6px; text-decoration:none; color:inherit; transition:all .15s; }'
        '.search-row:hover { background:rgba(99,102,241,0.1); transform:translateX(4px) }'
        '.search-icon { font-size:14px; text-align:center; font-weight:700; color:#8b5cf6 }'
        '.search-title { font-weight:600; font-size:13px }'
        '.search-sub { font-size:11px; color:var(--muted) }'
        '.search-num { font-weight:700; font-size:13px }'
        '</style>'
    )
    return web.Response(text=_base("Qidiruv", "search", content), content_type="text/html")




@_require_auth
async def help_page(request):
    content = '''
<h1 style="margin-bottom:14px">Yordam markazi</h1>
<p style="color:var(--muted);margin-bottom:20px">Tez-tez beriladigan savollar va qo\'llanma</p>

<div class="card" style="margin-bottom:14px">
  <h2>Ombor boshqaruvi</h2>
  <details style="margin-top:10px">
    <summary class="help-q">Qanday qilib mahsulot qo\'shaman?</summary>
    <p class="help-a">"Mahsulotlar" -> "Yangi mahsulot" tugmasini bosing va formani to\'ldiring. Kategoriya, tur va miqdorni tanlang.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Mahsulot tugagandan keyin nima bo\'ladi?</summary>
    <p class="help-a">Mahsulot 0 ga tushganda bot avtomatik ravishda adminga xabar yuboradi va "saqlansin yoki o\'chirilsin?" deb so\'raydi. Agar mahsulot kerak emas bo\'lsa, butunlay o\'chiriladi.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Xromazes yo\'nalishi nima?</summary>
    <p class="help-a">Xromazes qayerga ishlatilishini ko\'rsatadi: "Tiger kesish" (to\'g\'ridan-to\'g\'ri) yoki "Zagatovka chain" (zagatovka -> gofra kley -> tiger). Bu filtrlash uchun kerak.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Qism nima (tepa/past/yon/paddo)?</summary>
    <p class="help-a">Adyol va Pastel karobkalar 3 qismdan iborat: Adyol = TEPA + PAST + YON (2 ta), Pastel = TEPA + PAST + PADDO. Har bir qism alohida saqlanadi.</p>
  </details>
</div>

<div class="card" style="margin-bottom:14px">
  <h2>Ishchilar va maoshlar</h2>
  <details style="margin-top:10px">
    <summary class="help-q">Avans qanday beriladi?</summary>
    <p class="help-a">"Avans" sahifasiga kiring, ishchini tanlang, summani kiriting. Yoki bot orqali admin "Avans berish" tugmasini ishlatadi.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Jarima qanday qo\'yiladi?</summary>
    <p class="help-a">"Jarimalar" sahifasida ishchi, jarima turi (jarima yoki mukofot), summa va sababni kiriting.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Maosh hisoboti qaerdan ko\'rinadi?</summary>
    <p class="help-a">"Maoshlar" sahifasida har oy uchun maoshlar ko\'rinadi. "Maosh hisoboti" Excel formatida yuklab olish mumkin.</p>
  </details>
</div>

<div class="card" style="margin-bottom:14px">
  <h2>Ishlab chiqarish</h2>
  <details style="margin-top:10px">
    <summary class="help-q">Ish bosqichlari qanday?</summary>
    <p class="help-a">Rulon -> Gofra ishlab -> Xromazeslar (yonalish) -> Laminat / Zagatovka -> Gofra kley -> Tiger -> Tikish -> Qoqish. Har bosqichda ishchi material tanlaydi.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Ish nima uchun rad etilishi mumkin?</summary>
    <p class="help-a">Nazoratchi yoki admin sifatsiz ishni rad etadi. Rad etilganda omborda ayirilgan material qaytariladi.</p>
  </details>
  <details style="margin-top:10px">
    <summary class="help-q">Smena qanday ochiladi/yopiladi?</summary>
    <p class="help-a">Ishchi "Smena boshlash" -> material tanlash -> ish boshlanadi. "Smena tugatish" bilan tugaydi va ishlar saqlanadi.</p>
  </details>
</div>

<div class="card">
  <h2>Tezkor harakatlar</h2>
  <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px;margin-top:10px">
    <a href="/web/warehouse" class="help-link">Ombor</a>
    <a href="/web/workers" class="help-link">Ishchilar</a>
    <a href="/web/reports" class="help-link">Hisobotlar</a>
    <a href="/web/inventory-health" class="help-link">Kritik holatlar</a>
    <a href="/web/notifications" class="help-link">Bildirishnomalar</a>
    <a href="/web/quality" class="help-link">Sifat tahlili</a>
  </div>
</div>

<style>
.help-q {
  cursor: pointer;
  padding: 10px;
  background: rgba(30,41,59,0.4);
  border-radius: 6px;
  font-weight: 600;
  font-size: 13px;
  list-style: none;
  transition: background .15s;
}
.help-q:hover { background: rgba(99,102,241,0.1) }
.help-q::before { content: "+ "; color: #6366f1; font-weight: 700; }
details[open] .help-q::before { content: "- " }
.help-a {
  padding: 10px 16px;
  margin: 4px 0 0 0;
  color: var(--muted);
  font-size: 13px;
  line-height: 1.6;
}
.help-link {
  background: rgba(99,102,241,0.1);
  border: 1px solid rgba(99,102,241,0.2);
  border-radius: 8px;
  padding: 10px 14px;
  font-size: 13px;
  text-decoration: none;
  color: var(--fg);
  text-align: center;
  transition: all .15s;
}
.help-link:hover {
  background: rgba(99,102,241,0.2);
  transform: translateY(-2px);
}
</style>
'''
    return web.Response(text=_base("Yordam", "help", content), content_type="text/html")


@_require_auth
async def avans(request: web.Request):
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(Advance)
            .where(Advance.oy == oy, Advance.yil == yil)
            .order_by(Advance.created_at.desc())
        )
        advances = r.scalars().all()
        workers = await get_users_by_role(db, UserRole.ishchi)
        worker_map = {w.id: w for w in workers}
        total = sum(float(a.summa) for a in advances)

    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    worker_opts = ''.join(f'<option value="{w.id}">{h(w.full_name)}</option>' for w in workers)
    rows = ""
    for adv in advances:
        w = worker_map.get(adv.worker_id)
        rows += f"""<tr>
<td class="t-xs t-muted">{adv.created_at.strftime('%d.%m.%Y %H:%M') if adv.created_at else '—'}</td>
<td><strong>{h(w.full_name if w else '?')}</strong></td>
<td class="td-n cv-yellow">{fmt(adv.summa)} so'm</td>
<td class="t-sm">{h(adv.izoh or '—')}</td>
</tr>"""
    if not rows:
        rows = '<tr><td colspan="4" class="empty-state">💳 Bu oyda avans yo\'q</td></tr>'
    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">💳 Avans boshqaruvi</span>
  <div style="display:flex;gap:7px;align-items:center">
    <form method="GET" style="display:flex;gap:6px">
      <select name="oy" class="fsel" onchange="this.form.submit()">{month_sel}</select>
    </form>
    <button class="btn btn-p btn-sm" onclick="openModal('modal-avans')">➕ Avans berish</button>
  </div>
</div>
<div class="stats-grid sg-2" style="margin-bottom:11px">
  <div class="stat-card cl-yellow">
    <div class="s-label">Bu oyda jami avans</div>
    <div class="s-value cv-yellow">{fmt(total)}</div>
    <div class="s-sub">{len(advances)} ta to'lov</div>
  </div>
  <div class="stat-card cl-blue">
    <div class="s-label">Limit</div>
    <div class="s-value cv-blue">{AVANS_MAX_PER_MONTH}</div>
    <div class="s-sub">ta/oy har ishchiga</div>
  </div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Sana</th><th>Ishchi</th><th>Summa</th><th>Izoh</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>

<div class="overlay" id="modal-avans" onclick="if(event.target===this)closeModal('modal-avans')">
  <div class="modal">
    <div class="modal-title">💳 Avans berish</div>
    <form method="POST" action="/web/avans/add">
      <input type="hidden" name="oy" value="{oy_str}">
      <div class="fg">
        <label class="fl">Ishchi *</label>
        <select name="worker_id" required>{worker_opts}</select>
      </div>
      <div class="fg">
        <label class="fl">Summa (so'm) *</label>
        <input type="number" name="summa" required min="1000" step="1000">
      </div>
      <div class="fg">
        <label class="fl">Izoh</label>
        <input type="text" name="izoh" placeholder="ixtiyoriy">
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-avans')">Bekor</button>
        <button type="submit" class="btn btn-p">💳 Berish</button>
      </div>
    </form>
  </div>
</div>
"""
    return web.Response(text=_base("💳 Avans", "avans", content), content_type="text/html")

async def avans_add(request: web.Request):
    data = await request.post()
    wid = int(data.get("worker_id", 0))
    try:
        summa = float(str(data.get("summa", "0")).replace(",", "").replace(" ", ""))
    except Exception:
        summa = 0.0
    izoh = data.get("izoh") or None
    oy_str = data.get("oy", "")
    oy, yil = parse_month(oy_str) if oy_str else (datetime.now().month, datetime.now().year)
    if wid > 0 and summa > 0:
        async with AsyncSessionLocal() as db:
            try:
                # Birinchi admin/superadmin foydalanuvchini topish
                r = await db.execute(
                    select(User).where(User.role.in_([UserRole.admin, UserRole.superadmin]))
                    .limit(1)
                )
                admin = r.scalar_one_or_none()
                admin_id = admin.id if admin else wid  # fallback: worker_id

                adv = Advance(
                    worker_id=wid, admin_id=admin_id,
                    summa=summa, izoh=izoh, oy=oy, yil=yil
                )
                db.add(adv)
                await db.commit()
            except Exception as e:
                logger.error("Avans qo'shishda xato: %s", e)
                await db.rollback()
    raise web.HTTPFound(f"/web/avans?oy={oy_str}")

# ─── PENALTIES ─────────────────────────────────────────────────────────────

@_require_auth
async def penalties(request: web.Request):
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(Penalty)
            .where(extract("month", Penalty.created_at) == oy,
                   extract("year", Penalty.created_at) == yil)
            .order_by(Penalty.created_at.desc())
        )
        pens = r.scalars().all()
        workers = await get_users_by_role(db, UserRole.ishchi)
        worker_map = {w.id: w for w in workers}
        total = sum(float(p.summa) for p in pens)

    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    worker_opts = ''.join(f'<option value="{w.id}">{h(w.full_name)}</option>' for w in workers)
    rows = ""
    for pen in pens:
        w = worker_map.get(pen.worker_id)
        ptype = pen.penalty_type.value if hasattr(pen.penalty_type, 'value') else str(pen.penalty_type)
        conf = '<span class="badge bg">✅ Ko\'rdi</span>' if pen.worker_confirmed else '<span class="badge by">⏳ Kutmoqda</span>'
        rows += f"""<tr>
<td class="t-xs t-muted">{pen.created_at.strftime('%d.%m.%Y') if pen.created_at else '—'}</td>
<td><strong>{h(w.full_name if w else '?')}</strong></td>
<td>{h(PENALTY_LABELS.get(ptype, ptype))}</td>
<td class="td-n cv-red">{fmt(pen.summa)} so'm</td>
<td class="t-sm">{h(pen.sabab or '—')}</td>
<td>{conf}</td>
</tr>"""
    if not rows:
        rows = '<tr><td colspan="6" class="empty-state">✅ Bu oyda jarima yo\'q</td></tr>'
    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">⚠️ Jarimalar</span>
  <div style="display:flex;gap:7px;align-items:center">
    <form method="GET" style="display:flex;gap:6px">
      <select name="oy" class="fsel" onchange="this.form.submit()">{month_sel}</select>
    </form>
    <button class="btn btn-d btn-sm" onclick="openModal('modal-pen')">➕ Jarima qo'shish</button>
  </div>
</div>
<div class="stat-card cl-red" style="margin-bottom:11px">
  <div class="s-label">Bu oyda jami jarima</div>
  <div class="s-value cv-red">{fmt(total)}</div>
  <div class="s-sub">{len(pens)} ta jarima</div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Sana</th><th>Ishchi</th><th>Tur</th><th>Summa</th><th>Sabab</th><th>Holat</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>

<div class="overlay" id="modal-pen" onclick="if(event.target===this)closeModal('modal-pen')">
  <div class="modal">
    <div class="modal-title">⚠️ Jarima qo'shish</div>
    <form method="POST" action="/web/penalties/add">
      <input type="hidden" name="oy" value="{oy_str}">
      <div class="fg">
        <label class="fl">Ishchi *</label>
        <select name="worker_id" required>{worker_opts}</select>
      </div>
      <div class="fg">
        <label class="fl">Jarima turi</label>
        <select name="penalty_type">
          <option value="jarima">💸 Jarima (pul)</option>
          <option value="xaypsan1">⚠️ 1-xaypsan</option>
          <option value="xaypsan2">🚫 2-xaypsan</option>
        </select>
      </div>
      <div class="fg">
        <label class="fl">Summa (so'm)</label>
        <input type="number" name="summa" value="0" min="0" step="1000">
        <div class="input-hint">Xaypsan uchun 0 kiriting</div>
      </div>
      <div class="fg">
        <label class="fl">Sabab *</label>
        <textarea name="sabab" required placeholder="Jarima sababi..."></textarea>
      </div>
      <div class="modal-footer">
        <button type="button" class="btn btn-s" onclick="closeModal('modal-pen')">Bekor</button>
        <button type="submit" class="btn btn-d">⚠️ Qo'shish</button>
      </div>
    </form>
  </div>
</div>
"""
    return web.Response(text=_base("⚠️ Jarimalar", "penalties", content), content_type="text/html")

async def penalties_add(request: web.Request):
    data = await request.post()
    wid = int(data.get("worker_id", 0))
    try:
        summa = float(str(data.get("summa", "0")).replace(",", "").replace(" ", ""))
    except Exception:
        summa = 0.0
    sabab = data.get("sabab", "")
    oy_str = data.get("oy", "")
    try:
        ptype = PenaltyType(data.get("penalty_type", "jarima"))
    except ValueError:
        ptype = PenaltyType.jarima
    if wid > 0 and sabab:
        async with AsyncSessionLocal() as db:
            try:
                # Real admin/inspector ni topish
                r = await db.execute(
                    select(User).where(User.role.in_([UserRole.admin, UserRole.superadmin, UserRole.nazoratchi]))
                    .limit(1)
                )
                admin = r.scalar_one_or_none()
                inspector_id = admin.id if admin else wid

                pen = Penalty(
                    worker_id=wid, inspector_id=inspector_id,
                    penalty_type=ptype, summa=summa, sabab=sabab,
                )
                db.add(pen)
                await db.commit()
            except Exception as e:
                logger.error("Penalty qo'shishda xato: %s", e)
                await db.rollback()
    raise web.HTTPFound(f"/web/penalties?oy={oy_str}")

# ─── REPORTS ───────────────────────────────────────────────────────────────



@_require_auth
async def salary_projection(request: web.Request):
    """Maoshlar prognozi — kelajak oy uchun bashorat."""
    today = date.today()
    month_start = today.replace(day=1)
    days_passed = today.day
    
    # Joriy oy oxirgi kuni
    import calendar as cal
    last_day = cal.monthrange(today.year, today.month)[1]
    days_remaining = last_day - days_passed

    async with AsyncSessionLocal() as db:
        workers_list = await get_users_by_role(db, UserRole.ishchi)
        
        projections = []
        for u in workers_list:
            # Joriy oygacha bo'lgan daromad
            r = await db.execute(
                select(func.coalesce(func.sum(WorkEntry.jami_summa), 0))
                .where(
                    WorkEntry.worker_id == u.id,
                    WorkEntry.work_date >= month_start,
                    WorkEntry.work_date <= today,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            current = float(r.scalar() or 0)
            
            # Kunlik o'rtacha
            daily_avg = current / days_passed if days_passed > 0 else 0
            # Prognoz
            projected = current + (daily_avg * days_remaining)
            
            # Jarima va avans
            r_pen = await db.execute(
                select(func.coalesce(func.sum(Penalty.summa), 0))
                .where(
                    Penalty.worker_id == u.id,
                    func.extract('month', Penalty.created_at) == today.month,
                    func.extract('year',  Penalty.created_at) == today.year,
                )
            )
            penalty = float(r_pen.scalar() or 0)
            
            r_adv = await db.execute(
                select(func.coalesce(func.sum(Advance.summa), 0))
                .where(Advance.worker_id == u.id, Advance.oy == today.month, Advance.yil == today.year)
            )
            advance = float(r_adv.scalar() or 0)
            
            projected_net = projected - penalty - advance
            current_net = current - penalty - advance
            
            projections.append({
                "u": u, "current": current, "projected": projected,
                "penalty": penalty, "advance": advance,
                "current_net": current_net, "projected_net": projected_net,
                "daily_avg": daily_avg,
            })
        
        projections.sort(key=lambda x: x["projected"], reverse=True)
        total_current = sum(p["current"] for p in projections)
        total_projected = sum(p["projected"] for p in projections)
        total_penalty = sum(p["penalty"] for p in projections)
        total_advance = sum(p["advance"] for p in projections)

    parts = []
    parts.append('<h1 style="margin-bottom:4px">💼 Maosh prognozi</h1>')
    parts.append(f'<p style="color:var(--muted);margin-bottom:20px">Oy oxirigacha: <b>{days_remaining}</b> kun qoldi | {days_passed}/{last_day}</p>')
    
    # Top cards
    progress_pct = (days_passed / last_day * 100) if last_day > 0 else 0
    parts.append('<div class="stats-grid" style="margin-bottom:20px">')
    parts.append(f'<div class="stat-card"><div class="stat-label">💰 Joriy daromad</div><div class="stat-value">{fmt(total_current)}</div><div class="stat-trend">jami ishlangan</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--green)"><div class="stat-label">🔮 Prognoz</div><div class="stat-value" style="color:var(--green)">{fmt(total_projected)}</div><div class="stat-trend">oy oxiriga</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid var(--red)"><div class="stat-label">⚠️ Jarima</div><div class="stat-value" style="color:var(--red)">-{fmt(total_penalty)}</div><div class="stat-trend">bu oyda</div></div>')
    parts.append(f'<div class="stat-card" style="border-left:4px solid #f59e0b"><div class="stat-label">💳 Avans</div><div class="stat-value" style="color:#f59e0b">-{fmt(total_advance)}</div><div class="stat-trend">bu oyda</div></div>')
    parts.append('</div>')
    
    # Progress bar
    parts.append('<div class="card" style="margin-bottom:16px">')
    parts.append(f'<h3 style="margin-bottom:10px">📅 Oy davomiyligi: {progress_pct:.0f}%</h3>')
    parts.append(f'<div style="background:rgba(30,41,59,0.5);border-radius:12px;height:30px;overflow:hidden;position:relative">')
    parts.append(f'<div style="background:linear-gradient(90deg,#6366f1,#8b5cf6);height:100%;width:{progress_pct:.1f}%;transition:width .5s;display:flex;align-items:center;justify-content:flex-end;padding-right:10px;color:#fff;font-weight:700;font-size:13px">{days_passed}/{last_day}</div>')
    parts.append('</div></div>')
    
    # Jadval
    parts.append('<div class="card">')
    parts.append('<h2>📊 Ishchilar bo\'yicha</h2>')
    parts.append('<div class="table-wrap" style="margin-top:10px">')
    parts.append('<table style="width:100%;font-size:13px">')
    parts.append('<thead><tr><th>Ishchi</th><th style="text-align:right">Joriy</th><th style="text-align:right">Kunlik o\'rtacha</th><th style="text-align:right">⚠️</th><th style="text-align:right">💳</th><th style="text-align:right">Prognoz (sof)</th></tr></thead>')
    parts.append('<tbody>')
    
    for p in projections:
        u = p["u"]
        diff_color = "var(--green)" if p["projected_net"] > p["current_net"] else "var(--muted)"
        parts.append('<tr>')
        parts.append(f'<td><a href="/web/workers/{u.id}" style="color:inherit;text-decoration:none;font-weight:600">{h(u.full_name or "—")}</a></td>')
        parts.append(f'<td style="text-align:right">{fmt(p["current"])}</td>')
        parts.append(f'<td style="text-align:right;color:var(--muted)">{fmt(p["daily_avg"])}/kun</td>')
        parts.append(f'<td style="text-align:right;color:var(--red)">-{fmt(p["penalty"])}</td>')
        parts.append(f'<td style="text-align:right;color:#f59e0b">-{fmt(p["advance"])}</td>')
        parts.append(f'<td style="text-align:right;color:{diff_color};font-weight:700">{fmt(p["projected_net"])}</td>')
        parts.append('</tr>')
    
    parts.append('</tbody></table>')
    parts.append('</div></div>')
    
    content = "\n".join(parts)
    return web.Response(text=_base("Maosh prognozi", "salary_projection", content), content_type="text/html")


@_require_auth
async def reports(request: web.Request):
    tab = request.rel_url.query.get("tab", "daily")
    report_date_str = request.rel_url.query.get("date", date.today().isoformat())
    try:
        report_date = date.fromisoformat(report_date_str)
    except Exception:
        report_date = date.today()
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    async with AsyncSessionLocal() as db:
        if tab == "daily":
            r = await db.execute(
                select(WorkEntry, User)
                .join(User, WorkEntry.worker_id == User.id)
                .where(WorkEntry.work_date == report_date)
                .order_by(User.full_name, WorkEntry.created_at)
            )
            entries = r.all()
            rows = ""
            total = 0.0
            for entry, worker in entries:
                st = entry.status.value if hasattr(entry.status, 'value') else str(entry.status)
                cls = STATUS_CLS.get(st, "bgr")
                lbl = STATUS_LABELS.get(st, st)
                if st in ("approved", "adjusted"):
                    total += float(entry.jami_summa or 0)
                rows += f"""<tr>
<td><strong>{h(worker.full_name)}</strong></td>
<td class="t-xs">{h(WORK_TYPE_LABELS.get(entry.work_type.value if hasattr(entry.work_type,'value') else str(entry.work_type), ''))}</td>
<td class="t-xs">{h(entry.mahsulot_nomi or '—')}</td>
<td class="t-xs">{h(entry.razmer or '—')}</td>
<td class="td-n">{entry.soni}</td>
<td class="td-n">{fmt(entry.birlik_narx or 0)}</td>
<td class="td-n fw7">{fmt(entry.jami_summa or 0)}</td>
<td><span class="badge {cls}">{lbl}</span></td>
</tr>"""
            if not rows:
                rows = '<tr><td colspan="8" class="empty-state">📭 Bu kun ish yo\'q</td></tr>'
            body = f"""
<div style="display:flex;gap:7px;align-items:center;margin-bottom:11px;flex-wrap:wrap">
  <input type="date" value="{report_date_str}" onchange="window.location='/web/reports?tab=daily&date='+this.value"
    style="background:var(--bg3);border:1px solid var(--border);border-radius:7px;padding:6px 10px;color:var(--text);font-size:12px">
  <a href="/web/reports/download/daily?date={report_date_str}" class="btn btn-cy btn-sm">📥 Excel</a>
  <span class="badge bg fw7">Jami: {fmt(total)} so'm</span>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Ishchi</th><th>Tur</th><th>Mahsulot</th><th>Razmer</th><th>Soni</th><th>Narx</th><th>Summa</th><th>Holat</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>"""
        elif tab == "monthly":
            r = await db.execute(
                select(WorkEntry, User)
                .join(User, WorkEntry.worker_id == User.id)
                .where(extract("month", WorkEntry.work_date) == oy,
                       extract("year", WorkEntry.work_date) == yil)
                .order_by(User.full_name, WorkEntry.work_date.desc())
            )
            entries = r.all()
            rows = ""
            total = 0.0
            for entry, worker in entries:
                st = entry.status.value if hasattr(entry.status, 'value') else str(entry.status)
                cls = STATUS_CLS.get(st, "bgr")
                if st in ("approved", "adjusted"):
                    total += float(entry.jami_summa or 0)
                rows += f"""<tr>
<td class="t-xs t-muted">{entry.work_date}</td>
<td><strong>{h(worker.full_name)}</strong></td>
<td class="t-xs">{h(WORK_TYPE_LABELS.get(entry.work_type.value if hasattr(entry.work_type,'value') else str(entry.work_type), ''))}</td>
<td class="td-n">{entry.soni}</td>
<td class="td-n fw7">{fmt(entry.jami_summa or 0)}</td>
<td><span class="badge {cls}">{STATUS_LABELS.get(st,'')}</span></td>
</tr>"""
            if not rows:
                rows = '<tr><td colspan="6" class="empty-state">📭 Bu oy ish yo\'q</td></tr>'
            body = f"""
<div style="display:flex;gap:7px;align-items:center;margin-bottom:11px;flex-wrap:wrap">
  <select class="fsel" onchange="window.location='/web/reports?tab=monthly&oy='+this.value">{month_sel}</select>
  <a href="/web/reports/download/monthly?oy={oy_str}" class="btn btn-cy btn-sm">📥 Excel</a>
  <span class="badge bg fw7">Jami: {fmt(total)} so'm</span>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Sana</th><th>Ishchi</th><th>Tur</th><th>Soni</th><th>Summa</th><th>Holat</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>"""
        else:
            body = ""

    tab_html = ""
    for tk, tl in [("daily", "📅 Kunlik"), ("monthly", "🗓 Oylik")]:
        active = " active" if tab == tk else ""
        tab_html += f'<a href="/web/reports?tab={tk}" class="tab{active}">{tl}</a>'
    content = f"""
<div class="tabs">{tab_html}</div>
{body}
"""
    return web.Response(text=_base("📊 Ish hisobotlari", "reports", content), content_type="text/html")

@_require_auth
async def ombor_report(request: web.Request):
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(WarehouseLog)
            .where(extract("month", WarehouseLog.created_at) == oy,
                   extract("year", WarehouseLog.created_at) == yil)
            .order_by(WarehouseLog.created_at.desc())
        )
        logs = r.scalars().all()
        kirim_total = sum(float(lg.miqdor) for lg in logs if lg.amal == "kirim")
        chiqim_total = sum(float(lg.miqdor) for lg in logs if lg.amal == "chiqim")
        rows = ""
        for lg in logs:
            product = await get_product_by_id(db, lg.product_id)
            pname = h(product.name) if product else f"id={lg.product_id}"
            badge = '<span class="badge bg">📥 Kirim</span>' if lg.amal == "kirim" else '<span class="badge br">📤 Chiqim</span>'
            ts = lg.created_at.strftime("%d.%m.%Y %H:%M") if lg.created_at else "—"
            rows += f"""<tr>
<td class="t-xs t-muted">{ts}</td>
<td><strong>{pname}</strong></td>
<td>{badge}</td>
<td class="td-n">{lg.miqdor}</td>
<td class="td-n t-muted">{lg.oldin} → {lg.keyin}</td>
<td class="t-xs t-muted">{h(lg.izoh or '—')}</td>
</tr>"""
    if not rows:
        rows = '<tr><td colspan="6" class="empty-state">📭 Bu oyda harakat yo\'q</td></tr>'
    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">🏭 Ombor harakati hisoboti</span>
  <form method="GET" style="display:flex;gap:6px">
    <select name="oy" class="fsel" onchange="this.form.submit()">{month_sel}</select>
  </form>
</div>
<div class="stats-grid sg-2" style="margin-bottom:11px">
  <div class="stat-card cl-green">
    <div class="s-label">Jami kirim</div>
    <div class="s-value cv-green">{fmt(kirim_total)}</div>
    <div class="s-sub">birlik</div>
  </div>
  <div class="stat-card cl-red">
    <div class="s-label">Jami chiqim</div>
    <div class="s-value cv-red">{fmt(chiqim_total)}</div>
    <div class="s-sub">birlik</div>
  </div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Vaqt</th><th>Mahsulot</th><th>Amal</th><th>Miqdor</th><th>Oldin → Keyin</th><th>Izoh</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>
"""
    return web.Response(text=_base("🏭 Ombor hisoboti", "ombor-report", content), content_type="text/html")

@_require_auth
async def maosh_report(request: web.Request):
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    async with AsyncSessionLocal() as db:
        reports_list = await get_monthly_reports(db, oy, yil)
        rows = ""
        grand_ish = grand_jarima = grand_avans = grand_sof = 0.0
        for rep in reports_list:
            worker = await get_user_by_id(db, rep.worker_id)
            st = "✅ Tasdiqlangan" if rep.admin_tasdiqladi else "⏳ Kutmoqda"
            cls = "bg" if rep.admin_tasdiqladi else "by"
            grand_ish += float(rep.jami_ish_summa)
            grand_jarima += float(rep.jami_jarima)
            grand_avans += float(rep.jami_avans)
            grand_sof += float(rep.sof_maosh)
            rows += f"""<tr>
<td><strong>{h(worker.full_name if worker else '?')}</strong></td>
<td class="td-n">{fmt(rep.jami_ish_summa)}</td>
<td class="td-n cv-red">{fmt(rep.jami_jarima)}</td>
<td class="td-n cv-yellow">{fmt(rep.jami_avans)}</td>
<td class="td-n cv-green fw7">{fmt(rep.sof_maosh)}</td>
<td><span class="badge {cls}">{st}</span></td>
</tr>"""
        await db.commit()
    if not rows:
        rows = '<tr><td colspan="6" class="empty-state">📭 Maosh hisoboti yo\'q</td></tr>'
    content = f"""
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">💼 Maosh hisoboti</span>
  <div style="display:flex;gap:7px;align-items:center">
    <form method="GET" style="display:flex;gap:6px">
      <select name="oy" class="fsel" onchange="this.form.submit()">{month_sel}</select>
    </form>
    <a href="/web/reports/download/monthly?oy={oy_str}" class="btn btn-cy btn-sm">📥 Excel</a>
  </div>
</div>
<div class="stats-grid sg-4" style="margin-bottom:11px">
  <div class="stat-card cl-blue">
    <div class="s-label">Jami ish</div>
    <div class="s-value cv-blue">{fmt(grand_ish)}</div>
    <div class="s-sub">so'm</div>
  </div>
  <div class="stat-card cl-red">
    <div class="s-label">Jarimalar</div>
    <div class="s-value cv-red">{fmt(grand_jarima)}</div>
    <div class="s-sub">so'm</div>
  </div>
  <div class="stat-card cl-yellow">
    <div class="s-label">Avanslar</div>
    <div class="s-value cv-yellow">{fmt(grand_avans)}</div>
    <div class="s-sub">so'm</div>
  </div>
  <div class="stat-card cl-green">
    <div class="s-label">Jami sof maosh</div>
    <div class="s-value cv-green">{fmt(grand_sof)}</div>
    <div class="s-sub">so'm</div>
  </div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Ishchi</th><th>Ish summasi</th><th>Jarimalar</th><th>Avanslar</th><th>Sof maosh</th><th>Holat</th></tr></thead>
      <tbody>{rows}</tbody>
      <tfoot><tr style="background:var(--bg3)">
        <td class="fw7">JAMI</td>
        <td class="td-n fw7">{fmt(grand_ish)}</td>
        <td class="td-n fw7 cv-red">{fmt(grand_jarima)}</td>
        <td class="td-n fw7 cv-yellow">{fmt(grand_avans)}</td>
        <td class="td-n fw7 cv-green">{fmt(grand_sof)}</td>
        <td></td>
      </tr></tfoot>
    </table>
  </div>
</div>
"""
    return web.Response(text=_base("💼 Maosh hisoboti", "maosh-report", content), content_type="text/html")

# ─── SALARY PANEL ──────────────────────────────────────────────────────────

@_require_auth
async def salary(request: web.Request):
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    oy, yil = parse_month(oy_str)
    month_sel = ''.join(
        f'<option value="{m["val"]}" {"selected" if m["val"]==oy_str else ""}>{h(m["label"])}</option>'
        for m in months_list()
    )
    async with AsyncSessionLocal() as db:
        workers = await get_users_by_role(db, UserRole.ishchi)
        rows = ""
        grand = 0.0
        for worker in workers:
            rep = await calculate_and_save_salary(db, worker.id, oy, yil)
            st_cls = "bg" if rep.admin_tasdiqladi else "by"
            st_lbl = "✅ Tasdiqlangan" if rep.admin_tasdiqladi else "⏳ Kutmoqda"
            grand += float(rep.sof_maosh)
            btn = "" if rep.admin_tasdiqladi else f"""
<form method="POST" action="/web/salary/confirm/{rep.id}?oy={oy_str}" style="display:inline">
  <button class="btn btn-g btn-xs">✅ Tasdiqlash</button>
</form>"""
            rows += f"""<tr>
<td><strong>{h(worker.full_name)}</strong></td>
<td class="td-n">{fmt(rep.jami_ish_summa)}</td>
<td class="td-n cv-red">{fmt(rep.jami_jarima)}</td>
<td class="td-n cv-yellow">{fmt(rep.jami_avans)}</td>
<td class="td-n cv-green fw7">{fmt(rep.sof_maosh)}</td>
<td><span class="badge {st_cls}">{st_lbl}</span></td>
<td>{btn}</td>
</tr>"""
        await db.commit()

    pending_count = 0
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(func.count(SalaryReport.id))
            .where(SalaryReport.oy == oy, SalaryReport.yil == yil,
                   SalaryReport.admin_tasdiqladi == False)
        )
        pending_count = r.scalar() or 0

    if not rows:
        rows = '<tr><td colspan="7" class="empty-state">📭 Ishchilar topilmadi</td></tr>'
    warn = f'<div class="alert alert-w">⏳ {pending_count} ta ishchi maoshi tasdiqlanmagan!</div>' if pending_count else ""
    content = f"""
{warn}
<div class="card-hd" style="margin-bottom:11px">
  <span class="card-title">💰 Maosh paneli</span>
  <div style="display:flex;gap:7px;align-items:center">
    <form method="GET" style="display:flex;gap:6px">
      <select name="oy" class="fsel" onchange="this.form.submit()">{month_sel}</select>
    </form>
    <form method="POST" action="/web/salary/confirm-all?oy={oy_str}">
      <button type="submit" class="btn btn-g btn-sm">✅ Barchasini tasdiqlash</button>
    </form>
  </div>
</div>
<div class="stat-card cl-green" style="margin-bottom:11px">
  <div class="s-label">Jami sof maosh</div>
  <div class="s-value cv-green">{fmt(grand)}</div>
  <div class="s-sub">so'm — {MONTHS_UZ[oy]} {yil}</div>
</div>
<div class="card" style="padding:0">
  <div class="tbl-wrap">
    <table>
      <thead><tr><th>Ishchi</th><th>Ish</th><th>Jarima</th><th>Avans</th><th>Sof maosh</th><th>Holat</th><th></th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>
"""
    return web.Response(text=_base("💰 Maosh paneli", "salary", content), content_type="text/html")

async def salary_confirm(request: web.Request):
    rid = int(request.match_info["id"])
    oy_str = request.rel_url.query.get("oy", "")
    async with AsyncSessionLocal() as db:
        r = await db.execute(select(SalaryReport).where(SalaryReport.id == rid))
        rep = r.scalar_one_or_none()
        if rep:
            rep.admin_tasdiqladi = True
            rep.tasdiq_vaqti = datetime.now()
            await db.commit()
    raise web.HTTPFound(f"/web/salary?oy={oy_str}")

async def salary_confirm_all(request: web.Request):
    oy_str = request.rel_url.query.get("oy", "")
    now = datetime.now()
    oy, yil = parse_month(oy_str) if oy_str else (now.month, now.year)
    async with AsyncSessionLocal() as db:
        r = await db.execute(
            select(SalaryReport)
            .where(SalaryReport.oy == oy, SalaryReport.yil == yil,
                   SalaryReport.admin_tasdiqladi == False)
        )
        for rep in r.scalars().all():
            rep.admin_tasdiqladi = True
            rep.tasdiq_vaqti = datetime.now()
        await db.commit()
    raise web.HTTPFound(f"/web/salary?oy={oy_str}")

# ─── PRICES ────────────────────────────────────────────────────────────────

@_require_auth
async def prices(request: web.Request):
    """Narxlar — har ish turi va variant uchun alohida narx + saqlash."""
    from constants import get_variants, get_work_name

    async with AsyncSessionLocal() as db:
        all_prices = await get_all_prices(db)

    price_map = {}
    for p in all_prices:
        wt = p.work_type.value if hasattr(p.work_type, "value") else str(p.work_type)
        rt = (p.razmer_turi or "Standart")
        price_map[(wt, rt)] = p.narx

    parts = []
    parts.append('<h1 style="margin-bottom:4px">Narxlar</h1>')
    parts.append('<p style="color:var(--muted);margin-bottom:20px">Har ish turi va razmer uchun alohida narx belgilang</p>')

    for wt in WorkType:
        wt_val = wt.value
        variants = get_variants(wt_val)
        wt_name = get_work_name(wt_val)

        parts.append('<div class="card" style="margin-bottom:14px">')
        parts.append('<h2 style="margin-bottom:12px">' + wt_name + '</h2>')

        for variant in variants:
            current = price_map.get((wt_val, variant), 0)
            current_str = str(int(current)) if current else ""
            input_id = ("price_" + wt_val + "_" + variant).replace(" ", "_").replace("'", "")
            row = '<div style="display:flex;align-items:center;gap:10px;padding:8px 0;border-bottom:1px solid var(--border);flex-wrap:wrap">'
            row += '<div style="flex:0 0 160px;font-weight:600;font-size:13px">' + variant + '</div>'
            row += '<input type="number" id="' + input_id + '" value="' + current_str + '" placeholder="narx" style="flex:1;max-width:180px;padding:8px;border-radius:8px;border:1px solid var(--border);background:var(--bg2);color:var(--fg)">'
            row += '<span style="color:var(--muted);font-size:13px">som</span>'
            row += '<button class="btn btn-cy btn-xs" onclick="savePrice(\'' + wt_val + '\', \'' + variant + '\', \'' + input_id + '\')">Saqlash</button>'
            row += '</div>'
            parts.append(row)

        parts.append('</div>')

    js = """
<script>
function savePrice(workType, variant, inputId) {
  var val = document.getElementById(inputId).value;
  if (val === "" || isNaN(val)) { alert("Narx kiriting"); return; }
  var btn = event.target;
  fetch('/web/prices/save', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({work_type: workType, variant: variant, narx: parseFloat(val)})
  })
  .then(function(r){ return r.json(); })
  .then(function(d){
    if (d.ok) {
      var old = btn.textContent;
      btn.textContent = "Saqlandi";
      btn.style.background = "#10b981";
      setTimeout(function(){ btn.textContent = old; btn.style.background = ""; }, 1500);
    } else { alert("Xato: " + (d.error || "?")); }
  })
  .catch(function(e){ alert("Tarmoq xatosi: " + e); });
}
</script>
"""
    parts.append(js)
    content = "\n".join(parts)
    return web.Response(text=_base("Narxlar", "prices", content), content_type="text/html")


@_require_auth
async def prices_save(request: web.Request):
    """Narxni saqlash (work_type + variant)."""
    try:
        data = await request.json()
        wt_val  = data["work_type"]
        variant = data["variant"]
        narx    = float(data["narx"])
    except Exception as e:
        return web.json_response({"ok": False, "error": str(e)})

    db_variant = None if variant == "Standart" else variant

    async with AsyncSessionLocal() as db:
        from sqlalchemy import select
        wt_enum = WorkType(wt_val)
        r = await db.execute(
            select(WorkPrice).where(
                WorkPrice.work_type == wt_enum,
                WorkPrice.razmer_turi == db_variant,
            )
        )
        existing = r.scalar_one_or_none()
        if existing:
            existing.narx = narx
            existing.is_active = True
        else:
            db.add(WorkPrice(work_type=wt_enum, razmer_turi=db_variant, narx=narx, is_active=True))
        await db.commit()
    return web.json_response({"ok": True})


async def prices_set(request: web.Request):
    data = await request.post()
    try:
        wt = WorkType(data.get("work_type", ""))
        narx = float(str(data.get("narx", "0")).replace(",", "").replace(" ", ""))
        rzm = data.get("razmer_turi", "").strip() or None
        birlik = data.get("birlik", "dona")
    except Exception as e:
        logger.error("Narx saqlashda xato: %s", e)
        raise web.HTTPFound("/web/prices")
    async with AsyncSessionLocal() as db:
        price = await set_price(db, wt, narx, razmer_turi=rzm)
        price.birlik = birlik
        await db.commit()
    raise web.HTTPFound("/web/prices")

# ─── EXCEL DOWNLOADS ───────────────────────────────────────────────────────

async def reports_download(request: web.Request):
    rtype = request.match_info["t"]
    now = datetime.now()
    oy_str = request.rel_url.query.get("oy", f"{now.year}-{now.month:02d}")
    date_str = request.rel_url.query.get("date", date.today().isoformat())
    async with AsyncSessionLocal() as db:
        if rtype == "daily":
            try:
                report_date = date.fromisoformat(date_str)
            except Exception:
                report_date = date.today()
            data = await generate_daily_excel(db, report_date)
            fname = f"kunlik_{report_date}.xlsx"
        elif rtype == "weekly":
            data = await generate_weekly_excel(db)
            fname = f"haftalik_{date.today()}.xlsx"
        elif rtype == "monthly":
            oy, yil = parse_month(oy_str)
            data = await generate_monthly_excel(db, oy, yil)
            fname = f"oylik_{oy}_{yil}.xlsx"
        else:
            raise web.HTTPNotFound()
    return web.Response(
        body=data.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── CSV IMPORT ─────────────────────────────────────────────────────────────

async def warehouse_csv_import(request: web.Request):
    """
    CSV orqali mahsulotlarni ommaviy kiritish.
    GET  → forma ko'rsatish
    POST → import qilish

    CSV format (header bilan):
    category,name,razmer,rang,tur,birlik,miqdor,min_threshold,yellow_threshold

    Misol:
    yarim_tayyor,Tiger uchun karobka,Katta,Oq,adyol_tikish_uchun,dona,150,5,20
    adyol_zapchast,Ruchka,,Qora,,dona,200,10,30
    """
    if request.method == "GET":
        html = """<!DOCTYPE html><html><head>
<meta charset="utf-8"><title>CSV Import — Quti Tsexi</title>
<style>
*{box-sizing:border-box}
body{font-family:'Segoe UI',sans-serif;max-width:960px;margin:40px auto;padding:0 24px;background:#1a1a2e;color:#e0e0e0}
h2{color:#4fc3f7;margin-bottom:4px}
p.sub{color:#90a4ae;font-size:14px;margin-top:0}
.card{background:#16213e;border-radius:14px;padding:28px;margin:20px 0;box-shadow:0 4px 20px rgba(0,0,0,.3)}
textarea{width:100%;height:180px;background:#0d0d1a;color:#e0e0e0;border:1px solid #2a3a4a;border-radius:8px;padding:14px;font-family:monospace;font-size:13px;resize:vertical}
.btn{background:#4fc3f7;color:#000;border:none;padding:11px 28px;border-radius:8px;cursor:pointer;font-weight:700;font-size:14px;transition:.2s}
.btn:hover{background:#81d4fa}
.btn-g{background:#66bb6a;color:#000}
.btn-g:hover{background:#81c784}
.info{background:#0d2137;border-left:4px solid #4fc3f7;padding:14px 16px;border-radius:4px;margin:14px 0;font-size:13px;line-height:1.7}
.warn{background:#1a1200;border-left:4px solid #ffa726;padding:10px 14px;border-radius:4px;font-size:13px}
pre{background:#0d0d1a;padding:14px;border-radius:8px;overflow-x:auto;font-size:12px;line-height:1.6;border:1px solid #1e2d3d}
a.back{color:#4fc3f7;text-decoration:none;font-size:14px;display:inline-block;margin-bottom:20px}
a.back:hover{text-decoration:underline}
label{display:block;font-size:13px;color:#90a4ae;margin-bottom:6px;margin-top:16px}
input[type=file]{color:#e0e0e0;font-size:13px;padding:8px 0}
hr{border:none;border-top:1px solid #1e2d3d;margin:20px 0}
</style></head><body>
<a href="/web/warehouse" class="back">← Omborga qaytish</a>
<h2>📥 CSV orqali mahsulot import</h2>
<p class="sub">Bir vaqtda yuzlab mahsulotni kiritish imkoniyati</p>

<div class="card">
<div class="info">
<b>CSV ustun tartibi (header majburiy):</b><br>
<code>category, name, razmer, rang, tur, birlik, miqdor, min_threshold, yellow_threshold</code>
<br><br>
<b>Kategoriyalar:</b> rulon · gofra · gofra_zagatovka · xromazes · laminat_xromazes · yarim_tayyor · qolip · tayyor_mahsulot · adyol_zapchast · uskuna_zapchast<br>
<b>Birliklar:</b> dona · kg · top · rulon · m · m2 · litr<br>
<b>Razmer:</b> avtomatik normallanadi — "90 x 110" → "90x110"<br>
<b>Mavjud mahsulotlar:</b> takrorlanmaydi (o'tkazib yuboriladi)
</div>

<pre>category,name,razmer,rang,tur,birlik,miqdor,min_threshold,yellow_threshold
yarim_tayyor,Tiger uchun karobka,Katta,Oq,adyol_tikish_uchun,dona,150,5,20
yarim_tayyor,Tiger uchun karobka,O'rta,Oq,adyol_tikish_uchun,dona,80,5,20
adyol_zapchast,Ruchka,,Qora,,dona,200,10,30
adyol_zapchast,IP,,,,,100,10,50
tayyor_mahsulot,Tayyor Adyol karobka,Katta,,,dona,0,2,10</pre>

<form method="POST" enctype="multipart/form-data">
<label>📁 Fayl yuklash (Excel .xlsx yoki .csv):</label>
<input type="file" name="csv_file" accept=".csv,.txt,.xlsx"><br>
<div class="info">💡 Excel shablonni to'ldirib yuklang — eng tez yo'l. Shablonni adminдан so'rang yoki tayyor faylni ishlating.</div>

<div class="warn">⚠️ Yoki pastdagi maydonga CSV matnini to'g'ridan kiriting:</div>
<label>📝 CSV matni:</label>
<textarea name="csv_text" placeholder="category,name,razmer,rang,tur,birlik,miqdor,min_threshold,yellow_threshold
yarim_tayyor,Misol mahsulot,Katta,,adyol_tikish_uchun,dona,100,5,20"></textarea>
<br><br>
<button type="submit" class="btn">📥 Import qilish</button>
</form>
</div>
</body></html>"""
        return web.Response(text=html, content_type="text/html")

    # POST — import jarayoni
    import csv as _csv

    reader_mp = await request.multipart()
    csv_text = ""
    xlsx_rows = None
    while True:
        field = await reader_mp.next()
        if field is None:
            break
        if field.name == "csv_file":
            raw = await field.read()
            if raw:
                fname = (field.filename or "").lower()
                if fname.endswith(".xlsx") or raw[:2] == b"PK":
                    # Excel fayl — openpyxl bilan o'qiymiz
                    try:
                        import io
                        from openpyxl import load_workbook
                        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
                        ws = wb["Mahsulotlar"] if "Mahsulotlar" in wb.sheetnames else wb.active
                        xlsx_rows = []
                        for row in ws.iter_rows(values_only=True):
                            xlsx_rows.append(row)
                    except Exception as e:
                        xlsx_rows = None
                        csv_text = ""
                else:
                    csv_text = raw.decode("utf-8-sig", errors="replace")
        elif field.name == "csv_text":
            raw = await field.read()
            if raw and not csv_text and xlsx_rows is None:
                csv_text = raw.decode("utf-8-sig", errors="replace")

    if not csv_text.strip() and not xlsx_rows:
        raise web.HTTPFound("/web/warehouse/csv-import")

    from sqlalchemy import select as _sel
    valid_cats = {c.value for c in ProductCategory}

    # Excel qatorlarini CSV matnga aylantirib, bir xil quvurdan o'tkazamiz
    if xlsx_rows:
        EXPECTED = ["category", "name", "tur", "qism", "razmer", "rang", "gramaj", "birlik", "miqdor", "min_threshold", "yellow_threshold"]
        ncol = len(EXPECTED)
        lines = [",".join(EXPECTED)]
        for row in xlsx_rows:
            if row is None:
                continue
            vals = ["" if v is None else str(v).strip() for v in row][:ncol]
            if not any(vals):
                continue
            first = vals[0].lower() if vals else ""
            # Sarlavha / izoh / namuna qatorlarini o'tkazib yuboramiz
            if first in ("category", "bo'lim*", "bo'lim", "bolim"):
                continue
            # faqat haqiqiy kategoriya bo'lsa qabul qilamiz
            if first not in valid_cats:
                continue
            while len(vals) < ncol:
                vals.append("")
            safe = ['"' + v.replace('"', '""') + '"' for v in vals]
            lines.append(",".join(safe))
        csv_text = "\n".join(lines)

    added   = 0
    skipped = 0
    errors  = []

    async with AsyncSessionLocal() as db:
        lines      = csv_text.strip().splitlines()
        csv_reader = _csv.DictReader(lines)

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                cat_val = row.get("category", "").strip()
                name    = row.get("name", "").strip()

                if not cat_val or not name:
                    errors.append(f"Qator {row_num}: category yoki name bo'sh — o'tkazildi")
                    continue
                if cat_val not in valid_cats:
                    errors.append(f"Qator {row_num}: noto'g'ri category '{cat_val}'")
                    continue

                cat      = ProductCategory(cat_val)
                razmer   = _normalize_razmer(row.get("razmer", "").strip() or None)
                rang     = row.get("rang",    "").strip() or None
                tur      = row.get("tur",     "").strip() or None
                qism     = (row.get("qism", "") or "").strip().lower() or None
                birlik   = row.get("birlik",  "dona").strip() or "dona"
                # gramaj — qalinlik maydoniga
                _gram_raw = (row.get("gramaj", "") or row.get("qalinlik", "") or "").replace(",", ".").strip()
                try:
                    qalinlik = float(_gram_raw) if _gram_raw else None
                except ValueError:
                    qalinlik = None
                miqdor   = float((row.get("miqdor",          "0") or "0").replace(",", "."))
                min_t    = float((row.get("min_threshold",   "2") or "2").replace(",", "."))
                yellow_t = float((row.get("yellow_threshold","5") or "5").replace(",", "."))

                # Mavjudligini tekshirish
                q = _sel(WarehouseProduct).where(
                    WarehouseProduct.category == cat,
                    WarehouseProduct.name     == name,
                    WarehouseProduct.is_active == True,
                )
                if razmer: q = q.where(WarehouseProduct.razmer == razmer)
                if rang:   q = q.where(WarehouseProduct.rang   == rang)
                if tur:    q = q.where(WarehouseProduct.tur    == tur)
                if qism:   q = q.where(WarehouseProduct.qism   == qism)

                existing = (await db.execute(q.limit(1))).scalar_one_or_none()
                if existing:
                    skipped += 1
                    continue

                db.add(WarehouseProduct(
                    category=cat, name=name,
                    razmer=razmer, rang=rang, tur=tur, qism=qism,
                    qalinlik=qalinlik,
                    birlik=birlik, miqdor=miqdor,
                    min_threshold=min_t,
                    yellow_threshold=yellow_t,
                ))
                added += 1

            except Exception as e:
                errors.append(f"Qator {row_num}: {e}")

        await db.commit()

    err_rows = "".join(
        f"<li style='color:#ef9a9a'>{e}</li>" for e in errors[:30]
    )
    err_block = (
        f"<div class='card' style='border-left:4px solid #ef5350'>"
        f"<b>❌ Xatolar ({len(errors)} ta):</b><ul>{err_rows}"
        + ("<li>...va boshqalar</li>" if len(errors) > 30 else "")
        + "</ul></div>"
    ) if errors else ""

    html = f"""<!DOCTYPE html><html><head>
<meta charset="utf-8"><title>Import natijasi</title>
<style>
body{{font-family:'Segoe UI',sans-serif;max-width:700px;margin:40px auto;padding:0 24px;background:#1a1a2e;color:#e0e0e0}}
h2{{color:#66bb6a}}.card{{background:#16213e;border-radius:14px;padding:24px;margin:16px 0}}
.info{{background:#0a2200;border-left:4px solid #66bb6a;padding:14px;border-radius:4px;font-size:15px;line-height:2}}
.btn{{background:#4fc3f7;color:#000;border:none;padding:11px 24px;border-radius:8px;cursor:pointer;font-weight:700;margin-right:10px}}
.btn:hover{{background:#81d4fa}}.btn-g{{background:#66bb6a}}.btn-g:hover{{background:#81c784}}
a{{text-decoration:none}}
</style></head><body>
<h2>✅ Import yakunlandi</h2>
<div class="card">
<div class="info">
✅ Qo'shildi: <b>{added}</b> ta yangi mahsulot<br>
⏭ Mavjud edi (o'tkazildi): <b>{skipped}</b> ta<br>
❌ Xatolar: <b>{len(errors)}</b> ta
</div>
</div>
{err_block}
<a href="/web/warehouse"><button class="btn">← Omborga qaytish</button></a>
<a href="/web/warehouse/csv-import"><button class="btn btn-g">📥 Yana import</button></a>
</body></html>"""
    return web.Response(text=html, content_type="text/html")




# ═══ OMBOR BO'LIMLARI — KATEGORIYA BROWSER ════════════════════════════════════

# Kategoriya konfiguratsiyasi
_CAT_CFG = {
    "rulon": {
        "title": "🌀 Rulonlar",
        "turlar": {"yangi":"🆕 Yangi rulon","oralgan":"🔄 O'ralgan rulon","salafanli":"🎁 Salafanli rulon"},
        "extra_cols": [("Gramm","razmer"),("Rang","rang"),("Hajm m³","qalinlik")],
    },
    "gofra": {
        "title": "📋 Gofralar",
        "turlar": {"Yirik":"🔵 Yirik gofra","Mayin":"🟢 Mayin gofra"},
        "extra_cols": [],
    },
    "gofra_zagatovka": {
        "title": "✂️ Zagatovka gofralar",
        "turlar": {"adyol":"🛏 Adyol","pastel":"💼 Pastel","poyabzal":"👟 Poyabzal","shirinlik":"🍰 Shirinlik","fast_food":"🍔 Fast food","boshqa":"📝 Boshqa"},
        "extra_cols": [("Razmer","razmer"),("Sloy","sloy")],
    },
    "xromazes": {
        "title": "🖨️ Xromazeslar",
        "turlar": {"adyol":"🛏 Adyol","pastel":"💼 Pastel","poyabzal":"👟 Poyabzal","shirinlik":"🍰 Shirinlik","fast_food":"🍔 Fast food","boshqa":"📝 Boshqa"},
        "extra_cols": [("Razmer","razmer"),("Rang","rang")],
    },
    "laminat_xromazes": {
        "title": "✨ Laminat xromazeslar",
        "turlar": {"adyol":"🛏 Adyol (laminat)","pastel":"💼 Pastel (laminat)","poyabzal":"👟 Poyabzal","shirinlik":"🍰 Shirinlik","fast_food":"🍔 Fast food","boshqa":"📝 Boshqa"},
        "extra_cols": [("Razmer","razmer"),("Rang","rang")],
    },
    "yarim_tayyor": {
        "title": "🧩 Yarim tayyor",
        "turlar": {
            "tiger_uchun":"✂️ Tiger kesish","gofra_kley_zagatovka":"🔨 Gofra kley — zagatovka","gofra_kley_xromazes":"🔨 Gofra kley — xromazes",
            "stepler_uchun":"📌 Stepler","salafan_uchun":"🎁 Salafan",
            "yopish_uchun":"🔗 Yopish",
            "adyol_tikish_uchun":"🧵 Adyol tikish",
            "pastel_tikish_uchun":"💼 Pastel tikish","adyol_qoqish_uchun":"📫 Adyol qoqish",
            "pastel_qoqish_uchun":"📬 Pastel qoqish","xom_komple":"📦 Xom komple",
            "gofra_uchun_rulon":"🌀 Gofra uchun rulon","list_qogoz_uchun_rulon":"📄 List qog'oz uchun rulon",
            "zagatovka_uchun_gofra":"✂️ Zagatovka uchun gofra",
            "kapalak":"🦋 Kapalak","boshqa":"📝 Boshqa",
        },
        "extra_cols": [("Razmer","razmer"),("Rang","rang")],
    },
    "tayyor_mahsulot": {
        "title": "📦 Tayyor mahsulot",
        "turlar": {"adyol":"🛏 Adyol","pastel":"💼 Pastel","diplomat":"🧳 Diplomat","poyabzal":"👟 Poyabzal","shirinlik":"🍰 Shirinlik","fast_food":"🍔 Fast food","tushli":"🍱 Tushli","blok":"🧱 Blok","boshqa":"📝 Boshqa"},
        "extra_cols": [("Razmer","razmer"),("Rang","rang")],
    },
    "qolip": {
        "title": "🔲 Qoliplar",
        "turlar": {"fast_food":"🍔 Fast food","tushli":"🍱 Tushlik","shirinlik":"🍰 Shirinlik","blok_4quloqli":"📦 Blok 4 quloqli","adyol_3qism":"🛏 Adyol 3 qism","pastel_3qism":"💼 Pastel 3 qism","boshqa":"📝 Boshqa"},
        "extra_cols": [("Razmer","razmer"),("Holat","holat")],
        "has_holat": True,
    },
    "adyol_zapchast": {
        "title": "🧩 Adyol zapchastlari",
        "turlar": {"quluf":"🔒 Qulflar","ruchka":"🖐 Ruchka","piston":"⚙️ Pistonlar","ip_tesma":"🧵 IP/Tesma","boshqa":"📝 Boshqa"},
        "extra_cols": [("Rang","rang")],
    },
    "uskuna_zapchast": {
        "title": "🔧 Stanok ehtiyot qismlari",
        "turlar": {"motor":"⚡ Motor","tasma_kamar":"🔄 Tasma/Kamar","podshipnik":"🔩 Podshipnik","boshqa":"📝 Boshqa"},
        "extra_cols": [],
    },
}

_HOLAT_BADGE = {
    "yaroqli":     '<span style="background:rgba(16,185,129,.15);color:#10b981;padding:1px 7px;border-radius:10px;font-size:11px;font-weight:700">✅ Yaroqli</span>',
    "tamir_talab": '<span style="background:rgba(245,158,11,.15);color:#f59e0b;padding:1px 7px;border-radius:10px;font-size:11px;font-weight:700">🔧 Tamir talab</span>',
    "yaroqsiz":    '<span style="background:rgba(239,68,68,.15);color:#ef4444;padding:1px 7px;border-radius:10px;font-size:11px;font-weight:700">❌ Yaroqsiz</span>',
}


def _adyol_komple(items) -> int:
    """To'liq komple (to'plam) soni.
    Adyol: TEPA + PAST + YON(×2) → min(tepa, past, yon//2).
    Pastel: TEPA + PAST + PADDO → min(tepa, past, paddo).
    """
    bq = {}
    for it in items:
        if it.qism:
            bq[it.qism] = bq.get(it.qism, 0) + float(it.miqdor or 0)
    if not bq:
        return 0
    tepa = bq.get("tepa", 0)
    past = bq.get("past", 0)
    if "yon" in bq:
        return int(min(tepa, past, bq.get("yon", 0) // 2))
    if "paddo" in bq:
        return int(min(tepa, past, bq.get("paddo", 0)))
    return int(min(bq.values()))


QISM_ORD_W = {"tepa": 0, "past": 1, "yon": 2, "paddo": 2}
QISM_LBL_W = {"tepa": "TEPA", "past": "PAST", "yon": "YON", "paddo": "PADDO"}
QISM_ICO_W = {"tepa": "⬆️", "past": "⬇️", "yon": "↔️", "paddo": "🔲"}




def _add_form_fields(cat_key: str, cfg: dict, tur_filter: str = "") -> str:
    """Kategoriyaga qarab mahsulot qo'shish forma maydonlarini generatsiya qilish."""
    fields = ""

    # Tur (barcha kategoriyalar uchun)
    turlar = cfg.get("turlar", {})
    if turlar:
        opts = f'<option value="">— Tanlang —</option>'
        for tk, tl in turlar.items():
            sel = " selected" if tk == tur_filter else ""
            opts += f'<option value="{tk}"{sel}>{tl}</option>'
        fields += f"""
        <div class="fg"><label class="fl">Tur *</label>
        <select name="tur" class="fsel" required>{opts}</select></div>"""

    # Nom (barcha uchun)
    fields += """
    <div class="fg"><label class="fl">Nomi *</label>
    <input type="text" name="name" placeholder="Mahsulot nomi" required></div>"""

    # Kategoriyaga maxsus maydonlar
    extra_cols = cfg.get("extra_cols", [])
    col_attrs  = [c[1] for c in extra_cols]

    if cat_key == "rulon":
        fields += """
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Gramm (g/m²)</label>
          <input type="text" name="razmer" placeholder="80gr"></div>
          <div class="fg"><label class="fl">Rang</label>
          <input type="text" name="rang" placeholder="Oq, Ko'k..."></div>
          <div class="fg"><label class="fl">Hajm (m³)</label>
          <input type="number" step="0.01" name="qalinlik" placeholder="1.5"></div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Miqdor</label>
          <input type="number" step="0.1" name="miqdor" value="0" min="0"></div>
          <div class="fg"><label class="fl">Birlik</label>
          <select name="birlik" class="fsel">
            <option value="rulon">Rulon</option><option value="m">Metr</option>
            <option value="kg">Kg</option><option value="dona">Dona</option>
          </select></div>
        </div>"""
    elif cat_key in ("xromazes", "laminat_xromazes", "gofra_zagatovka"):
        fields += """
        <div class="info" style="font-size:11px;margin-bottom:8px">
          📐 <b>Aniq razmer</b> (40x60, 98x62.5) — sinxronizatsiya uchun<br>
          📦 <b>O'lcham</b> (Katta/O'rta/Kichik) — tiger narxi uchun
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Aniq razmer (40x60, 98×62.5) *</label>
          <input type="text" name="razmer" placeholder="98x62.5" required></div>
          <div class="fg"><label class="fl">O'lcham (tiger narxi uchun)</label>
          <select name="razmer_tur" class="fsel">
            <option value="">— Tanlang —</option>
            <option value="Katta">Katta</option>
            <option value="O'rta">O'rta</option>
            <option value="Kichik">Kichik</option>
          </select></div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Rang</label>
          <input type="text" name="rang" placeholder="Ixtiyoriy"></div>
          <div class="fg"><label class="fl">Miqdor</label>
          <input type="number" step="1" name="miqdor" value="0" min="0"></div>
        </div>
        <div class="fg"><label class="fl">Birlik</label>
        <select name="birlik" class="fsel">
          <option value="dona">Dona</option><option value="top">Top</option>
        </select></div>"""
    elif cat_key == "qolip":
        fields += """
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Razmer (masalan 40x40x60)</label>
          <input type="text" name="razmer" placeholder="40x40x60"></div>
          <div class="fg"><label class="fl">Holat</label>
          <select name="holat" class="fsel">
            <option value="yaroqli">✅ Yaroqli</option>
            <option value="tamir_talab">🔧 Tamir talab</option>
            <option value="yaroqsiz">❌ Yaroqsiz</option>
          </select></div>
        </div>
        <div class="fg"><label class="fl">Holat izohi (tamir uchun)</label>
        <input type="text" name="holat_izoh" placeholder="Nima buzilgan?"></div>"""
    elif cat_key in ("adyol_zapchast", "uskuna_zapchast"):
        fields += """
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Rang</label>
          <input type="text" name="rang" placeholder="Qora, Oq..."></div>
          <div class="fg"><label class="fl">Miqdor</label>
          <input type="number" step="1" name="miqdor" value="0" min="0"></div>
          <div class="fg"><label class="fl">Birlik</label>
          <select name="birlik" class="fsel">
            <option value="dona">Dona</option><option value="qop">Qop</option>
            <option value="pachka">Pachka</option><option value="kg">Kg</option>
          </select></div>
        </div>"""
    elif cat_key == "yarim_tayyor":
        # Adyol/pastel turlari uchun qism tanlash
        fields += """
        <div class="fg"><label class="fl">Qism (adyol/pastel uchun)</label>
        <select name="qism" class="fsel">
          <option value="">— Qism yo'q —</option>
          <option value="tepa">⬆️ Tepa qism</option>
          <option value="past">⬇️ Past qism</option>
          <option value="yon">↔️ Yon qism (×2)</option>
          <option value="paddo">🔲 Paddo (ichki qism)</option>
        </select></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Razmer</label>
          <input type="text" name="razmer" placeholder="Katta / 40x60 / ..."></div>
          <div class="fg"><label class="fl">Rang</label>
          <input type="text" name="rang" placeholder="Ixtiyoriy"></div>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Miqdor</label>
          <input type="number" step="1" name="miqdor" value="0" min="0"></div>
          <div class="fg"><label class="fl">Birlik</label>
          <select name="birlik" class="fsel">
            <option value="dona">Dona</option><option value="top">Top</option>
          </select></div>
        </div>"""
    else:
        fields += """
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px">
          <div class="fg"><label class="fl">Razmer</label>
          <input type="text" name="razmer" placeholder="Ixtiyoriy"></div>
          <div class="fg"><label class="fl">Miqdor</label>
          <input type="number" step="1" name="miqdor" value="0" min="0"></div>
          <div class="fg"><label class="fl">Birlik</label>
          <select name="birlik" class="fsel">
            <option value="dona">Dona</option><option value="kg">Kg</option>
            <option value="m">Metr</option><option value="top">Top</option>
          </select></div>
        </div>"""

    # Min/Yellow threshold
    fields += """
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
      <div class="fg"><label class="fl">🔴 Min chegara</label>
      <input type="number" name="min_threshold" value="2" min="0"></div>
      <div class="fg"><label class="fl">🟡 Ogohlantirish</label>
      <input type="number" name="yellow_threshold" value="5" min="0"></div>
    </div>"""

    # Qismli mahsulot tanlovi (omborchi panel bilan bir xil) — barcha bo'limlarda
    if True:
        fields += """
    <div class="fg"><label class="fl">Necha qismli?</label>
    <select name="kind" id="adm-kind" onchange="admKindUI()">
      <option value="oddiy">1 — Oddiy (bitta mahsulot)</option>
      <option value="ikki">2 — TEPA + PAST (tort, shirinlik, qozon...)</option>
      <option value="adyol">3 — Adyol (TEPA, PAST, YON ×2)</option>
      <option value="pastel">3 — Pastel (TEPA, PAST, PADDO)</option>
    </select></div>
    <div id="adm-kind-info" style="display:none;background:rgba(99,102,241,.12);border:1px solid rgba(99,102,241,.35);color:#4F46E5;border-radius:10px;padding:9px 11px;margin:8px 0;font-size:12px;font-weight:600"></div>
    <div id="adm-qism-box" style="display:none">
      <div style="font-size:12px;font-weight:700;color:#4F46E5;margin-bottom:6px">Har qism — razmer va soni:</div>
      <div style="display:flex;gap:6px;align-items:center;margin-bottom:6px">
        <span style="min-width:48px;font-size:11px;font-weight:800;color:#4F46E5">TEPA</span>
        <input name="razmer_tepa" placeholder="razmer" style="flex:1">
        <input name="qism_tepa" type="number" step="any" placeholder="soni" value="0" style="flex:1"></div>
      <div style="display:flex;gap:6px;align-items:center;margin-bottom:6px">
        <span style="min-width:48px;font-size:11px;font-weight:800;color:#4F46E5">PAST</span>
        <input name="razmer_past" placeholder="razmer" style="flex:1">
        <input name="qism_past" type="number" step="any" placeholder="soni" value="0" style="flex:1"></div>
      <div style="display:flex;gap:6px;align-items:center;margin-bottom:6px" id="adm-q3-row">
        <span style="min-width:48px;font-size:11px;font-weight:800;color:#4F46E5" id="adm-q3-lbl">YON</span>
        <input name="razmer_3" placeholder="razmer" style="flex:1">
        <input name="qism_3" type="number" step="any" placeholder="soni" value="0" style="flex:1"></div>
    </div>
    <script>
    function admKindUI(){
      var form=document.getElementById('adm-kind').closest('form')||document;
      var kind=document.getElementById('adm-kind').value;
      var box=document.getElementById('adm-qism-box');
      var info=document.getElementById('adm-kind-info');
      var q3=document.getElementById('adm-q3-row');
      var miq=form.querySelector('[name=miqdor]');
      var qsel=form.querySelector('select[name=qism]');
      if(kind==='oddiy'){box.style.display='none';info.style.display='none';
        if(miq)miq.closest('.fg').style.display='';if(qsel)qsel.closest('.fg').style.display='';return;}
      box.style.display='';info.style.display='';
      if(miq)miq.closest('.fg').style.display='none';if(qsel)qsel.closest('.fg').style.display='none';
      if(kind==='ikki'){q3.style.display='none';
        info.textContent='📦 2 qism (TEPA, PAST) — tort/shirinlik/qozon karobka';}
      else{q3.style.display='flex';
        info.textContent=(kind==='pastel')?'💼 Pastel — 3 qism (TEPA, PAST, PADDO)':'🛏 Adyol — 3 qism (TEPA, PAST, YON)';
        document.getElementById('adm-q3-lbl').textContent=(kind==='pastel')?'PADDO':'YON';}
    }
    (function(){
      var form=document.getElementById('adm-kind').closest('form')||document;
      var turSel=form.querySelector('[name=tur]');
      function auto(){
        var t=(turSel&&turSel.value)||'';
        var kind='oddiy';
        if(t.indexOf('adyol')===0)kind='adyol';
        else if(t.indexOf('pastel')===0)kind='pastel';
        document.getElementById('adm-kind').value=kind;
        admKindUI();
      }
      if(turSel){turSel.addEventListener('change',auto);}
      auto();
    })();
    </script>"""

    return fields



@_require_auth
@_require_auth
async def tayyor_chiqim(request: web.Request):
    """Tayyor mahsulot CHIQIMI: nima chiqdi, kim chiqardi, qachon, izoh."""
    async with AsyncSessionLocal() as db:
        from sqlalchemy import desc as _desc
        from database.models import WarehouseLog, User as _User
        rows = (await db.execute(
            select(WarehouseLog, WarehouseProduct, _User)
            .join(WarehouseProduct, WarehouseLog.product_id == WarehouseProduct.id)
            .outerjoin(_User, WarehouseLog.user_id == _User.id)
            .where(
                WarehouseProduct.category == ProductCategory.tayyor_mahsulot,
                WarehouseLog.amal == "chiqim",
            )
            .order_by(_desc(WarehouseLog.created_at))
            .limit(300)
        )).all()

        tm_turlar = _CAT_CFG.get("tayyor_mahsulot", {}).get("turlar", {})
        body = ""
        jami = 0.0
        for log, prod, user in rows:
            jami += float(log.miqdor or 0)
            tur_lbl = tm_turlar.get(prod.tur or "", prod.tur or "—")
            dt = log.created_at.strftime("%d.%m.%Y %H:%M") if log.created_at else "—"
            who = h(user.full_name) if user else "—"
            body += f"""
            <tr>
              <td class="t-sm">{dt}</td>
              <td><b>{h(prod.name)}</b>{(' <span class="t-muted t-sm">'+h(prod.razmer)+'</span>') if prod.razmer else ''}</td>
              <td><span class="badge">{tur_lbl}</span></td>
              <td class="td-n cv-red">−{log.miqdor:.0f} {h(prod.birlik or 'dona')}</td>
              <td>👤 {who}</td>
              <td class="t-sm t-muted">{h(log.izoh or '')}</td>
            </tr>"""
        if not rows:
            body = '<tr><td colspan="6" class="empty-state">📭 Hozircha chiqim yo\'q</td></tr>'

        content = f"""
<div class="card-hd mb12">
  <div>
    <h1 style="font-size:19px">📤 Tayyor mahsulot chiqimi</h1>
    <div class="t-sm t-muted">Oxirgi 300 ta chiqim · jami −{jami:.0f}</div>
  </div>
  <a href="/web/ombor/tayyor_mahsulot" class="btn btn-s">📦 Tayyor mahsulot bo'limi</a>
</div>
<div class="card"><div class="tbl-wrap">
<table>
  <tr><th>Sana</th><th>Mahsulot</th><th>Tur</th><th>Miqdor</th><th>Kim chiqardi</th><th>Izoh</th></tr>
  {body}
</table>
</div></div>"""
        return web.Response(text=_base("Tayyor chiqimi", "tayyor-chiqim", content),
                            content_type="text/html")


async def ombor_cats(request: web.Request):
    """Barcha ombor bo'limlari overview."""
    async with AsyncSessionLocal() as db:
        from sqlalchemy import select, func as sf
        rows_html = ""
        total_kam = 0

        for cat_key, cfg in _CAT_CFG.items():
            cat_enum = ProductCategory(cat_key)
            r = await db.execute(
                select(
                    sf.count(WarehouseProduct.id),
                    sf.sum(
                        sa_case(
                            (WarehouseProduct.miqdor <= WarehouseProduct.min_threshold, 1),
                            else_=0,
                        )
                    ),
                    sf.sum(
                        sa_case(
                            (WarehouseProduct.miqdor <= WarehouseProduct.yellow_threshold, 1),
                            else_=0,
                        )
                    ),
                ).where(
                    WarehouseProduct.category == cat_enum,
                    WarehouseProduct.is_active == True,
                )
            )
            row = r.one()
            cnt, kam, ogoh = row[0] or 0, row[1] or 0, row[2] or 0
            total_kam += kam

            status_html = ""
            if kam:   status_html += f'<span class="badge red ml4">🔴 {kam} ta</span>'
            elif ogoh: status_html += f'<span class="badge yellow ml4">🟡 {ogoh} ta</span>'
            else:      status_html += '<span class="badge green ml4">🟢 OK</span>'

            rows_html += f"""
            <tr>
              <td><a href="/web/ombor/{cat_key}" style="color:var(--accent);font-weight:700;text-decoration:none">
                {cfg["title"]}
              </a></td>
              <td style="font-family:monospace;font-weight:700">{cnt}</td>
              <td>{len(cfg["turlar"])} ta</td>
              <td>{status_html}</td>
              <td>
                <a href="/web/ombor/{cat_key}" class="btn btn-sm btn-blue">Ko'rish</a>
                <a href="/web/warehouse?cat={cat_key}" class="btn btn-sm btn-outline">+ Kirim</a>
              </td>
            </tr>"""

        warn_bar = ""
        if total_kam:
            warn_bar = f'<div class="alert alert-yellow mb12">⚠️ Jami <b>{total_kam}</b> ta mahsulot kritik darajada kam!</div>'

        content = f"""
{warn_bar}
<div class="card">
  <div class="card-hd">
    <span class="card-title">🏭 Barcha ombor bo'limlari</span>
    <a href="/web/warehouse/csv-import" class="btn btn-blue">📥 CSV Import</a>
  </div>
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Bo'lim</th><th>Mahsulotlar</th><th>Turlar</th><th>Holat</th><th>Amallar</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>"""

    return web.Response(
        text=_base("Ombor bo'limlari", "ombor-cats", content),
        content_type="text/html",
    )


@_require_auth
async def ombor_cat_detail(request: web.Request):
    """Alohida kategoriya sahifasi — tur tablar va mahsulotlar ro'yxati."""
    cat_key = request.match_info.get("cat_key", "rulon")
    cfg = _CAT_CFG.get(cat_key)
    if not cfg:
        raise web.HTTPNotFound()

    tur_filter   = request.rel_url.query.get("tur", "")
    holat_filter = request.rel_url.query.get("holat", "")
    search_q     = request.rel_url.query.get("q", "")
    page         = max(0, int(request.rel_url.query.get("page", "0")))
    per_page     = 25

    async with AsyncSessionLocal() as db:
        from sqlalchemy import select, func as sf, or_

        cat_enum = ProductCategory(cat_key)
        q = select(WarehouseProduct).where(
            WarehouseProduct.category == cat_enum,
            WarehouseProduct.is_active == True,
        )

        if tur_filter:
            q = q.where(WarehouseProduct.tur == tur_filter)
        if holat_filter and cfg.get("has_holat"):
            try:
                from database.models import ProductHolat
                q = q.where(WarehouseProduct.holat == ProductHolat(holat_filter))
            except ValueError:
                pass
        if search_q:
            q = q.where(or_(
                WarehouseProduct.name.ilike(f"%{search_q}%"),
                WarehouseProduct.razmer.ilike(f"%{search_q}%"),
                WarehouseProduct.rang.ilike(f"%{search_q}%"),
            ))

        total = (await db.execute(
            select(sf.count()).select_from(q.subquery())
        )).scalar() or 0

        products = (await db.execute(
            q.order_by(WarehouseProduct.tur, WarehouseProduct.name, WarehouseProduct.razmer)
            .limit(per_page).offset(page * per_page)
        )).scalars().all()

        # Tur bo'yicha sonlar
        tur_counts = {}
        for tk in cfg["turlar"]:
            r = await db.execute(
                select(sf.count(WarehouseProduct.id))
                .where(WarehouseProduct.category == cat_enum,
                       WarehouseProduct.is_active == True,
                       WarehouseProduct.tur == tk)
            )
            tur_counts[tk] = r.scalar() or 0

        # Tur tablar
        base_url = f"/web/ombor/{cat_key}"
        q_params = f"&q={search_q}" if search_q else ""
        h_params = f"&holat={holat_filter}" if holat_filter else ""

        tur_tabs_html = f'<a href="{base_url}?{h_params}{q_params}" class="cat-tab{" active" if not tur_filter else ""}">Barchasi ({total if not tur_filter else ""})</a>'
        for tk, tl in cfg["turlar"].items():
            cnt = tur_counts.get(tk, 0)
            active = " active" if tur_filter == tk else ""
            tur_tabs_html += f'<a href="{base_url}?tur={tk}{h_params}{q_params}" class="cat-tab{active}">{tl} ({cnt})</a>'

        # Holat tablar (faqat qoliplar uchun)
        holat_tabs_html = ""
        if cfg.get("has_holat"):
            from database.models import ProductHolat
            h_counts = {}
            for hv in ["yaroqli","tamir_talab","yaroqsiz"]:
                r = await db.execute(
                    select(sf.count(WarehouseProduct.id))
                    .where(WarehouseProduct.category == cat_enum,
                           WarehouseProduct.is_active == True,
                           WarehouseProduct.holat == ProductHolat(hv))
                )
                h_counts[hv] = r.scalar() or 0
            t_params = f"&tur={tur_filter}" if tur_filter else ""
            holat_tabs_html = f"""
            <div style="display:flex;gap:5px;flex-wrap:wrap;margin-bottom:10px">
              <a href="{base_url}?{t_params}{q_params}" class="cat-tab{" active" if not holat_filter else ""}">Barchasi</a>
              <a href="{base_url}?holat=yaroqli{t_params}{q_params}" class="cat-tab{" active" if holat_filter=="yaroqli" else ""}">✅ Yaroqli ({h_counts.get("yaroqli",0)})</a>
              <a href="{base_url}?holat=tamir_talab{t_params}{q_params}" class="cat-tab{" active" if holat_filter=="tamir_talab" else ""}" style="{"color:#f59e0b" if holat_filter!="tamir_talab" else ""}">🔧 Tamir talab ({h_counts.get("tamir_talab",0)})</a>
              <a href="{base_url}?holat=yaroqsiz{t_params}{q_params}" class="cat-tab{" active" if holat_filter=="yaroqsiz" else ""}" style="{"color:#ef4444" if holat_filter!="yaroqsiz" else ""}">❌ Yaroqsiz ({h_counts.get("yaroqsiz",0)})</a>
            </div>"""

        # Mahsulotlar jadvali
        extra_cols = cfg.get("extra_cols", [])
        th_extra = "".join(f"<th>{col[0]}</th>" for col in extra_cols)

        def _td_extra(p):
            s = ""
            for _, attr in extra_cols:
                if attr == "holat":
                    hval = p.holat.value if p.holat else None
                    s += f"<td>{_HOLAT_BADGE.get(hval, '<span style=\"color:var(--text2)\">—</span>')}</td>"
                elif attr == "qalinlik":
                    s += f"<td>{p.qalinlik if p.qalinlik else '—'}</td>"
                else:
                    val = getattr(p, attr, None) or "—"
                    s += f"<td>{h(str(val))}</td>"
            return s

        def _status(p):
            m = float(p.miqdor)
            if m <= float(p.min_threshold):      return "🔴", "cv-red"
            elif m <= float(p.yellow_threshold): return "🟡", "cv-yellow"
            return "🟢", ""

        show_kochir = cat_key != "yarim_tayyor"
        def _actions(p):
            mv = (f'<button onclick="admMove({p.id},\'{h(p.name)}\')" class="btn btn-sm" style="background:#0e7490;color:#fff" title="Yarim tayyorga ko\'chirish">→</button>'
                  if show_kochir else "")
            return (
                f'<button onclick="quickKirim({p.id},\'{h(p.name)}\')" class="btn btn-sm btn-green">+</button>'
                f'<button onclick="quickChiqim({p.id},\'{h(p.name)}\')" class="btn btn-sm btn-red">−</button>'
                + mv +
                f'<button onclick="deleteProduct({p.id},\'{h(p.name)}\')" class="btn btn-sm" style="background:#7c2d12;color:#fca5a5">🗑</button>'
            )

        # Guruhlash: qismli (nom,rang,tur) → ramka; qismsiz → oddiy qator
        groups, order = {}, []
        for p in products:
            key = ("V", p.name, p.rang or "", p.tur or "") if p.qism else ("S", p.id)
            if key not in groups:
                groups[key] = []
                order.append(key)
            groups[key].append(p)

        rows_html = ""
        for key in order:
            items = groups[key]
            if key[0] == "V":
                items.sort(key=lambda x: QISM_ORD_W.get(x.qism or "", 9))
                first = items[0]
                komple = _adyol_komple(items)
                tur_lbl = cfg["turlar"].get(first.tur or "", first.tur or "—")
                rows_html += f"""
                <tr style="background:rgba(99,102,241,.08)">
                  <td colspan="99" style="border-left:3px solid #6366f1">
                    🛏 <b>{h(first.name)}</b>{(' | '+h(first.rang)) if first.rang else ''}
                    <span class="badge">{tur_lbl}</span>
                    <span style="background:rgba(16,185,129,.15);color:#10b981;padding:1px 8px;border-radius:10px;font-size:11px;font-weight:700;margin-left:6px">📦 Komple: {komple}</span>
                  </td>
                </tr>"""
                for p in items:
                    st, sc = _status(p)
                    qico = QISM_ICO_W.get(p.qism, "")
                    qlbl = QISM_LBL_W.get(p.qism, (p.qism or "").upper())
                    rows_html += f"""
                    <tr style="border-left:3px solid #6366f1">
                      <td style="padding-left:18px">
                        <span style="background:rgba(99,102,241,.15);color:#4F46E5;padding:1px 7px;border-radius:8px;font-size:11px;font-weight:700">{qico} {qlbl}</span>
                        <span style="color:var(--text2);font-size:12px;margin-left:4px">{h(p.razmer or '')}</span>
                      </td>
                      <td></td>
                      {_td_extra(p)}
                      <td>
                        <span class="{sc}" style="font-family:monospace;font-weight:700">{st} {p.miqdor:.0f}</span>
                        <span style="color:var(--text2);font-size:11px">{p.birlik}</span>
                      </td>
                      <td>{_actions(p)}</td>
                    </tr>"""
            else:
                p = items[0]
                st, sc = _status(p)
                tur_lbl = cfg["turlar"].get(p.tur or "", p.tur or "—")
                rows_html += f"""
                <tr>
                  <td><b>{h(p.name)}</b>
                    {f'<br><span style="font-size:10px;color:var(--text2)">{h(p.holat_izoh)}</span>' if p.holat_izoh else ""}
                  </td>
                  <td><span class="badge">{tur_lbl}</span></td>
                  {_td_extra(p)}
                  <td>
                    <span class="{sc}" style="font-family:monospace;font-weight:700">{st} {p.miqdor:.0f}</span>
                    <span style="color:var(--text2);font-size:11px">{p.birlik}</span>
                  </td>
                  <td>{_actions(p)}</td>
                </tr>"""

        if not products:
            rows_html = '<tr><td colspan="99" style="text-align:center;color:var(--text2);padding:24px">📭 Mahsulot topilmadi</td></tr>'

        # Pagination
        pages_total = max(1, (total - 1) // per_page + 1)
        pg_html = ""
        if pages_total > 1:
            t_p = f"&tur={tur_filter}" if tur_filter else ""
            h_p = f"&holat={holat_filter}" if holat_filter else ""
            q_p = f"&q={search_q}" if search_q else ""
            pg_html = '<div style="display:flex;gap:5px;justify-content:center;margin-top:12px;flex-wrap:wrap">'
            for i in range(pages_total):
                act = ' style="background:var(--accent);color:#fff"' if i == page else ""
                pg_html += f'<a href="{base_url}?page={i}{t_p}{h_p}{q_p}" class="btn btn-sm btn-outline"{act}>{i+1}</a>'
            pg_html += "</div>"

        # Yarim tayyor turlari — ko'chirish modali uchun
        _yt_turlar = _CAT_CFG.get("yarim_tayyor", {}).get("turlar", {})
        yt_opts = "".join(f'<option value="{h(k)}">{h(v)}</option>' for k, v in _yt_turlar.items())

        content = f"""
<div class="card-hd mb12">
  <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
    <a href="/web/ombor" style="color:var(--text2);text-decoration:none;font-size:12px">← Bo'limlar</a>
    <h2 style="font-size:15px;font-weight:800;margin:0">{cfg["title"]}</h2>
    <span style="color:var(--text2);font-size:12px">Jami: {total} ta</span>
  </div>
</div>

<div class="card">
  <!-- Qidiruv + Qo'shish -->
  <div style="display:flex;gap:8px;align-items:flex-start;flex-wrap:wrap;margin-bottom:10px">
    <form method="get" class="toolbar" style="flex:1;margin-bottom:0">
      <input type="hidden" name="tur" value="{tur_filter}">
      <input type="hidden" name="holat" value="{holat_filter}">
      <div class="search-wrap">
        <span class="si">🔍</span>
        <input type="text" name="q" value="{h(search_q)}" placeholder="Nom, razmer, rang...">
      </div>
      <button type="submit" class="btn btn-blue">Qidirish</button>
      {f'<a href="{base_url}" class="btn btn-outline">Tozalash</a>' if search_q or tur_filter or holat_filter else ""}
    </form>
    <button onclick="openModal('addModal')" class="btn btn-green" style="white-space:nowrap">➕ Mahsulot qo'shish</button>
    <a href="/web/warehouse/csv-import" class="btn btn-outline" style="white-space:nowrap">📥 CSV</a>
  </div>

  <!-- Yangi mahsulot modal -->
  <div id="addModal" class="overlay">
    <div class="modal" style="max-width:520px">
      <div class="modal-hd">
        <span class="modal-title">➕ Yangi mahsulot — {cfg["title"]}</span>
        <button onclick="closeModal('addModal')" class="btn-close">✕</button>
      </div>
      <form method="post" action="/web/ombor/{cat_key}/add">
        <input type="hidden" name="cat_key" value="{cat_key}">
        {_add_form_fields(cat_key, cfg, tur_filter)}
        <div style="display:flex;gap:8px;margin-top:12px">
          <button type="submit" class="btn btn-green" style="flex:1">✅ Saqlash</button>
          <button type="button" onclick="closeModal('addModal')" class="btn btn-outline">Bekor</button>
        </div>
      </form>
    </div>
  </div>

  <!-- Tur tablar -->
  <div class="cat-tabs" style="margin-bottom:8px">{tur_tabs_html}</div>

  <!-- Holat tablar (qoliplar uchun) -->
  {holat_tabs_html}

  <!-- Jadval -->
  <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Nomi</th><th>Tur</th>
        {th_extra}
        <th>Qoldiq</th><th>Amallar</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
  {pg_html}
</div>

<div id="adm-mv-overlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:300;align-items:center;justify-content:center;padding:16px" onclick="if(event.target===this)this.style.display='none'">
  <div style="background:var(--card,#1a1a2e);border:1px solid var(--border,#333);border-radius:16px;padding:20px;width:100%;max-width:380px" onclick="event.stopPropagation()">
    <div style="font-size:16px;font-weight:800;color:#22d3ee;margin-bottom:4px">→ Yarim tayyorga ko'chirish</div>
    <div id="adm-mv-name" style="font-size:14px;margin-bottom:14px;font-weight:600"></div>
    <label class="fl">Qaysi turga?</label>
    <select id="adm-mv-tur" class="fsel" style="width:100%;margin-bottom:10px">{yt_opts}</select>
    <label class="fl">Miqdor</label>
    <input id="adm-mv-miqdor" type="number" step="any" min="0" placeholder="0" style="width:100%">
    <div id="adm-mv-err" style="color:#f87171;font-size:12px;min-height:16px;margin-top:6px"></div>
    <div style="display:flex;gap:10px;margin-top:14px">
      <button onclick="document.getElementById('adm-mv-overlay').style.display='none'" class="btn btn-outline" style="flex:1">Bekor</button>
      <button id="adm-mv-ok" onclick="doAdmMove()" class="btn" style="flex:2;background:#0e7490;color:#fff">Ko'chirish</button>
    </div>
  </div>
</div>

<script>

function deleteProduct(id, nom) {{
  if (!confirm(`Haqiqatan ham "${{nom}}" mahsulotini butunlay o'chirmoqchimisiz?\n\nBu amalni qaytarib bo'lmaydi!`)) return;
  fetch('/web/warehouse/delete', {{
    method:'POST', headers:{{"Content-Type":"application/json"}},
    body: JSON.stringify({{product_id: id}})
  }}).then(r=>r.json()).then(d=>{{
    if(d.ok) {{ alert(`✅ "${{nom}}" o'chirildi!`); location.reload() }}
    else alert("❌ Xato: " + (d.error||""))
  }})
}}

function quickKirim(id, nom) {{
  const m = prompt(`+Kirim: ${{nom}}\nMiqdor kiriting:`)
  if (!m || isNaN(m) || +m <= 0) return
  fetch('/web/warehouse/kirim', {{
    method:'POST', headers:{{"Content-Type":"application/json"}},
    body: JSON.stringify({{product_id: id, miqdor: +m}})
  }}).then(r=>r.json()).then(d=>{{
    if(d.ok) {{ alert(`✅ Kirim qo'shildi! Yangi qoldiq: ${{d.new_miqdor}}`); location.reload() }}
    else alert("❌ Xato: " + (d.error||""))
  }})
}}
function quickChiqim(id, nom) {{
  const m = prompt(`−Chiqim: ${{nom}}\nMiqdor kiriting:`)
  if (!m || isNaN(m) || +m <= 0) return
  fetch('/web/warehouse/chiqim', {{
    method:'POST', headers:{{"Content-Type":"application/json"}},
    body: JSON.stringify({{product_id: id, miqdor: +m}})
  }}).then(r=>r.json()).then(d=>{{
    if(d.ok) {{ alert(`✅ Chiqim qo'shildi! Yangi qoldiq: ${{d.new_miqdor}}`); location.reload() }}
    else alert("❌ Xato: " + (d.error||""))
  }})
}}
var admMvId=null;
function admMove(id, nom) {{
  admMvId=id;
  document.getElementById('adm-mv-name').textContent=nom;
  document.getElementById('adm-mv-miqdor').value='';
  document.getElementById('adm-mv-err').textContent='';
  document.getElementById('adm-mv-overlay').style.display='flex';
}}
function doAdmMove() {{
  var tur=document.getElementById('adm-mv-tur').value;
  var miq=parseFloat(document.getElementById('adm-mv-miqdor').value);
  var err=document.getElementById('adm-mv-err');
  if(!miq||miq<=0){{err.textContent='Miqdor kiriting';return;}}
  var btn=document.getElementById('adm-mv-ok');btn.disabled=true;
  fetch('/web/ombor/quick-transfer',{{
    method:'POST',headers:{{"Content-Type":"application/json"}},
    body:JSON.stringify({{product_id:admMvId,dst_tur:tur,miqdor:miq}})
  }}).then(r=>r.json()).then(d=>{{
    btn.disabled=false;
    if(d.ok){{ document.getElementById('adm-mv-overlay').style.display='none'; location.reload(); }}
    else err.textContent=(d.error||'Xato');
  }}).catch(e=>{{btn.disabled=false;err.textContent='Tarmoq xatosi';}});
}}
</script>"""

    return web.Response(
        text=_base(cfg["title"], cat_key, content),
        content_type="text/html",
    )





@_require_auth
@_require_auth
async def ombor_admin_quick_transfer(request: web.Request):
    """Admin panel: '→ Ko'chir' — asosiy ombordan yarim tayyorga ko'chiradi."""
    sess = _current(request)
    uid = sess.get("user_id") if sess else None
    try:
        body = await request.json()
        pid = int(body.get("product_id"))
        dst_tur = (body.get("dst_tur") or "").strip()
        miqdor = float(body.get("miqdor") or 0)
    except (ValueError, TypeError, Exception):
        return web.json_response({"ok": False, "error": "Noto'g'ri ma'lumot"}, status=400)
    if miqdor <= 0:
        return web.json_response({"ok": False, "error": "Miqdor 0 dan katta bo'lsin"})
    if not dst_tur:
        return web.json_response({"ok": False, "error": "Tur tanlanmadi"})

    from database.queries import update_product_miqdor as _upd
    async with AsyncSessionLocal() as db:
        from sqlalchemy import select as _sel
        src = await db.get(WarehouseProduct, pid)
        if not src:
            return web.json_response({"ok": False, "error": "Mahsulot topilmadi"})
        if float(src.miqdor or 0) < miqdor:
            return web.json_response({"ok": False, "error": f"Yetarli emas (mavjud: {float(src.miqdor or 0):.0f})"})
        await _upd(db, pid, -miqdor, uid, f"Ko'chirildi → yarim_tayyor/{dst_tur}")
        dst = (await db.execute(_sel(WarehouseProduct).where(
            WarehouseProduct.is_active == True,
            WarehouseProduct.category == ProductCategory.yarim_tayyor,
            WarehouseProduct.name == src.name,
            WarehouseProduct.razmer == src.razmer,
            WarehouseProduct.rang == src.rang,
            WarehouseProduct.qism == src.qism,
            WarehouseProduct.tur == dst_tur,
        ).limit(1))).scalar_one_or_none()
        if dst:
            await _upd(db, dst.id, miqdor, uid, f"Ko'chirildi ← {src.category.value if src.category else '?'}")
        else:
            db.add(WarehouseProduct(
                category=ProductCategory.yarim_tayyor, name=src.name, razmer=src.razmer,
                rang=src.rang, tur=dst_tur, qism=src.qism, birlik=src.birlik, miqdor=miqdor,
                min_threshold=src.min_threshold, yellow_threshold=src.yellow_threshold,
                qalinlik=getattr(src, "qalinlik", None), is_active=True,
            ))
        await db.commit()
        src2 = await db.get(WarehouseProduct, pid)
        return web.json_response({"ok": True, "new_miqdor": float(src2.miqdor or 0)})


async def ombor_cat_add(request: web.Request):
    """Kategoriyaga yangi mahsulot qo'shish."""
    cat_key = request.match_info.get("cat_key", "")
    cfg = _CAT_CFG.get(cat_key)
    if not cfg:
        raise web.HTTPNotFound()

    data = await request.post()

    def _f(k, default=""):
        v = data.get(k, default)
        return str(v).strip() if v else default

    from utils.razmer import normalize_razmer as _norm

    async with AsyncSessionLocal() as db:
        from database.models import ProductHolat
        cat_enum = ProductCategory(cat_key)

        holat = None
        if cfg.get("has_holat"):
            try:
                holat = ProductHolat(_f("holat", "yaroqli"))
            except ValueError:
                holat = None

        razmer_raw = _f("razmer") or None
        razmer_norm = _norm(razmer_raw) if razmer_raw else None

        try:
            miqdor = float(_f("miqdor", "0").replace(",",".") or 0)
            min_t  = float(_f("min_threshold", "2").replace(",",".") or 2)
            yel_t  = float(_f("yellow_threshold", "5").replace(",",".") or 5)
            qalinlik_raw = _f("qalinlik")
            qalinlik = float(qalinlik_raw.replace(",",".")) if qalinlik_raw else None
        except ValueError:
            miqdor = min_t = 0; yel_t = 5; qalinlik = None

        # "Necha qismli?" tanlovi (oddiy / ikki / adyol / pastel)
        tur_v = _f("tur")
        kind = _f("kind") or "oddiy"
        if kind not in ("oddiy", "ikki", "adyol", "pastel"):
            kind = "oddiy"

        if kind in ("adyol", "pastel", "ikki"):
            def _fnum(k):
                try:
                    return float((_f(k, "0") or "0").replace(",", "."))
                except ValueError:
                    return 0.0
            qismlar = [
                ("tepa", _f("razmer_tepa") or razmer_raw, _fnum("qism_tepa")),
                ("past", _f("razmer_past") or razmer_raw, _fnum("qism_past")),
            ]
            if kind != "ikki":
                third = "paddo" if kind == "pastel" else "yon"
                qismlar.append((third,  _f("razmer_3") or razmer_raw, _fnum("qism_3")))
            for qism_v, rz, miq in qismlar:
                db.add(WarehouseProduct(
                    category=cat_enum, name=_f("name") or "Noma'lum", tur=tur_v or None,
                    razmer=rz, razmer_normalized=(_norm(rz) if rz else None),
                    razmer_tur=_f("razmer_tur") or None, qism=qism_v,
                    rang=_f("rang") or None, qalinlik=qalinlik, birlik=_f("birlik", "dona"),
                    miqdor=miq, min_threshold=min_t, yellow_threshold=yel_t,
                    holat=holat, holat_izoh=_f("holat_izoh") or None,
                ))
            await db.commit()
            redirect = f"/web/ombor/{cat_key}"
            if tur_v:
                redirect += f"?tur={tur_v}"
            raise web.HTTPFound(redirect)

        product = WarehouseProduct(
            category          = cat_enum,
            name              = _f("name") or "Noma'lum",
            tur               = _f("tur") or None,
            razmer            = razmer_raw,
            razmer_normalized = razmer_norm,
            razmer_tur        = _f("razmer_tur") or None,   # Katta/O'rta/Kichik
            qism              = _f("qism") or None,
            rang              = _f("rang") or None,
            qalinlik          = qalinlik,
            birlik            = _f("birlik", "dona"),
            miqdor            = miqdor,
            min_threshold     = min_t,
            yellow_threshold  = yel_t,
            holat             = holat,
            holat_izoh        = _f("holat_izoh") or None,
        )
        db.add(product)
        await db.commit()

    tur = _f("tur")
    redirect = f"/web/ombor/{cat_key}"
    if tur:
        redirect += f"?tur={tur}"
    raise web.HTTPFound(redirect)





async def token_login(request: web.Request):
    """Bot bergan token orqali kirish: /w/{token}"""
    token = request.match_info.get("token", "")
    if not token:
        raise web.HTTPFound("/web/login")

    async with AsyncSessionLocal() as db:
        from sqlalchemy import select
        r = await db.execute(select(User).where(User.web_token == token, User.is_active == True))
        user = r.scalar_one_or_none()

    if not user:
        return web.Response(
            text="<h2 style='font-family:sans-serif;text-align:center;margin-top:50px'>Havola yaroqsiz yoki eskirgan.<br>Botdan yangi havola oling.</h2>",
            content_type="text/html",
        )

    role = user.role.value if hasattr(user.role, "value") else str(user.role)

    sess_token = _make_token(str(user.id))
    _sessions[sess_token] = {
        "expires": time.time() + _SESSION_TTL,
        "user_id": user.id,
        "role":    role,
        "name":    user.full_name,
    }

    resp = web.HTTPFound(_role_home(role))
    resp.set_cookie("wpsession", sess_token, max_age=_SESSION_TTL, httponly=True, samesite="Lax")
    raise resp


async def web_login(request: web.Request):
    """Login sahifasi."""
    error = ""
    if request.method == "POST":
        form     = await request.post()
        password = form.get("password", "")
        if password == WEB_PASSWORD:
            token = _make_token(request.remote)
            _sessions[token] = {"expires": time.time() + _SESSION_TTL}
            resp  = web.HTTPFound("/web/")
            resp.set_cookie("wpsession", token, httponly=True, max_age=_SESSION_TTL)
            raise resp
        else:
            error = "Noto'g'ri parol!"

    html = f"""<!DOCTYPE html>
<html lang="uz">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Quti Tsexi — Kirish</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0f172a;display:flex;align-items:center;justify-content:center;min-height:100vh;font-family:system-ui}}
.card{{background:#1e293b;border-radius:16px;padding:40px;width:340px;box-shadow:0 25px 50px rgba(0,0,0,.5)}}
h1{{color:#f1f5f9;font-size:22px;margin-bottom:8px;text-align:center}}
p{{color:#94a3b8;font-size:13px;text-align:center;margin-bottom:28px}}
label{{color:#535C78;font-size:12px;font-weight:600;display:block;margin-bottom:6px}}
input{{width:100%;background:#0f172a;border:1px solid #334155;border-radius:8px;
       padding:10px 14px;color:#f1f5f9;font-size:14px;outline:none}}
input:focus{{border-color:#6366f1}}
.btn{{width:100%;background:#6366f1;color:#fff;border:none;border-radius:8px;
      padding:11px;font-size:14px;font-weight:700;cursor:pointer;margin-top:16px}}
.btn:hover{{background:#4f46e5}}
.err{{color:#f87171;font-size:12px;text-align:center;margin-top:10px}}
.logo{{text-align:center;font-size:32px;margin-bottom:16px}}
</style>
</head>
<body>
<div class="card">
  <div class="logo">📦</div>
  <h1>Quti Tsexi</h1>
  <p>Admin panel</p>
  <form method="post">
    <label>Parol</label>
    <input type="password" name="password" placeholder="••••••••" autofocus>
    <button class="btn" type="submit">Kirish</button>
    {f'<p class="err">{error}</p>' if error else ""}
  </form>
</div>
</body>
</html>"""
    return web.Response(text=html, content_type="text/html")


async def web_logout(request: web.Request):
    token = request.cookies.get("wpsession")
    if token:
        _sessions.pop(token, None)
    resp = web.HTTPFound("/web/login")
    resp.del_cookie("wpsession")
    raise resp



# ─── APP SETUP ─────────────────────────────────────────────────────────────


# ════════════════════════════════════════════════════════════════════════════
# YANGI SAHIFALAR
# ════════════════════════════════════════════════════════════════════════════

@_require_auth
async def stats_advanced(request: web.Request):
    """Kengaytirilgan statistika sahifasi."""
    async with AsyncSessionLocal() as db:
        from datetime import datetime, timedelta
        from sqlalchemy import func as sf

        today = datetime.now().date()
        days_data = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            r = await db.execute(
                select(
                    sf.count(WorkEntry.id),
                    sf.coalesce(sf.sum(WorkEntry.jami_summa), 0),
                ).where(
                    sf.date(WorkEntry.created_at) == d,
                    WorkEntry.status == WorkStatus.approved,
                )
            )
            cnt, summ = r.one()
            days_data.append({"kun": d.strftime("%d.%m"), "soni": cnt or 0, "summa": float(summ or 0)})

        max_soni = max((d["soni"] for d in days_data), default=1) or 1

        oy_boshi = today.replace(day=1)
        r = await db.execute(
            select(
                User.id, User.full_name,
                sf.count(WorkEntry.id).label("cnt"),
                sf.coalesce(sf.sum(WorkEntry.jami_summa), 0).label("summa"),
            ).join(WorkEntry, WorkEntry.worker_id == User.id)
             .where(WorkEntry.created_at >= oy_boshi, WorkEntry.status == WorkStatus.approved)
             .group_by(User.id, User.full_name)
             .order_by(sf.sum(WorkEntry.jami_summa).desc())
             .limit(10)
        )
        rating = r.all()

    chart_html = '<div style="display:flex;gap:8px;align-items:flex-end;height:200px;padding:16px;background:#0f172a;border-radius:10px">'
    for d in days_data:
        bh = int(d["soni"] / max_soni * 180) if max_soni else 0
        chart_html += f'<div style="flex:1;display:flex;flex-direction:column;align-items:center;gap:4px"><div style="font-size:10px;color:#94a3b8">{d["soni"]}</div><div style="width:100%;height:{bh}px;background:linear-gradient(180deg,#6366f1,#4f46e5);border-radius:6px 6px 0 0;min-height:4px"></div><div style="font-size:10px;color:#64748b;margin-top:4px">{d["kun"]}</div></div>'
    chart_html += '</div>'

    medals = {0:"🥇",1:"🥈",2:"🥉"}
    rating_html = '<table style="width:100%;border-collapse:collapse"><thead><tr><th style="text-align:left;padding:8px;color:#94a3b8;font-size:11px">#</th><th style="text-align:left;padding:8px;color:#94a3b8;font-size:11px">ISHCHI</th><th style="text-align:right;padding:8px;color:#94a3b8;font-size:11px">ISHLAR</th><th style="text-align:right;padding:8px;color:#94a3b8;font-size:11px">SUMMA</th></tr></thead><tbody>'
    for i, (uid, name, cnt, summa) in enumerate(rating):
        medal = medals.get(i, f"#{i+1}")
        rating_html += f'<tr style="border-top:1px solid rgba(99,102,241,.1)"><td style="padding:8px;font-weight:700">{medal}</td><td style="padding:8px">{h(name)}</td><td style="padding:8px;text-align:right">{cnt}</td><td style="padding:8px;text-align:right;color:#34d399;font-weight:700">{summa:,.0f}</td></tr>'
    rating_html += '</tbody></table>'

    content = f'<h1>📊 Kengaytirilgan statistika</h1><div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:20px"><div class="card"><h3>📈 So\'nggi 7 kun</h3>{chart_html}</div><div class="card"><h3>🏆 Oylik reyting</h3>{rating_html if rating else "<p>Ma\'lumot yo\'q</p>"}</div></div>'
    return web.Response(text=_base("Statistika", "stats_advanced", content), content_type="text/html")


@_require_auth
async def material_flow(request: web.Request):
    """Material oqimi sahifasi."""
    async with AsyncSessionLocal() as db:
        from sqlalchemy import func as sf
        cats_data = []
        for cat in ProductCategory:
            r = await db.execute(
                select(
                    sf.count(WarehouseProduct.id),
                    sf.coalesce(sf.sum(WarehouseProduct.miqdor), 0),
                    sf.sum(sa_case(
                        (WarehouseProduct.miqdor <= WarehouseProduct.min_threshold, 1),
                        else_=0,
                    )),
                ).where(WarehouseProduct.category == cat, WarehouseProduct.is_active == True)
            )
            cnt, jami, warn = r.one()
            cats_data.append({"name": cat.value, "cnt": cnt or 0, "jami": float(jami or 0), "warn": int(warn or 0)})

    icons = {"rulon":"🌀","gofra":"📋","gofra_zagatovka":"✂️","xromazes":"🖨️","laminat_xromazes":"✨","yarim_tayyor":"🧩","tayyor_mahsulot":"📦","qolip":"🔲","adyol_zapchast":"🧵","uskuna_zapchast":"🔧"}

    cards_html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-top:20px">'
    for c in cats_data:
        icon = icons.get(c["name"], "📦")
        warn_badge = f'<span style="background:rgba(239,68,68,.15);color:#f87171;padding:2px 8px;border-radius:8px;font-size:10px;font-weight:700">⚠️ {c["warn"]} kam</span>' if c["warn"] > 0 else ''
        cards_html += f'<div class="card" style="cursor:pointer" onclick="location.href=\'/web/ombor/{c["name"]}\'"><div style="display:flex;justify-content:space-between"><div style="font-size:36px">{icon}</div>{warn_badge}</div><div style="font-size:14px;font-weight:700;color:#f1f5f9;margin-top:10px;text-transform:capitalize">{c["name"].replace("_"," ")}</div><div style="display:flex;justify-content:space-between;margin-top:10px;font-size:11px;color:#94a3b8"><span>{c["cnt"]} ta</span><span style="color:#34d399;font-weight:700">{c["jami"]:,.0f}</span></div></div>'
    cards_html += '</div>'

    content = f'<h1>🔄 Material oqimi</h1><p style="color:#94a3b8;font-size:13px">Barcha kategoriyalar</p>{cards_html}'
    return web.Response(text=_base("Material oqimi", "material_flow", content), content_type="text/html")


@_require_auth
async def help_page(request: web.Request):
    """Yordam sahifasi."""
    sections = [
        ("🏠 Boshlash", [("Dashboard","Asosiy ko'rsatkichlar"),("Statistika","Grafiklar va reyting"),("Material oqimi","Kategoriyalar")]),
        ("📦 Ombor", [("Mahsulotlar","Mahsulotlar ro'yxati"),("Kirim/Chiqim","Tezkor amallar"),("Tarix","Amaliyotlar tarixi")]),
        ("👷 Ishchilar", [("Ro'yxat","Ishchilar va maoshi"),("Avanslar","Avans berish"),("Jarimalar","Jarima boshqarish")]),
        ("📊 Hisobotlar", [("Kunlik/Haftalik/Oylik","Davriy hisobotlar"),("Maosh","Oylik maosh"),("Excel","Yuklab olish")]),
        ("⚙️ Sozlamalar", [("Narxlar","Ish narxlari"),("Threshold","Minimal qoldiq")]),
    ]
    html = ""
    for title, items in sections:
        items_html = ""
        for name, desc in items:
            items_html += f'<div style="padding:10px 14px;border-left:3px solid #6366f1;margin-bottom:8px;background:rgba(99,102,241,.05);border-radius:0 8px 8px 0"><div style="font-weight:700;color:#f1f5f9;font-size:13px">{name}</div><div style="color:#94a3b8;font-size:12px;margin-top:2px">{desc}</div></div>'
        html += f'<div class="card" style="margin-bottom:16px"><h3>{title}</h3>{items_html}</div>'

    content = f'<h1>❓ Yordam</h1>{html}<div class="card" style="background:linear-gradient(135deg,rgba(99,102,241,.15),rgba(139,92,246,.1))"><h3 style="color:#4F46E5">💡 Maslahat</h3><p style="color:#535C78;font-size:13px">Web panel va bot avtomatik sinxronlashadi.</p></div>'
    return web.Response(text=_base("Yordam", "help", content), content_type="text/html")





@_require_auth
async def notifications_page(request: web.Request):
    """Bildirishnomalar sahifasi."""
    async with AsyncSessionLocal() as db:
        from sqlalchemy import func as sf

        # Past qoldiq mahsulotlar
        # Past qoldiq — faqat xabar yoqilgan mahsulotlar
        r = await db.execute(
            select(WarehouseProduct).where(
                WarehouseProduct.miqdor <= WarehouseProduct.min_threshold,
                WarehouseProduct.is_active == True,
                WarehouseProduct.alert_enabled == True,
            ).order_by(WarehouseProduct.miqdor).limit(50)
        )
        low_stock = r.scalars().all()

        # Sariq zona — faqat xabar yoqilgan mahsulotlar
        r = await db.execute(
            select(WarehouseProduct).where(
                WarehouseProduct.miqdor > WarehouseProduct.min_threshold,
                WarehouseProduct.miqdor <= WarehouseProduct.yellow_threshold,
                WarehouseProduct.is_active == True,
                WarehouseProduct.alert_enabled == True,
            ).order_by(WarehouseProduct.miqdor).limit(30)
        )
        yellow = r.scalars().all()

        # Xabar o'chirilgan mahsulotlar soni
        r_off = await db.execute(
            select(func.count(WarehouseProduct.id)).where(
                WarehouseProduct.is_active == True,
                WarehouseProduct.alert_enabled == False,
            )
        )
        alerts_off_count = r_off.scalar() or 0

        # So'nggi rad etilgan ishlar
        r = await db.execute(
            select(WorkEntry, User.full_name)
            .join(User, User.id == WorkEntry.worker_id)
            .where(WorkEntry.status == WorkStatus.rejected)
            .order_by(WorkEntry.created_at.desc()).limit(20)
        )
        rejected = r.all()

    # Past qoldiq
    low_html = ""
    for p in low_stock:
        low_html += f'''
        <div style="display:flex;justify-content:space-between;align-items:center;padding:10px 14px;
             background:rgba(239,68,68,.08);border-left:3px solid #ef4444;border-radius:0 8px 8px 0;margin-bottom:6px">
          <div>
            <div style="font-weight:700;color:#f1f5f9;font-size:13px">🔴 {h(p.name)}</div>
            <div style="color:#94a3b8;font-size:11px">Kategoriya: {p.category.value if hasattr(p.category,"value") else p.category}</div>
          </div>
          <div style="text-align:right">
            <div style="color:#f87171;font-weight:700;font-size:14px">{p.miqdor:.0f} {p.birlik}</div>
            <div style="color:#94a3b8;font-size:10px">Min: {p.min_threshold:.0f}</div>
          </div>
        </div>'''

    # Sariq zona
    yellow_html = ""
    for p in yellow:
        yellow_html += f'''
        <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 14px;
             background:rgba(245,158,11,.08);border-left:3px solid #f59e0b;border-radius:0 8px 8px 0;margin-bottom:6px">
          <div style="font-weight:600;color:#f1f5f9;font-size:12px">🟡 {h(p.name)}</div>
          <div style="color:#fbbf24;font-weight:600;font-size:12px">{p.miqdor:.0f} {p.birlik}</div>
        </div>'''

    # Rad etilgan
    rej_html = ""
    for entry, wname in rejected:
        rej_html += f'''
        <div style="padding:10px 14px;background:rgba(99,102,241,.05);border-radius:8px;margin-bottom:6px">
          <div style="display:flex;justify-content:space-between;font-size:12px">
            <span style="color:#f1f5f9;font-weight:600">{h(wname)}</span>
            <span style="color:#94a3b8;font-size:10px">{entry.created_at.strftime("%d.%m %H:%M")}</span>
          </div>
          <div style="color:#94a3b8;font-size:11px;margin-top:4px">
            Ish: {entry.work_type.value if hasattr(entry.work_type,"value") else entry.work_type} · Soni: {entry.soni}
          </div>
        </div>'''

    content = f'''
    <h1>🔔 Bildirishnomalar</h1>

    <div class="grid-3" style="display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:20px 0">
      <div class="stat-card" style="background:rgba(239,68,68,.08);border-left:4px solid #ef4444">
        <div class="stat-num" style="color:#f87171">{len(low_stock)}</div>
        <div class="stat-label">🔔 Past qoldiq</div>
      </div>
      <div class="stat-card" style="background:rgba(245,158,11,.08);border-left:4px solid #f59e0b">
        <div class="stat-num" style="color:#fbbf24">{len(yellow)}</div>
        <div class="stat-label">🔔 Sariq zona</div>
      </div>
      <div class="stat-card" style="background:rgba(99,102,241,.08);border-left:4px solid #6366f1">
        <div class="stat-num" style="color:#4F46E5">{len(rejected)}</div>
        <div class="stat-label">Rad etilgan ishlar</div>
      </div>
      <div class="stat-card" style="background:rgba(100,116,139,.08);border-left:4px solid #64748b">
        <div class="stat-num" style="color:#94a3b8">{alerts_off_count}</div>
        <div class="stat-label">🔕 Xabar o&#39;chirilgan</div>
      </div>
    </div>

    <div class="grid-2" style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
      <div class="card">
        <h3 style="color:#f87171">🔴 Past qoldiq mahsulotlar</h3>
        {low_html if low_html else '<div class="empty-state"><div class="empty-state-icon">✅</div><div class="empty-state-text">Hammasi yetarli</div></div>'}
      </div>
      <div class="card">
        <h3 style="color:#fbbf24">🟡 Sariq zona</h3>
        {yellow_html if yellow_html else '<div class="empty-state"><div class="empty-state-icon">✅</div><div class="empty-state-text">Yo\'q</div></div>'}
      </div>
    </div>

    <div class="card" style="margin-top:16px">
      <h3>❌ So\'nggi rad etilgan ishlar</h3>
      {rej_html if rej_html else '<div class="empty-state"><div class="empty-state-icon">✨</div><div class="empty-state-text">Hech narsa yo\'q</div></div>'}
    </div>
    '''
    return web.Response(text=_base("Bildirishnomalar", "notifications", content), content_type="text/html")


@_require_auth
async def system_info(request: web.Request):
    """Tizim haqida ma'lumot."""
    import platform, sys
    from datetime import datetime

    async with AsyncSessionLocal() as db:
        from sqlalchemy import func as sf
        # Statistika
        total_users   = (await db.execute(select(sf.count(User.id)).where(User.is_active == True))).scalar() or 0
        total_works   = (await db.execute(select(sf.count(WorkEntry.id)))).scalar() or 0
        total_products= (await db.execute(select(sf.count(WarehouseProduct.id)).where(WarehouseProduct.is_active == True))).scalar() or 0
        total_summa   = (await db.execute(select(sf.coalesce(sf.sum(WorkEntry.jami_summa), 0)).where(WorkEntry.status == WorkStatus.approved))).scalar() or 0

    content = f'''
    <h1>ℹ️ Tizim haqida</h1>

    <div class="grid-2" style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:20px">
      <div class="card">
        <h3>📊 Umumiy statistika</h3>
        <table style="width:100%;margin-top:10px">
          <tr><td style="padding:8px;color:#94a3b8">Faol ishchilar</td><td style="padding:8px;text-align:right;font-weight:700">{total_users}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Jami ishlar</td><td style="padding:8px;text-align:right;font-weight:700">{total_works:,}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Ombor mahsulotlari</td><td style="padding:8px;text-align:right;font-weight:700">{total_products:,}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Jami daromad</td><td style="padding:8px;text-align:right;font-weight:700;color:#34d399">{float(total_summa):,.0f} so\'m</td></tr>
        </table>
      </div>

      <div class="card">
        <h3>⚙️ Texnik ma\'lumot</h3>
        <table style="width:100%;margin-top:10px">
          <tr><td style="padding:8px;color:#94a3b8">Python</td><td style="padding:8px;text-align:right;font-size:12px">{sys.version.split()[0]}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Platforma</td><td style="padding:8px;text-align:right;font-size:12px">{platform.system()}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Server vaqti</td><td style="padding:8px;text-align:right;font-size:12px">{datetime.now().strftime("%d.%m.%Y %H:%M")}</td></tr>
          <tr><td style="padding:8px;color:#94a3b8">Bot status</td><td style="padding:8px;text-align:right"><span class="badge badge-success">🟢 Faol</span></td></tr>
        </table>
      </div>
    </div>

    <div class="card" style="margin-top:16px;background:linear-gradient(135deg,rgba(99,102,241,.15),rgba(139,92,246,.1))">
      <h3 style="color:#4F46E5">🎯 Quti Tsexi Bot v5.0</h3>
      <p style="color:#535C78;font-size:13px;line-height:1.7">
        Karobka ishlab chiqarish korxonasi uchun mo\'ljallangan ish boshqaruv tizimi.<br>
        Ishchilar, omborchi va admin uchun alohida funksionallik.<br>
        Real vaqtda statistika va hisobotlar.
      </p>
    </div>

    <div class="card" style="margin-top:16px">
      <h3>📝 So\'nggi yangiliklar</h3>
      <ul style="list-style:none;padding:0;margin:10px 0">
        <li style="padding:8px 0;border-bottom:1px solid rgba(99,102,241,.1);color:#535C78;font-size:13px">✅ Mobil moslashtirish qo\'shildi</li>
        <li style="padding:8px 0;border-bottom:1px solid rgba(99,102,241,.1);color:#535C78;font-size:13px">✅ Kengaytirilgan statistika sahifasi</li>
        <li style="padding:8px 0;border-bottom:1px solid rgba(99,102,241,.1);color:#535C78;font-size:13px">✅ Material oqimi sahifasi</li>
        <li style="padding:8px 0;border-bottom:1px solid rgba(99,102,241,.1);color:#535C78;font-size:13px">✅ Bildirishnomalar markazi</li>
        <li style="padding:8px 0;color:#535C78;font-size:13px">✅ Yordam sahifasi</li>
      </ul>
    </div>
    '''
    return web.Response(text=_base("Tizim", "system_info", content), content_type="text/html")





@_require_auth
async def analytics_page(request: web.Request):
    """Tahlil sahifasi — ishchilar samaradorligi."""
    async with AsyncSessionLocal() as db:
        from sqlalchemy import func as sf
        from datetime import datetime, timedelta

        today = datetime.now().date()
        oy_boshi = today.replace(day=1)
        hafta_boshi = today - timedelta(days=today.weekday())

        # Ishchilar samaradorligi
        r = await db.execute(
            select(
                User.id, User.full_name,
                sf.count(WorkEntry.id).label("oylik_soni"),
                sf.coalesce(sf.sum(WorkEntry.jami_summa), 0).label("oylik_summa"),
            )
            .outerjoin(WorkEntry, (WorkEntry.worker_id == User.id) &
                       (WorkEntry.created_at >= oy_boshi) &
                       (WorkEntry.status == WorkStatus.approved))
            .where(User.role == UserRole.ishchi, User.is_active == True)
            .group_by(User.id, User.full_name)
            .order_by(sf.sum(WorkEntry.jami_summa).desc())
        )
        workers_data = r.all()

        # O'rtacha
        if workers_data:
            avg_soni = sum(w[2] or 0 for w in workers_data) / len(workers_data)
            avg_summa = sum(float(w[3] or 0) for w in workers_data) / len(workers_data)
        else:
            avg_soni = avg_summa = 0

    # Stat karta
    total_workers = len(workers_data)
    active_workers = sum(1 for w in workers_data if (w[2] or 0) > 0)

    workers_html = ""
    for uid, name, cnt, summa in workers_data:
        cnt = cnt or 0
        summa = float(summa or 0)
        diff_summa = ((summa - avg_summa) / avg_summa * 100) if avg_summa else 0
        diff_color = "#34d399" if diff_summa >= 0 else "#f87171"
        diff_sign = "+" if diff_summa >= 0 else ""

        initials = "".join([w[0].upper() for w in (name or "?").split()[:2]]) or "?"

        workers_html += f'''
        <div class="list-item" onclick="location.href=\'/web/workers/{uid}\'">
          <div style="display:flex;align-items:center;gap:12px">
            <div class="avatar">{initials}</div>
            <div style="flex:1">
              <div style="font-weight:700;color:#f1f5f9;font-size:13px">{h(name)}</div>
              <div style="color:#94a3b8;font-size:11px;margin-top:2px">{cnt} ta ish · {summa:,.0f} so\'m</div>
            </div>
            <div style="text-align:right">
              <div style="color:{diff_color};font-weight:700;font-size:12px">{diff_sign}{diff_summa:.0f}%</div>
              <div style="color:#64748b;font-size:10px">o\'rtachadan</div>
            </div>
          </div>
        </div>'''

    content = f'''
    <h1>📈 Ishchilar tahlili</h1>

    <div class="grid-3" style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin:20px 0">
      <div class="stat-card stat-card-gradient">
        <div class="stat-num" style="color:#4F46E5">{total_workers}</div>
        <div class="stat-label">Jami ishchilar</div>
      </div>
      <div class="stat-card stat-card-gradient">
        <div class="stat-num" style="color:#34d399">{active_workers}</div>
        <div class="stat-label">Faol (oylik)</div>
      </div>
      <div class="stat-card stat-card-gradient">
        <div class="stat-num" style="color:#fbbf24">{avg_summa:,.0f}</div>
        <div class="stat-label">O\'rtacha daromad</div>
      </div>
    </div>

    <div class="card">
      <h3>👥 Ishchilar reytingi (oylik)</h3>
      <p style="color:#94a3b8;font-size:12px;margin-bottom:14px">
        O\'rtachadan farqi % ko\'rsatilgan
      </p>
      {workers_html if workers_html else '<div class="empty-state"><div class="empty-state-icon">👥</div><div>Ma\'lumot yo\'q</div></div>'}
    </div>
    '''
    return web.Response(text=_base("Tahlil", "analytics", content), content_type="text/html")


@_require_auth
async def quick_actions(request: web.Request):
    """Tezkor amallar sahifasi."""
    actions = [
        ("📦", "Tezkor kirim",       "/web/warehouse?action=kirim",   "Omborga kirim qo\'shish"),
        ("📤", "Tezkor chiqim",      "/web/warehouse?action=chiqim",  "Ombordan chiqim"),
        ("🔍", "Mahsulot qidirish",  "/web/warehouse",                "Ombor qidirish"),
        ("💳", "Avans berish",       "/web/avans",                    "Ishchiga avans"),
        ("⚠️", "Jarima berish",      "/web/penalties",                "Jarima qo\'shish"),
        ("📊", "Kunlik hisobot",     "/web/reports/download/daily",   "Excel yuklab olish"),
        ("📈", "Statistika",          "/web/stats-advanced",           "To\'liq tahlil"),
        ("🔔", "Bildirishnomalar",   "/web/notifications",            "Past qoldiq va boshqalar"),
        ("⚙️", "Narxlar",             "/web/prices",                   "Ish narxlarini sozlash"),
        ("💰", "Maoshlar",            "/web/salary",                   "Maosh hisobotlari"),
        ("🏭", "Ombor bo\'limlari",  "/web/ombor",                    "Kategoriya bo\'yicha"),
        ("ℹ️", "Tizim ma\'lumoti",  "/web/system",                   "Tizim haqida"),
    ]

    actions_html = '<div class="quick-grid">'
    for icon, name, url, desc in actions:
        actions_html += f'''
        <a href="{url}" class="quick-action">
          <div class="quick-action-icon">{icon}</div>
          <div class="quick-action-label">{name}</div>
          <div style="color:#64748b;font-size:10px;margin-top:4px">{desc}</div>
        </a>'''
    actions_html += '</div>'

    content = f'''
    <h1>⚡ Tezkor amallar</h1>
    <p style="color:#94a3b8;font-size:13px;margin-bottom:20px">
      Eng ko\'p ishlatadigan amallar shu yerda
    </p>
    {actions_html}

    <div class="card" style="margin-top:20px">
      <h3>💡 Klaviatura yorliqlari</h3>
      <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-top:12px">
        <div style="display:flex;justify-content:space-between;padding:8px 12px;background:rgba(99,102,241,.08);border-radius:6px">
          <span style="font-size:12px">Sahifani yangilash</span>
          <span class="code">F5</span>
        </div>
        <div style="display:flex;justify-content:space-between;padding:8px 12px;background:rgba(99,102,241,.08);border-radius:6px">
          <span style="font-size:12px">Qidirish</span>
          <span class="code">Ctrl+F</span>
        </div>
        <div style="display:flex;justify-content:space-between;padding:8px 12px;background:rgba(99,102,241,.08);border-radius:6px">
          <span style="font-size:12px">Bosh sahifa</span>
          <span class="code">Home</span>
        </div>
        <div style="display:flex;justify-content:space-between;padding:8px 12px;background:rgba(99,102,241,.08);border-radius:6px">
          <span style="font-size:12px">Pastga</span>
          <span class="code">End</span>
        </div>
      </div>
    </div>
    '''
    return web.Response(text=_base("Tezkor amallar", "quick", content), content_type="text/html")



def create_app() -> web.Application:
    app = web.Application()

    app.router.add_get("/web/warehouse/csv-import",  warehouse_csv_import)
    app.router.add_post("/web/warehouse/csv-import", warehouse_csv_import)
    app.router.add_get("/web/ombor",                 ombor_cats)
    app.router.add_get("/web/tayyor-chiqim",          tayyor_chiqim)
    app.router.add_get("/web/ombor/{cat_key}",       ombor_cat_detail)
    app.router.add_post("/web/ombor/{cat_key}/add",  ombor_cat_add)
    app.router.add_post("/web/ombor/quick-transfer",  ombor_admin_quick_transfer)
    app.router.add_get("/web/login",         web_login)
    app.router.add_get("/w/{token}", token_login)
    # Omborchi va inspektor panellari — alohida modullar
    from web_panel_ombor import register_ombor_routes
    from web_panel_inspektor import register_inspektor_routes
    from web_panel_topshiriq import register_topshiriq_routes
    from web_panel_rulon import register_rulon_routes
    register_ombor_routes(app)
    register_inspektor_routes(app)
    register_topshiriq_routes(app)
    register_rulon_routes(app)
    app.router.add_post("/web/login",         web_login)
    app.router.add_get("/web/logout",         web_logout)
    app.router.add_get("/web/",                       dashboard)
    app.router.add_get("/web/production", production_analytics)
    app.router.add_get("/web/health", inventory_health)
    app.router.add_get("/web/activity", activity_feed)
    app.router.add_get("/web/quality", quality_dashboard)
    app.router.add_get("/web/salary-projection", salary_projection)
    app.router.add_get("/web/system-health", web_system_health)
    app.router.add_get("/web/stats-advanced",         stats_advanced)
    app.router.add_get("/web/material-flow",          material_flow)
    app.router.add_get("/web/help",                   help_page)
    app.router.add_get("/web/notifications",          notifications_page)
    app.router.add_get("/web/system",                 system_info)
    app.router.add_get("/web/analytics",              analytics_page)
    app.router.add_get("/web/quick",                  quick_actions)

    # Warehouse
    app.router.add_get("/web/warehouse",              warehouse)
    app.router.add_post("/web/warehouse/add",         warehouse_add)
    app.router.add_post("/web/warehouse/kirim",       warehouse_kirim)
    app.router.add_post("/web/warehouse/chiqim",      warehouse_chiqim)
    app.router.add_post("/web/warehouse/thresholds",  warehouse_thresholds)
    app.router.add_post("/web/warehouse/alert-toggle", warehouse_alert_toggle)
    app.router.add_get("/web/warehouse/export",       warehouse_export)
    app.router.add_get("/web/warehouse/logs",         warehouse_logs)
    app.router.add_post("/web/warehouse/delete",      warehouse_delete)
    app.router.add_get("/web/warehouse/delete-confirm/{id}", warehouse_delete_confirm)
    app.router.add_get("/web/zero-stock",             zero_stock_page)
    app.router.add_post("/web/warehouse/zero-keep",   warehouse_zero_keep)

    # Workers
    app.router.add_get("/web/workers",                workers)
    app.router.add_get("/web/workers/{worker_id}",    worker_detail)
    app.router.add_post("/web/workers/{id}/toggle",   worker_toggle)
    app.router.add_post("/web/workers/{id}/delete",   worker_delete)

    # Finance
    app.router.add_get("/web/avans",                  avans)
    app.router.add_post("/web/avans/add",             avans_add)
    app.router.add_get("/web/penalties",              penalties)
    app.router.add_post("/web/penalties/add",         penalties_add)

    # Reports
    app.router.add_get("/web/reports",                reports)
    app.router.add_get("/web/ombor-report",           ombor_report)
    app.router.add_get("/web/maosh-report",           maosh_report)
    app.router.add_get("/web/reports/download/{t}",   reports_download)

    # Salary
    app.router.add_get("/web/salary",                 salary)
    app.router.add_post("/web/salary/confirm/{id}",   salary_confirm)
    app.router.add_post("/web/salary/confirm-all",    salary_confirm_all)

    # Prices
    app.router.add_get("/web/prices",                 prices)
    app.router.add_post("/web/prices/set",            prices_set)
    app.router.add_post("/web/prices/save",           prices_save)

    # Root redirect
    app.router.add_get("/", lambda r: web.HTTPFound("/web/"))

    return app

async def start_web(app: web.Application):
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, WEB_HOST, WEB_PORT)
    await site.start()
    logger.info("Web panel: http://%s:%s/web/", WEB_HOST, WEB_PORT)
    return runner

