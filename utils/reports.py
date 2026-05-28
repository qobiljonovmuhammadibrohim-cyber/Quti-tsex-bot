"""utils/reports.py — Excel hisobotlar."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, extract
from datetime import date, timedelta

from database.models import User, UserRole, WorkEntry, WorkStatus, WorkType, WorkPrice, Penalty, SalaryReport
from database.queries import get_users_by_role, get_monthly_reports, get_penalty_sum, get_advance_sum

THIN = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
MONTHS_UZ = ["","Yanvar","Fevral","Mart","Aprel","May","Iyun","Iyul","Avgust","Sentabr","Oktabr","Noyabr","Dekabr"]

def _border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row,max_row=max_row,min_col=min_col,max_col=max_col):
        for cell in row: cell.border = THIN

def _header_cell(cell, text, bg="2E86AB"):
    cell.value = text; cell.font = Font(color="FFFFFF",bold=True,size=11)
    cell.fill  = PatternFill(start_color=bg,end_color=bg,fill_type="solid")
    cell.alignment = Alignment(horizontal='center',vertical='center')

async def generate_daily_excel(db: AsyncSession, report_date: date = None) -> io.BytesIO:
    if not report_date: report_date = date.today()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Kunlik hisobot"
    ws.merge_cells("A1:I1"); ws["A1"].value = f"KUNLIK HISOBOT — {report_date.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True,size=14); ws["A1"].alignment = Alignment(horizontal='center')
    headers = ["#","Ishchi","Ish turi","Mahsulot","Razmer","Soni","Narx","Summa","Holat"]
    for col, h in enumerate(headers,1): _header_cell(ws.cell(row=3,column=col), h)
    r = await db.execute(
        select(WorkEntry, User).join(User, WorkEntry.worker_id==User.id)
        .where(WorkEntry.work_date==report_date).order_by(User.full_name, WorkEntry.created_at))
    entries = r.all()
    STATUS_NAMES = {WorkStatus.pending:"Kutmoqda",WorkStatus.approved:"Tasdiqlangan",
                    WorkStatus.adjusted:"Tuzatilgan",WorkStatus.rejected:"Rad etilgan"}
    GREEN = PatternFill(start_color="C8E6C9",end_color="C8E6C9",fill_type="solid")
    RED   = PatternFill(start_color="FFCDD2",end_color="FFCDD2",fill_type="solid")
    row = 4; total = 0.0
    for i,(entry,worker) in enumerate(entries,1):
        vals = [i,worker.full_name,entry.work_type.value.replace("_"," ").title(),
                entry.mahsulot_nomi or "",entry.razmer or "",entry.soni,
                entry.birlik_narx or 0,entry.jami_summa or 0,STATUS_NAMES.get(entry.status,"")]
        fill = None
        if entry.status in (WorkStatus.approved,WorkStatus.adjusted): fill=GREEN; total+=entry.jami_summa or 0
        elif entry.status == WorkStatus.rejected: fill=RED
        for col,val in enumerate(vals,1):
            cell = ws.cell(row=row,column=col,value=val)
            if fill: cell.fill = fill
        row += 1
    ws.cell(row=row+1,column=7,value="JAMI TASDIQLANGAN:").font = Font(bold=True)
    ws.cell(row=row+1,column=8,value=total).font = Font(bold=True)
    for i,w_ in enumerate([5,25,22,25,12,8,12,15,18],1): ws.column_dimensions[get_column_letter(i)].width = w_
    if row > 3: _border(ws,3,row-1,1,9)
    out = io.BytesIO(); wb.save(out); out.seek(0); return out


async def generate_weekly_excel(db: AsyncSession) -> io.BytesIO:
    today = date.today(); week_start = today - timedelta(days=today.weekday())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Haftalik hisobot"
    ws.merge_cells("A1:G1"); ws["A1"].value = f"HAFTALIK HISOBOT — {week_start.strftime('%d.%m')} – {today.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True,size=14); ws["A1"].alignment = Alignment(horizontal='center')
    headers = ["#","Ishchi","Jami ishlar","Tasdiqlangan","Rad etilgan","Summa","Jarimalar"]
    for col,h in enumerate(headers,1): _header_cell(ws.cell(row=3,column=col),h,bg="E65100")
    workers = await get_users_by_role(db,UserRole.ishchi)
    row = 4
    for i,worker in enumerate(workers,1):
        apr = await db.execute(
            select(func.count(WorkEntry.id),func.coalesce(func.sum(WorkEntry.jami_summa),0))
            .where(WorkEntry.worker_id==worker.id,WorkEntry.work_date>=week_start,
                   WorkEntry.work_date<=today,WorkEntry.status.in_([WorkStatus.approved,WorkStatus.adjusted])))
        apr_row = apr.one()
        rej = (await db.execute(select(func.count(WorkEntry.id)).where(WorkEntry.worker_id==worker.id,
               WorkEntry.work_date>=week_start,WorkEntry.status==WorkStatus.rejected))).scalar() or 0
        total_count = (await db.execute(select(func.count(WorkEntry.id)).where(
               WorkEntry.worker_id==worker.id,WorkEntry.work_date>=week_start))).scalar() or 0
        pen_r = await db.execute(select(func.coalesce(func.sum(Penalty.summa),0))
            .where(Penalty.worker_id==worker.id,Penalty.created_at>=week_start,Penalty.created_at<=today))
        pen_sum = float(pen_r.scalar())
        ws.cell(row=row,column=1,value=i); ws.cell(row=row,column=2,value=worker.full_name)
        ws.cell(row=row,column=3,value=total_count); ws.cell(row=row,column=4,value=apr_row[0] or 0)
        ws.cell(row=row,column=5,value=rej); ws.cell(row=row,column=6,value=float(apr_row[1]))
        ws.cell(row=row,column=7,value=pen_sum); row += 1
    for i,w_ in enumerate([5,25,15,15,15,18,15],1): ws.column_dimensions[get_column_letter(i)].width = w_
    if row > 3: _border(ws,3,row-1,1,7)
    out = io.BytesIO(); wb.save(out); out.seek(0); return out


async def generate_monthly_excel(db: AsyncSession, oy: int, yil: int) -> io.BytesIO:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Oylik maosh"
    ws.merge_cells("A1:H1"); ws["A1"].value = f"OYLIK MAOSH — {MONTHS_UZ[oy]} {yil}"
    ws["A1"].font = Font(bold=True,size=14); ws["A1"].alignment = Alignment(horizontal='center')
    headers = ["#","Ishchi","Ish summasi","Jarimalar","Avanslar","Sof maosh","Holat","Tasdiqlangan"]
    for col,h in enumerate(headers,1): _header_cell(ws.cell(row=3,column=col),h,bg="1B5E20")
    reports = await get_monthly_reports(db,oy,yil)
    row = 4; grand = 0.0
    for i,rep in enumerate(reports,1):
        from database.queries import get_user_by_id
        w = await get_user_by_id(db, rep.worker_id)
        ws.cell(row=row,column=1,value=i); ws.cell(row=row,column=2,value=w.full_name if w else "?")
        ws.cell(row=row,column=3,value=rep.jami_ish_summa); ws.cell(row=row,column=4,value=rep.jami_jarima)
        ws.cell(row=row,column=5,value=rep.jami_avans); ws.cell(row=row,column=6,value=rep.sof_maosh)
        ws.cell(row=row,column=7,value="Tasdiqlangan" if rep.admin_tasdiqladi else "Kutmoqda")
        ws.cell(row=row,column=8,value=rep.tasdiq_vaqti.strftime("%d.%m.%Y %H:%M") if rep.tasdiq_vaqti else "")
        grand += rep.sof_maosh; row += 1
    ws.cell(row=row+1,column=5,value="JAMI:").font = Font(bold=True)
    ws.cell(row=row+1,column=6,value=grand).font = Font(bold=True)
    for i,w_ in enumerate([5,25,18,15,15,18,18,22],1): ws.column_dimensions[get_column_letter(i)].width = w_
    if row > 3: _border(ws,3,row-1,1,8)
    out = io.BytesIO(); wb.save(out); out.seek(0); return out
